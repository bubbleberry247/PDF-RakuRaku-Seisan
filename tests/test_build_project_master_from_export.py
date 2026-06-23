from __future__ import annotations

import importlib.util
from pathlib import Path

import openpyxl
import pytest


ROOT_DIR = Path(__file__).resolve().parents[1]
MODULE_PATH = ROOT_DIR / "tools" / "build_project_master_from_export.py"


def load_module():
    spec = importlib.util.spec_from_file_location(
        "build_project_master_from_export", MODULE_PATH
    )
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def write_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def read_workbook(path: Path) -> list[list[object]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def test_template_master_preserves_headers_and_extra_fields(tmp_path: Path) -> None:
    module = load_module()
    source_path = tmp_path / "source.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "project_master.xlsx"

    write_workbook(
        source_path,
        [
            ["工事番号", "工事名称", "部署", "顧客"],
            ["1001", "A棟修繕工事", "修繕", "東海A"],
            ["1002", "B棟新築計画", "新築", "東海B"],
        ],
    )
    write_workbook(
        template_path,
        [
            ["工事番号", "工事名称", "キーワード", "部署", "status", "source", "顧客", "備考"],
            ["1001", "旧A", "旧キーワード", "旧部署", "inactive", "legacy", "旧顧客", "残す"],
        ],
    )

    exit_code = module.main(
        [
            "--input",
            str(source_path),
            "--output",
            str(output_path),
            "--template-master",
            str(template_path),
        ]
    )

    rows = read_workbook(output_path)
    assert exit_code == 0
    assert rows[0] == ["工事番号", "工事名称", "キーワード", "部署", "status", "source", "顧客", "備考"]
    assert rows[1][0:5] == ["1001", "A棟修繕工事", "A棟修繕工事,東海A", "修繕", "active"]
    assert rows[1][5:] == ["legacy", "旧顧客", "残す"]
    assert rows[2][0:5] == ["1002", "B棟新築計画", "B棟新築計画,東海B", "新築", "active"]
    assert rows[2][5:] == ["export:source.xlsx", "東海B", None]


def test_template_master_requires_normalized_columns(tmp_path: Path) -> None:
    module = load_module()
    source_path = tmp_path / "source.xlsx"
    template_path = tmp_path / "template_missing.xlsx"
    output_path = tmp_path / "project_master.xlsx"

    write_workbook(
        source_path,
        [
            ["工事番号", "工事名称", "部署"],
            ["1001", "A棟修繕工事", "修繕"],
        ],
    )
    write_workbook(
        template_path,
        [
            ["工事番号", "工事名称", "キーワード", "部署"],
            ["1001", "旧A", "旧キーワード", "旧部署"],
        ],
    )

    with pytest.raises(ValueError, match="template master missing required columns"):
        module.main(
            [
                "--input",
                str(source_path),
                "--output",
                str(output_path),
                "--template-master",
                str(template_path),
            ]
        )
