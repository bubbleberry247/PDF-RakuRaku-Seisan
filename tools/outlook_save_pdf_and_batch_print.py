# -*- coding: utf-8 -*-
"""
Scenario 12/13: 受信メールのPDFを保存・一括印刷（Outlook COM）

目的:
- Outlook受信メールからPDF添付（請求書）を漏れなく保存する
- 必要ならPDFを結合し、一括印刷する
- 例外（暗号化PDFのパスワード不明、URLのみで未対応など）は処理を止めて通知する（TPS: 自働化/アンドン）

安全:
- 既定はドライラン（--execute を付けない限り保存/印刷しない）
- 上書き禁止（ファイル名衝突は自動で退避名にする）

Usage examples:
    python outlook_save_pdf_and_batch_print.py --config C:\\...\\tool_config.json --dry-run --dry-run-mail
    python outlook_save_pdf_and_batch_print.py --config C:\\...\\tool_config.json --execute
    python outlook_save_pdf_and_batch_print.py --config C:\\...\\tool_config.json --execute --print
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import importlib.util
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import traceback
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence, TypeVar
from unicodedata import normalize

sys.path.insert(0, str(Path(__file__).parent))

from common.email_notifier import (
    OutlookEmail,
    build_simple_html,
    html_link_to_path,
    send_outlook,
)

try:
    import win32com.client as win32  # type: ignore

    _OUTLOOK_AVAILABLE = True
except Exception:
    _OUTLOOK_AVAILABLE = False

try:
    import pythoncom  # type: ignore
    import pywintypes  # type: ignore

    _PYWIN32_COM_AVAILABLE = True
except Exception:
    _PYWIN32_COM_AVAILABLE = False

try:
    from web_invoice_downloader import (
        WebDownloadConfig, VendorWebConfig, WebDownloadResult,
        dispatch_web_download, cleanup_web_sessions,
    )
    _WEB_DOWNLOAD_AVAILABLE = True
except Exception:
    _WEB_DOWNLOAD_AVAILABLE = False


T = TypeVar("T")


_COM_MESSAGE_FILTER_INSTALLED = False


class _RetryMessageFilter:
    """
    COM message filter to automatically retry when Outlook is busy.

    This prevents sporadic failures like:
      (-2147418111, '呼び出し先が呼び出しを拒否しました。', None, None)
    """

    def HandleInComingCall(  # noqa: N802 - required COM interface name
        self, htaskCaller: Any, dwTickCount: int, dwCallType: int
    ) -> int:
        return pythoncom.SERVERCALL_ISHANDLED  # type: ignore[name-defined]

    def RetryRejectedCall(  # noqa: N802 - required COM interface name
        self, htaskCaller: Any, dwTickCount: int, dwRejectType: int
    ) -> int:
        # Retry later: return a delay in milliseconds.
        if dwRejectType == pythoncom.SERVERCALL_RETRYLATER:  # type: ignore[name-defined]
            return 250
        return -1

    def MessagePending(  # noqa: N802 - required COM interface name
        self, htaskCaller: Any, dwTickCount: int, dwPendingType: int
    ) -> int:
        return pythoncom.PENDINGMSG_WAITDEFPROCESS  # type: ignore[name-defined]


def _install_com_message_filter() -> None:
    global _COM_MESSAGE_FILTER_INSTALLED
    if _COM_MESSAGE_FILTER_INSTALLED:
        return
    if not _PYWIN32_COM_AVAILABLE:
        return
    try:
        # Make sure COM is initialized for this thread.
        pythoncom.CoInitialize()  # type: ignore[name-defined]
    except Exception:
        pass
    try:
        pythoncom.CoRegisterMessageFilter(_RetryMessageFilter())  # type: ignore[name-defined]
        _COM_MESSAGE_FILTER_INSTALLED = True
    except Exception:
        # Best-effort: tool can still work, but may hit RPC_E_CALL_REJECTED occasionally.
        pass


def _is_call_rejected_error(err: BaseException) -> bool:
    if not _PYWIN32_COM_AVAILABLE:
        return False
    if not isinstance(err, pywintypes.com_error):  # type: ignore[name-defined]
        return False
    if not err.args:
        return False
    hresult = err.args[0]
    # RPC_E_CALL_REJECTED / RPC_E_SERVERCALL_RETRYLATER
    return hresult in (-2147418111, -2147417846)


def _com_retry(fn: Callable[[], T], *, retries: int = 30) -> T:
    """
    Retry a COM call when Outlook is busy.

    The message filter handles many cases, but some calls still surface
    RPC_E_CALL_REJECTED. This keeps the automation stable.
    """

    delay = 0.2
    for _ in range(retries):
        try:
            return fn()
        except Exception as e:
            if _is_call_rejected_error(e):
                time.sleep(delay)
                delay = min(2.0, delay * 1.5)
                continue
            raise
    return fn()

try:
    from pypdf import PdfReader, PdfWriter  # type: ignore

    _PYPDF_AVAILABLE = True
except Exception:
    _PYPDF_AVAILABLE = False

try:
    import win32api  # type: ignore
    import win32print  # type: ignore

    _WIN32_PRINT_AVAILABLE = True
except Exception:
    _WIN32_PRINT_AVAILABLE = False

try:
    import fitz  # type: ignore  # PyMuPDF

    _PYMUPDF_AVAILABLE = True
except Exception:
    _PYMUPDF_AVAILABLE = False

try:
    import cv2 as _cv2  # type: ignore
    import numpy as _np  # type: ignore

    _CV2_AVAILABLE = True
except ImportError:
    _CV2_AVAILABLE = False

# easyocr is optional and loaded lazily (heavy).
_EASYOCR_AVAILABLE = importlib.util.find_spec("easyocr") is not None

# Do not import yomitoku at module import time (it can be heavy / slow).
_YOMITOKU_AVAILABLE = importlib.util.find_spec("yomitoku") is not None

try:
    import pyzipper  # type: ignore

    _PYZIPPER_AVAILABLE = True
except Exception:
    _PYZIPPER_AVAILABLE = False


def _find_tesseract_exe() -> Path | None:
    env = os.environ.get("TESSERACT_EXE")
    if env:
        p = Path(env)
        if p.exists():
            return p

    candidates = [
        Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
        Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
    ]
    for c in candidates:
        if c.exists():
            return c

    which = shutil.which("tesseract")
    return Path(which) if which else None


_TESSERACT_EXE = _find_tesseract_exe()
_TESSERACT_AVAILABLE = _TESSERACT_EXE is not None


DEFAULT_ARTIFACT_DIR = (
    r"C:\ProgramData\RK10\Robots\12・13受信メールのPDFを保存・一括印刷\artifacts"
)

URL_RE = re.compile(r"https?://[^\s<>\"]+")
PASSWORD_RE = re.compile(r"(?:パスワード|Password)\s*[:：]?\s*([^\s]+)", re.IGNORECASE)

INVALID_FILENAME_CHARS_RE = re.compile(r'[<>:"/\\\\|?*]+')
VENDOR_HINT_RE = re.compile(
    r"(株式会社|有限会社|合同会社|一般社団法人|一般財団法人|税理士法人|社会保険労務士法人|（株）|\(株\)|（有）|\(有\))"
)
DATE_YMD_RE = re.compile(
    r"(?P<y>20\d{2})[./\\年-](?P<m>\d{1,2})[./\\月-](?P<d>\d{1,2})"
)
DATE_LABEL_RE = re.compile(
    r"(請求(?:年月)?日|発行(?:年月)?日|作成(?:年月)?日|日付)\s*[:：]?\s*([0-9]{4}[^0-9]{0,2}[0-9]{1,2}[^0-9]{0,2}[0-9]{1,2})"
)
AMOUNT_LABEL_RE = re.compile(
    r"(支払決定金額|支払決定額|お支払い金額|支払金額|ご?請求金額|ご?請求額|請求金額|請求額|総合計|合計金額|合計)"
    r"(?:\s*[（(][^)）]*[)）])?"
    r"\s*[:：]?\s*[¥￥]?\s*([0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)"
)
INVOICE_NO_RE = re.compile(
    r"(請求書番号|請求書No\.?|請求No\.?|Invoice\s*No\.?|INV\.?)[^0-9A-Za-z]*([0-9A-Za-z\-_/]{3,})"
)

# Vendor name corrections for common OCR misrecognitions (pdf-ocr Skill knowledge).
# Key = OCR misread pattern (substring), Value = corrected name.
VENDOR_OCR_CORRECTIONS: dict[str, str] = {
    "緑エキスパート": "縁エキスパート",
    # Add more corrections as discovered during pilot testing
}

ZIP_MAX_MEMBERS = 80
ZIP_MAX_TOTAL_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
ZIP_MAX_MEMBER_UNCOMPRESSED_BYTES = 80 * 1024 * 1024
ZIP_PASSWORD_MAX_CANDIDATES = 20
PROCESSED_MANIFEST_NAME = "processed_attachments_manifest.jsonl"
PROJECT_NOISE_RE = re.compile(
    r"(請求書|御請求書|納品書|見積書|契約書類|契約書|電子決済サービス|査定表|チェックリスト|回収チェックリスト|精算分|本体工事精算分|控え)"
)
PROJECT_POSITIVE_MARKERS: tuple[str, ...] = (
    "工事",
    "新築",
    "改修",
    "修繕",
    "計画",
    "様邸",
    "共同住宅",
    "アパート",
    "マンション",
    "店舗",
    "ビル",
    "センター",
    "現場",
    "町",
    "丁目",
)


@dataclass(frozen=True)
class OutlookSelection:
    folder_path: str
    # Optional Outlook profile name (e.g. "RKM"). Useful when the default profile is not the target mailbox.
    profile_name: str | None = None
    unread_only: bool = True
    max_messages: int = 200
    received_within_days: int | None = 14
    subject_allow_regex: str | None = None
    subject_deny_regex: str | None = None
    mark_as_read_on_success: bool = False
    # If true, filter by FlagStatus (0 = no flag = unprocessed).
    use_flag_status: bool = False
    # If true, set FlagStatus = 2 (red flag) on successful processing.
    mark_flag_on_success: bool = False


@dataclass(frozen=True)
class MergeConfig:
    enabled: bool = True
    output_name: str = "merged.pdf"


@dataclass(frozen=True)
class PrintConfig:
    enabled: bool = False
    printer_name: str | None = None
    # shell: send to printer, pdf_copy: copy PDFs to a folder as "print output"
    method: str = "shell"
    pdf_output_dir: str | None = None


@dataclass(frozen=True)
class MailConfig:
    send_success: bool = True
    success_to: Sequence[str] = tuple()
    error_to: Sequence[str] = tuple()
    error_cc: Sequence[str] = tuple()


@dataclass(frozen=True)
class RenameConfig:
    enabled: bool = True
    # If true, create a sub-directory per vendor under save_dir.
    create_vendor_subdir: bool = True
    vendor_subdir_template: str = "{vendor_short}"
    file_name_template: str = "{vendor}_{issue_date}_{amount_comma}_{project}.pdf"
    # Regex to exclude recipient company names from vendor extraction (optional).
    company_deny_regex: str | None = None
    max_pages: int = 2
    # If true, try to extract invoice fields from filename before PDF extraction.
    filename_first: bool = True


@dataclass(frozen=True)
class RoutingRule:
    subdir: str
    keywords: Sequence[str] = tuple()


@dataclass(frozen=True)
class ProjectMasterEntry:
    kojiban:    str          # 工事番号（例: 2025-001。空欄可）
    kojimei:    str          # 工事名称（検索キーワード）
    busho:      str          # 部署（"新築" or "修繕"）
    keywords:   tuple[str, ...] = ()  # 追加キーワード（カンマ区切りで読み込み）
    start_date: str = ""
    end_date:   str = ""     # 空=有効
    status:     str = "active"


@dataclass(frozen=True)
class SenderRule:
    sender: str
    subdir: str
    keywords: tuple = ()


@dataclass(frozen=True)
class RoutingConfig:
    enabled: bool = False
    fallback_subdir: str = ""
    rules: Sequence[RoutingRule] = tuple()
    sender_rules: Sequence[SenderRule] = tuple()
    ban_senders: frozenset = frozenset()
    threshold: int = 1
    tie_margin: float | None = None
    vision_ocr_enabled: bool = False
    vision_ocr_provider: str = "openai"


@dataclass(frozen=True)
class ToolConfig:
    artifact_dir: str = DEFAULT_ARTIFACT_DIR
    save_dir: str | None = None
    outlook: OutlookSelection | None = None
    merge: MergeConfig = MergeConfig()
    print: PrintConfig = PrintConfig()
    mail: MailConfig = MailConfig()
    rename: RenameConfig = RenameConfig()
    routing: RoutingConfig = RoutingConfig()
    fail_on_url_only_mail: bool = True
    enable_zip_processing: bool = True
    enable_pdf_decryption: bool = True
    project_master: tuple = field(default_factory=tuple)
    project_master_path: str | None = None
    enable_web_download: bool = False
    web_download: Any = None  # WebDownloadConfig | None


@dataclass(frozen=True)
class PartialInvoiceFields:
    vendor: str | None = None
    issue_date: str | None = None
    amount: int | None = None
    project: str | None = None


@dataclass(frozen=True)
class PasswordNote:
    received_time: datetime
    subject: str
    password: str
    sender: str


@dataclass(frozen=True)
class SavedAttachment:
    message_entry_id: str
    message_subject: str
    sender: str
    received_time: str
    attachment_name: str
    # Final saved path (after rename/move). Prefer decrypted output when applicable.
    saved_path: str
    original_saved_path: str
    vendor: str | None
    issue_date: str | None
    amount: int | None
    invoice_no: str | None
    project: str | None
    sha256: str
    was_encrypted: bool
    decrypted_path: str | None
    encrypted_original_path: str | None
    route_subdir: str | None = None
    route_reason: str | None = None
    processing_status: str = "saved"


def _mask_url(url: str) -> str:
    """Mask sensitive parts of URLs for safe reporting."""
    from urllib.parse import urlparse

    try:
        parsed = urlparse(url)
        # Keep scheme + host, mask path/query
        masked = f"{parsed.scheme}://{parsed.netloc}/***masked***"
        return masked
    except Exception:
        return "***masked_url***"


@dataclass(frozen=True)
class ExtractedPdfFields:
    vendor: str
    issue_date: str  # YYYYMMDD
    amount: int
    invoice_no: str | None
    project: str | None = None
    raw_ocr_text: str = ""
    source: str = ""
    vendor_category: str | None = None
    requires_manual: bool = False
    review_reasons: tuple[str, ...] = ()


def _partial_from_extracted_fields(
    fields: ExtractedPdfFields,
    fallback: PartialInvoiceFields,
) -> PartialInvoiceFields:
    return PartialInvoiceFields(
        vendor=fields.vendor or fallback.vendor,
        issue_date=fields.issue_date or fallback.issue_date,
        amount=fields.amount if fields.amount is not None else fallback.amount,
        project=fields.project or fallback.project,
    )


def _apply_invoice_guardrails(fields: ExtractedPdfFields) -> ExtractedPdfFields:
    return _apply_invoice_guardrails_with_sender(fields, sender=None)


def _apply_invoice_guardrails_with_sender(
    fields: ExtractedPdfFields,
    *,
    sender: str | None,
) -> ExtractedPdfFields:
    resolved_vendor = fields.vendor
    try:
        from vendor_matching import canonicalize_vendor as _canonicalize_vendor

        canonical_vendor = _canonicalize_vendor(
            fields.vendor,
            sender=sender,
            context_text=fields.raw_ocr_text,
            drop_non_vendor=False,
        )
        if canonical_vendor:
            resolved_vendor = canonical_vendor
    except Exception:
        pass

    if fields.source == "vision_ocr" and resolved_vendor == fields.vendor:
        return fields

    try:
        from vision_ocr import evaluate_auto_pass as _evaluate_auto_pass

        vendor_category, requires_manual, review_reasons = _evaluate_auto_pass(
            vendor=resolved_vendor,
            issue_date=fields.issue_date,
            amount=fields.amount,
            confidence="high",
            context_text=fields.raw_ocr_text,
        )
    except Exception as e:
        vendor_category = None
        requires_manual = True
        review_reasons = (f"guardrail_eval_failed:{type(e).__name__}",)

    return ExtractedPdfFields(
        vendor=resolved_vendor,
        issue_date=fields.issue_date,
        amount=fields.amount,
        invoice_no=fields.invoice_no,
        project=fields.project,
        raw_ocr_text=fields.raw_ocr_text,
        source=fields.source,
        vendor_category=vendor_category,
        requires_manual=requires_manual,
        review_reasons=review_reasons,
    )


@dataclass(frozen=True)
class UrlOnlyTask:
    message_entry_id: str
    subject: str
    sender: str
    received_time: str
    urls: Sequence[str]


@dataclass(frozen=True)
class RunReport:
    run_id: str
    started_at: str
    finished_at: str
    dry_run: bool
    save_dir: str | None
    outlook_folder_path: str | None
    scanned_messages: int
    saved_attachments: Sequence[SavedAttachment]
    url_only_tasks: Sequence[UrlOnlyTask]
    merged_pdf_path: str | None
    printed_paths: Sequence[str]
    project_master_path: str | None
    new_project_candidates_path: str | None
    unresolved: Sequence[str]


def _now_run_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def _recreate_dir(p: Path) -> None:
    if p.exists():
        shutil.rmtree(p, ignore_errors=True)
    p.mkdir(parents=True, exist_ok=True)


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _sanitize_filename(name: str, max_len: int = 180) -> str:
    name = name.strip().replace("\u0000", "")
    name = INVALID_FILENAME_CHARS_RE.sub("_", name)
    name = re.sub(r"\s+", " ", name).strip()
    if len(name) > max_len:
        root, ext = os.path.splitext(name)
        keep = max_len - len(ext)
        name = root[:keep] + ext
    if not name:
        return "unnamed"
    return name


def _slug_subject(subject: str, max_len: int = 60) -> str:
    s = subject.strip()
    s = re.sub(r"\s+", " ", s)
    s = INVALID_FILENAME_CHARS_RE.sub("_", s)
    s = s.replace("【", "[").replace("】", "]")
    s = s.replace("／", "/")
    s = s.replace("/", "_")
    s = s.replace("\\", "_")
    s = s.strip(" ._")
    if len(s) > max_len:
        s = s[:max_len].rstrip(" ._")
    return s if s else "no_subject"


def _normalize_text(text: str) -> str:
    # Normalize full-width digits/punctuation -> ASCII where possible.
    text = normalize("NFKC", text)
    return text.replace("\u3000", " ")


def _clean_project_name(value: str | None) -> str | None:
    if value is None:
        return None
    s = _normalize_text(value)
    s = re.sub(r"\s+", " ", s).strip(" _-")
    return s or None


def _looks_like_new_project_candidate(project: str | None) -> bool:
    candidate = _clean_project_name(project)
    if not candidate:
        return False
    normalized = _normalize_for_match(candidate)
    if len(normalized) < 4:
        return False
    stripped = PROJECT_NOISE_RE.sub(" ", candidate)
    stripped = re.sub(r"\s+", " ", stripped).strip(" _-")
    stripped_norm = _normalize_for_match(stripped)
    if len(stripped_norm) < 4:
        return False
    if any(marker in stripped for marker in PROJECT_POSITIVE_MARKERS):
        return True
    return len(stripped_norm) >= 8


def _cleanup_rendered_filename(value: str) -> str:
    root, ext = os.path.splitext(value)
    root = re.sub(r"_+", "_", root).strip(" _.")
    if not root:
        root = "unnamed"
    if not ext:
        ext = ".pdf"
    return root + ext


def _normalize_for_match(value: str) -> str:
    s = _normalize_text(value or "").lower()
    # Keep letters/numbers/CJK only to make matching stable across separators.
    return re.sub(r"[\W_]+", "", s, flags=re.UNICODE)




def _resolve_route_by_sender(
    *,
    routing_cfg: RoutingConfig,
    sender: str,
    attachment_name: str,
    subject: str,
) -> str:
    """L1 sender-based routing. Returns subdir or empty string."""
    if not routing_cfg.enabled or not routing_cfg.sender_rules:
        return ""
    sender_lower = (sender or "").lower().strip()
    if not sender_lower:
        return ""
    for rule in routing_cfg.sender_rules:
        if rule.sender != sender_lower:
            continue
        subdir = _sanitize_filename(rule.subdir or "")
        if not subdir:
            continue
        if not rule.keywords:
            return subdir
        combined = _normalize_for_match(f"{attachment_name} {subject}")
        for kw in rule.keywords:
            if _normalize_for_match(str(kw)) in combined:
                return subdir
        return ""
    return ""


def _determine_route_decision(
    *,
    cfg: "ToolConfig",
    save_dir: Path | None,
    sender: str,
    attachment_name: str,
    subject: str,
    project: str | None,
    full_text: str,
) -> tuple[str, str]:
    """Unified routing with audit reason: BAN -> L1 sender -> L2 count -> existing folder -> fallback."""
    if not cfg.routing.enabled:
        return "", "routing_disabled"
    sender_lower = (sender or "").lower().strip()
    if sender_lower in cfg.routing.ban_senders:
        return cfg.routing.fallback_subdir, "ban_sender"
    route = _resolve_route_by_sender(
        routing_cfg=cfg.routing,
        sender=sender,
        attachment_name=attachment_name,
        subject=subject,
    )
    if route:
        return route, "sender_rule"
    route = _resolve_route_by_project_name(
        entries=cfg.project_master,
        project=project,
    )
    if route:
        return route, "project_master_project"
    route = _resolve_route_by_count(
        entries=cfg.project_master,
        full_text=full_text,
        threshold=cfg.routing.threshold,
        tie_margin=cfg.routing.tie_margin,
    )
    if route:
        return route, "project_master"
    route = _resolve_route_by_existing_folders(
        save_dir=save_dir,
        project=project,
        full_text=full_text,
        fallback_subdir=cfg.routing.fallback_subdir,
    )
    if route:
        return route, "existing_folder_match"
    fallback = cfg.routing.fallback_subdir
    if _looks_like_new_project_candidate(project):
        return _new_project_candidate_subdir(fallback), "new_project_candidate"
    if fallback:
        return fallback, "fallback"
    return "", "no_route"


def _resolve_route_by_project_name(
    *,
    entries,
    project: str | None,
) -> str:
    """Use extracted project text for a deterministic project_master hit before fuzzy count."""
    project_name = _clean_project_name(project)
    normalized_project = _normalize_for_match(project_name or "")
    if len(normalized_project) < 6:
        return ""

    scored: list[tuple[int, int, object]] = []
    for entry in entries:
        entry_name = _clean_project_name(getattr(entry, "kojimei", ""))
        entry_norm = _normalize_for_match(entry_name or "")
        if len(entry_norm) < 6:
            continue
        if entry_norm not in normalized_project:
            continue
        scored.append((len(entry_norm), len(entry_name or ""), entry))

    if not scored:
        return ""

    scored.sort(key=lambda item: (-item[0], -item[1], getattr(item[2], "kojiban", "")))
    if len(scored) >= 2 and scored[0][0] == scored[1][0] and scored[0][1] == scored[1][1]:
        return ""

    best_entry = scored[0][2]
    if best_entry.busho == "修繕":
        return "営繕"
    folder = f"{best_entry.kojiban}_{best_entry.kojimei}".strip("_")
    return folder


def _new_project_candidate_subdir(fallback_subdir: str) -> str:
    fallback = _sanitize_filename(fallback_subdir or "")
    m = re.search(r"(\d{8}_\d{4,6})$", fallback)
    if m:
        return f"新規案件候補_{m.group(1)}"
    return "新規案件候補"


def _resolve_route_by_count(
    *,
    entries,
    full_text: str,
    threshold: int = 1,
    tie_margin: float | None = None,
) -> str:
    """
    工事名称またはキーワードの出現回数が threshold 以上の ProjectMasterEntry を探して
    振り分け先フォルダ名を返す。見つからない場合は空文字を返す。

    tie_margin が指定されている場合、最高スコアと2位スコアの差が tie_margin 以下なら
    判定不能として空文字を返す（fallback に振られる）。
    """
    if not entries or not full_text:
        return ""
    normalized = _normalize_for_match(full_text)

    # 全エントリの最高ヒット数を集計
    scored: list[tuple[int, object]] = []  # (max_count, entry)
    for entry in entries:
        all_kw = [entry.kojimei] + list(entry.keywords)
        max_count = 0
        for raw_kw in all_kw:
            kw = _normalize_for_match(raw_kw)
            if not kw:
                continue
            if len(kw) < 3:
                continue
            if kw.isdigit() and len(kw) < 4:
                continue
            count = normalized.count(kw)
            if count > max_count:
                max_count = count
        if max_count >= threshold:
            scored.append((max_count, entry))

    if not scored:
        return ""

    # スコア降順ソート
    scored.sort(key=lambda x: x[0], reverse=True)

    # tie_margin チェック: 2位以上が存在し、差が margin 以下ならタイ判定
    if tie_margin is not None and len(scored) >= 2:
        top_score = scored[0][0]
        second_score = scored[1][0]
        if top_score > 0 and (top_score - second_score) / top_score <= tie_margin:
            return ""  # タイ → fallback

    best_entry = scored[0][1]
    if best_entry.busho == "修繕":
        return "営繕"
    folder = f"{best_entry.kojiban}_{best_entry.kojimei}".strip("_")
    return folder


def _folder_match_tokens(folder_name: str) -> tuple[str, ...]:
    cleaned = _clean_project_name(folder_name) or folder_name
    cleaned = re.sub(r"^\d{3,5}_?", "", cleaned).strip()
    raw_parts = re.split(r"[_\-\s()（）【】\[\]<>＜＞・/]+", cleaned)
    tokens: list[str] = []
    whole = _normalize_for_match(cleaned)
    if len(whole) >= 4:
        tokens.append(whole)
    for part in raw_parts:
        norm = _normalize_for_match(part)
        if len(norm) >= 3 and norm not in tokens:
            tokens.append(norm)
    return tuple(tokens)


def _resolve_route_by_existing_folders(
    *,
    save_dir: Path | None,
    project: str | None,
    full_text: str,
    fallback_subdir: str = "",
) -> str:
    """Use already-created project folders before falling back to unknown."""
    if save_dir is None or not save_dir.exists():
        return ""

    normalized_project = _normalize_for_match(project or "")
    normalized_full = _normalize_for_match(full_text or "")
    if not normalized_project and not normalized_full:
        return ""

    fallback_norm = _normalize_for_match(fallback_subdir or "")
    scored: list[tuple[int, int, str]] = []
    for child in save_dir.iterdir():
        if not child.is_dir():
            continue
        folder_name = child.name.strip()
        folder_norm = _normalize_for_match(folder_name)
        if not folder_norm or folder_norm == fallback_norm:
            continue
        tokens = _folder_match_tokens(folder_name)
        if not tokens:
            continue

        project_hits = [len(tok) for tok in tokens if normalized_project and tok in normalized_project]
        full_hits = [len(tok) for tok in tokens if normalized_full and tok in normalized_full]
        if not project_hits and not full_hits:
            continue

        score = 0
        if project_hits:
            score += 100 + max(project_hits)
        if full_hits:
            score += max(full_hits)
        scored.append((score, len(folder_name), folder_name))

    if not scored:
        return ""

    scored.sort(key=lambda item: (-item[0], -item[1], item[2]))
    if len(scored) >= 2 and scored[0][0] == scored[1][0]:
        return ""
    return scored[0][2]


def _is_ascii_path(path: Path) -> bool:
    try:
        str(path).encode("ascii")
        return True
    except UnicodeEncodeError:
        return False


def _vendor_short(vendor: str) -> str:
    v = vendor.strip()
    # Common Japanese corporate markers (prefix/suffix style).
    for token in [
        "株式会社",
        "有限会社",
        "合同会社",
        "一般社団法人",
        "一般財団法人",
        "税理士法人",
        "社会保険労務士法人",
    ]:
        v = v.replace(token, "")
    for token in ["（株）", "(株)", "（有）", "(有)"]:
        v = v.replace(token, "")
    return v.strip() or vendor.strip()


def _parse_ymd(value: str) -> str | None:
    s = _normalize_text(value)
    m = DATE_YMD_RE.search(s)
    if not m:
        return None
    y = int(m.group("y"))
    mo = int(m.group("m"))
    d = int(m.group("d"))
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return None
    return f"{y:04}{mo:02}{d:02}"


def _normalize_yyyymmdd_candidate(value: str) -> str | None:
    s = re.sub(r"\D", "", _normalize_text(value or ""))

    def _validate(y: int, mo: int, d: int) -> bool:
        if not (2000 <= y <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31):
            return False
        try:
            datetime(y, mo, d)
            return True
        except ValueError:
            return False

    if len(s) == 8:
        y, mo, d = int(s[0:4]), int(s[4:6]), int(s[6:8])
        if _validate(y, mo, d):
            return s
        return None
    if len(s) == 7:
        # YYYYMDD: 月1桁(1-9) + 日2桁。例: "2026125" → 2026-01-25
        y, mo, d = int(s[0:4]), int(s[4:5]), int(s[5:7])
        if _validate(y, mo, d):
            return f"{y:04d}{mo:02d}{d:02d}"
        return None
    if len(s) == 9:
        candidates: list[tuple[int, int, str]] = []
        today = datetime.now().date()
        for i in range(9):
            cand = s[:i] + s[i + 1 :]
            parsed = _normalize_yyyymmdd_candidate(cand)
            if parsed:
                year = int(parsed[0:4])
                month = int(parsed[4:6])
                day = int(parsed[6:8])
                cand_date = datetime(year, month, day).date()
                recent_penalty = 0 if cand_date <= today else 1
                candidates.append((recent_penalty, abs((today - cand_date).days), parsed))
        if candidates:
            candidates.sort(key=lambda item: (item[0], item[1], item[2]))
            return candidates[0][2]
    return None


def _parse_amount_token(value: str) -> int | None:
    s = _normalize_text(value or "").strip()
    s = s.replace("円", "").replace("¥", "").replace("￥", "")
    s = s.replace(",", "").replace("，", "")
    if not s.isdigit():
        return None
    amount = int(s)
    if amount <= 0:
        return None
    return amount


def _extract_partial_invoice_fields_from_filename(
    filename: str, company_deny_regex: str | None
) -> PartialInvoiceFields:
    normalized = normalize("NFKC", filename or "")
    stem = normalized.removesuffix(".pdf").strip()
    stem = stem.replace("＿", "_")
    parts = [p.strip() for p in re.split(r"[_\-]+", stem) if p.strip()]
    deny_re = re.compile(company_deny_regex) if company_deny_regex else None

    issue_date: str | None = None
    date_idx: int | None = None
    for idx, part in enumerate(parts):
        parsed = _parse_ymd(part) or _normalize_yyyymmdd_candidate(part)
        if parsed:
            issue_date = parsed
            date_idx = idx
            break

    amount: int | None = None
    amount_idx: int | None = None
    for idx, part in enumerate(parts):
        parsed_amount = _parse_amount_token(part)
        if parsed_amount is None:
            continue
        # If date exists, prefer amount after date.
        if date_idx is not None and idx <= date_idx:
            continue
        amount = parsed_amount
        amount_idx = idx
        break
    if amount is None:
        for idx, part in enumerate(parts):
            parsed_amount = _parse_amount_token(part)
            if parsed_amount is None:
                continue
            # Guard: YYYYMMDD 形式の数値（日付値）を amount として誤採用しない。
            # 20000101〜21001231 の範囲は日付リテラルとみなしてスキップ。
            if 20_000_000 <= parsed_amount <= 21_001_231:
                continue
            amount = parsed_amount
            amount_idx = idx
            break

    vendor: str | None = None
    if date_idx is not None and date_idx > 0:
        vendor = "_".join(parts[:date_idx]).strip()
    elif amount_idx is not None and amount_idx > 0:
        vendor = "_".join(parts[:amount_idx]).strip()
    elif parts:
        vendor = parts[0]

    if vendor:
        vendor = _clean_project_name(vendor)
        if vendor and deny_re and deny_re.search(vendor):
            vendor = None
        # Ignore vendor candidates that are effectively numeric.
        if vendor and re.fullmatch(r"[\d_ .-]+", vendor):
            vendor = None

    project: str | None = None
    if amount_idx is not None and amount_idx + 1 < len(parts):
        project = _clean_project_name("_".join(parts[amount_idx + 1 :]))
    elif date_idx is not None and date_idx + 1 < len(parts):
        project = _clean_project_name("_".join(parts[date_idx + 1 :]))

    return PartialInvoiceFields(
        vendor=vendor,
        issue_date=issue_date,
        amount=amount,
        project=project,
    )


def _build_review_filename(
    *,
    received_dt: datetime,
    attachment_name: str,
    partial: PartialInvoiceFields,
    review_reasons: Sequence[str] = (),
    route_reason: str | None = None,
) -> str:
    stem = normalize("NFKC", attachment_name or "").removesuffix(".pdf").strip()
    stem = _clean_project_name(stem) or "原本名不明"
    vendor = _clean_project_name(partial.vendor) or "vendor不明"
    issue_date = partial.issue_date or received_dt.strftime("%Y%m%d")
    amount = f"{partial.amount:,}" if partial.amount is not None else "金額不明"
    project = _clean_project_name(partial.project) or stem
    prefix = "要確認"
    if any(str(reason).startswith("non_invoice_signal:") for reason in review_reasons):
        prefix = "要確認_非請求書候補"
    elif route_reason == "new_project_candidate":
        prefix = "要確認_新規案件候補"
    candidate = f"{prefix}_{vendor}_{issue_date}_{amount}_{project}.pdf"
    return _cleanup_rendered_filename(_sanitize_filename(candidate))


def _processed_attachment_key(entry_id: str, attachment_name: str, sha256: str) -> str:
    raw = "\n".join(
        [
            str(entry_id or "").strip(),
            normalize("NFKC", str(attachment_name or "").strip()),
            str(sha256 or "").strip(),
        ]
    )
    return hashlib.sha1(raw.encode("utf-8", errors="ignore")).hexdigest()


def _load_processed_attachment_index(path: Path) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    if not path.exists():
        return index
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            key = str(record.get("key") or "").strip()
            saved_path = str(record.get("saved_path") or "").strip()
            if not key or not saved_path:
                continue
            index[key] = record
    return index


def _append_processed_attachment_record(path: Path, record: dict[str, Any]) -> None:
    _ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


def _is_within_dir(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except Exception:
        return False


def _extract_pdf_text_pypdf(path: Path, max_pages: int) -> str:
    if not _PYPDF_AVAILABLE:
        raise RuntimeError("pypdf が利用できません。PDF抽出/復号のために必要です。")
    reader = PdfReader(str(path))
    texts: list[str] = []
    for i, page in enumerate(reader.pages):
        if i >= max_pages:
            break
        try:
            t = page.extract_text() or ""
        except Exception:
            t = ""
        if t:
            texts.append(t)
    return _normalize_text("\n".join(texts))


def _extract_pdf_text_pymupdf(path: Path, max_pages: int) -> str:
    if not _PYMUPDF_AVAILABLE:
        raise RuntimeError("PyMuPDF (fitz) が利用できません。")
    tmp_dir: tempfile.TemporaryDirectory[str] | None = None
    open_path = path
    # Some native libraries have trouble with non-ASCII paths on Windows; use a temp ASCII path if needed.
    if not _is_ascii_path(path):
        tmp_dir = tempfile.TemporaryDirectory(prefix="s12_13_pymupdf_")
        open_path = Path(tmp_dir.name) / "input.pdf"
        shutil.copy2(path, open_path)

    doc = fitz.open(str(open_path))
    try:
        texts: list[str] = []
        for i in range(min(max_pages, int(doc.page_count))):
            try:
                page = doc.load_page(i)
                t = page.get_text("text") or ""
            except Exception:
                t = ""
            if t:
                texts.append(t)
        return _normalize_text("\n".join(texts))
    finally:
        try:
            doc.close()
        finally:
            if tmp_dir is not None:
                try:
                    tmp_dir.cleanup()
                except Exception:
                    pass


def _render_pdf_first_page_png(pdf_path: Path, out_path: Path, dpi: int = 200) -> None:
    if not _PYMUPDF_AVAILABLE:
        raise RuntimeError("PyMuPDF (fitz) が利用できません。")
    tmp_dir: tempfile.TemporaryDirectory[str] | None = None
    open_path = pdf_path
    if not _is_ascii_path(pdf_path):
        tmp_dir = tempfile.TemporaryDirectory(prefix="s12_13_render_")
        open_path = Path(tmp_dir.name) / "input.pdf"
        shutil.copy2(pdf_path, open_path)

    doc = fitz.open(str(open_path))
    try:
        page = doc.load_page(0)
        zoom = dpi / 72.0
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        pix.save(str(out_path))
    finally:
        try:
            doc.close()
        finally:
            if tmp_dir is not None:
                try:
                    tmp_dir.cleanup()
                except Exception:
                    pass


_EASYOCR_READERS: dict[tuple[tuple[str, ...], bool], Any] = {}


def _get_easyocr_reader(
    *,
    languages: Sequence[str] = ("ja", "en"),
    use_gpu: bool = False,
) -> Any:
    key = (tuple(languages), use_gpu)
    cached = _EASYOCR_READERS.get(key)
    if cached is not None:
        return cached

    if not _EASYOCR_AVAILABLE:
        raise RuntimeError("easyocr が利用できません。")
    try:
        import easyocr  # type: ignore
    except Exception as e:
        raise RuntimeError("easyocr が利用できません。") from e

    # Reader initialization can be heavy; keep it cached per (langs, gpu).
    reader = easyocr.Reader(list(languages), gpu=use_gpu, verbose=False)
    _EASYOCR_READERS[key] = reader
    return reader


def _extract_pdf_text_easyocr_ocr(pdf_path: Path, *, run_dir: Path) -> str:
    """
    OCR the first page using easyocr (fallback for PDFs where text extraction
    produces garbled characters due to font mappings).

    Returns the extracted text (best-effort) or raises on total failure.
    """

    if not (_EASYOCR_AVAILABLE and _PYMUPDF_AVAILABLE):
        raise RuntimeError("easyocr / PyMuPDF が利用できません。")

    # Use an ASCII temp directory to avoid native path issues on Windows.
    tmp_obj = tempfile.TemporaryDirectory(prefix="s12_13_easyocr_")
    tmp_dir = Path(tmp_obj.name)
    try:
        attempts: list[int] = [200, 300]
        best = ""
        best_dpi: int | None = None

        reader = _get_easyocr_reader(languages=("ja", "en"), use_gpu=False)

        for dpi in attempts:
            tmp_img = tmp_dir / f"p1_{dpi}.png"
            try:
                _render_pdf_first_page_png(pdf_path, tmp_img, dpi=dpi)
            except Exception:
                continue

            # Apply pdf-ocr Skill preprocessing (deskew + shadow removal).
            preproc_img = _preprocess_png_for_ocr(tmp_img)

            try:
                # detail=0 returns list[str]. Keep as lines to preserve signals.
                lines = reader.readtext(str(preproc_img), detail=0)
            except Exception:
                continue

            text = _normalize_text(
                "\n".join([str(ln).strip() for ln in lines if str(ln).strip()])
            )
            if len(text) > len(best):
                best = text
                best_dpi = dpi

            # If it already looks usable, stop early to save time.
            if VENDOR_HINT_RE.search(best) and DATE_YMD_RE.search(best) and (
                AMOUNT_LABEL_RE.search(best)
                or re.search(r"[0-9]{1,3}(?:,[0-9]{3})+", best)
            ):
                break

        if not best:
            raise RuntimeError("easyocr OCR でテキスト抽出できませんでした。")

        # Persist the chosen render for post-mortem debugging (best-effort).
        if best_dpi is not None:
            try:
                ocr_dir = run_dir / "ocr"
                _ensure_dir(ocr_dir)
                img_out = (
                    ocr_dir
                    / f"{_sanitize_filename(pdf_path.stem)}__easyocr_p1_{best_dpi}.png"
                )
                _render_pdf_first_page_png(pdf_path, img_out, dpi=best_dpi)
            except Exception:
                pass

        return best
    finally:
        try:
            tmp_obj.cleanup()
        except Exception:
            pass


def _get_tesseract_tessdata_dir() -> Path | None:
    """
    Determine tessdata directory.

    Priority:
    1) Adjacent directory next to this script: ./tessdata
    2) Environment variable TESSDATA_PREFIX (direct or +/tessdata)
    3) Tesseract install dir (exe_dir/tessdata)
    """

    local = Path(__file__).parent / "tessdata"
    if local.exists():
        return local

    env = os.environ.get("TESSDATA_PREFIX")
    if env:
        base = Path(env)
        if base.exists():
            if (base / "tessdata").exists():
                return base / "tessdata"
            return base

    if _TESSERACT_EXE is not None:
        guess = Path(_TESSERACT_EXE).parent / "tessdata"
        if guess.exists():
            return guess

    return None


def _tesseract_image_to_text(
    image_path: Path,
    *,
    languages: str,
    tessdata_dir: Path | None,
    timeout_s: int,
) -> str:
    if not _TESSERACT_AVAILABLE or _TESSERACT_EXE is None:
        raise RuntimeError("tesseract が利用できません。")

    cmd = [
        str(_TESSERACT_EXE),
        str(image_path),
        "stdout",
        "-l",
        languages,
        "--psm",
        "6",
    ]
    if tessdata_dir is not None:
        cmd.extend(["--tessdata-dir", str(tessdata_dir)])

    proc = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout_s,
    )
    if proc.returncode != 0:
        err = (proc.stderr or "").strip()
        raise RuntimeError(f"tesseract failed: rc={proc.returncode} stderr={err}")
    return proc.stdout or ""


def _preprocess_png_for_ocr(img_path: Path) -> Path:
    """Apply deskew and shadow removal to a rendered PNG before OCR.

    Implements algorithms from pdf-ocr Skill (pdf_preprocess.py template):
    - Deskew: Hough transform line detection → median angle → warpAffine
    - Shadow removal: block brightness std deviation → GaussianBlur background subtraction

    Returns path to preprocessed image (new file) or original if unchanged/unavailable.
    """
    if not _CV2_AVAILABLE:
        return img_path
    try:
        img = _cv2.imread(str(img_path))
        if img is None:
            return img_path

        modified = False

        # --- Deskew (Hough transform, from pdf_preprocess.py PDFPreprocessor.detect_skew_angle_hough) ---
        gray = _cv2.cvtColor(img, _cv2.COLOR_BGR2GRAY)
        clahe = _cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
        enhanced_gray = clahe.apply(gray)
        edges = _cv2.Canny(enhanced_gray, 50, 150, apertureSize=3)
        lines = _cv2.HoughLines(edges, 1, _np.pi / 180, 80)

        skew_angle = 0.0
        if lines is not None and len(lines) > 0:
            angles = []
            for line in lines:
                theta = line[0][1]
                angle_deg = float(_np.degrees(theta)) - 90.0
                if -20.0 <= angle_deg <= 20.0:
                    angles.append(angle_deg)
            if angles:
                skew_angle = float(_np.median(angles))

        if abs(skew_angle) >= 0.5:  # SKEW_THRESHOLD_DEFAULT from pdf_preprocess.py
            h, w = img.shape[:2]
            center = (w // 2, h // 2)
            M = _cv2.getRotationMatrix2D(center, skew_angle, 1.0)
            img = _cv2.warpAffine(
                img, M, (w, h),
                flags=_cv2.INTER_CUBIC,
                borderMode=_cv2.BORDER_CONSTANT,
                borderValue=(255, 255, 255),
            )
            modified = True

        # --- Shadow removal (from pdf_preprocess.py PDFPreprocessor.detect_shadow / remove_shadow) ---
        gray2 = _cv2.cvtColor(img, _cv2.COLOR_BGR2GRAY)
        h_img, w_img = gray2.shape
        block_h, block_w = max(1, h_img // 4), max(1, w_img // 4)
        brightnesses = []
        for bi in range(4):
            for bj in range(4):
                block = gray2[bi * block_h:(bi + 1) * block_h, bj * block_w:(bj + 1) * block_w]
                brightnesses.append(float(_np.mean(block)))
        has_shadow = float(_np.std(brightnesses)) > 30.0

        if has_shadow:
            kernel_size = max(h_img, w_img) // 10
            if kernel_size % 2 == 0:
                kernel_size += 1
            kernel_size = max(kernel_size, 51)
            background = _cv2.GaussianBlur(gray2, (kernel_size, kernel_size), 0)
            mean_bg = float(_np.mean(background))
            with _np.errstate(divide="ignore", invalid="ignore"):
                normalized = gray2.astype(_np.float32) * (mean_bg / background.astype(_np.float32))
                normalized = _np.clip(normalized, 0, 255).astype(_np.uint8)
            pct_lo, pct_hi = _np.percentile(normalized, [2, 98])
            if pct_hi > pct_lo:
                stretched = _np.clip(
                    (normalized - pct_lo) * 255.0 / (pct_hi - pct_lo), 0, 255
                ).astype(_np.uint8)
            else:
                stretched = normalized
            img = _cv2.cvtColor(stretched, _cv2.COLOR_GRAY2BGR)
            modified = True

        if not modified:
            return img_path

        preproc_path = img_path.parent / f"_preproc_{img_path.name}"
        _cv2.imwrite(str(preproc_path), img)
        return preproc_path

    except Exception:
        return img_path


def _extract_pdf_text_tesseract_ocr(pdf_path: Path, *, run_dir: Path) -> str:
    """
    OCR the first page using tesseract (no Torch dependency).

    Returns extracted text or raises on total failure.
    """

    if not (_TESSERACT_AVAILABLE and _PYMUPDF_AVAILABLE):
        raise RuntimeError("tesseract / PyMuPDF が利用できません。")

    tessdata_dir = _get_tesseract_tessdata_dir()
    # Prefer jpn+eng, but be robust if tessdata_dir does not contain eng.
    languages = "jpn+eng"
    if tessdata_dir is not None:
        if (tessdata_dir / "jpn.traineddata").exists() and not (
            tessdata_dir / "eng.traineddata"
        ).exists():
            languages = "jpn"

    # Use an ASCII temp directory to avoid native path issues on Windows.
    tmp_obj = tempfile.TemporaryDirectory(prefix="s12_13_tess_")
    tmp_dir = Path(tmp_obj.name)
    try:
        attempts: list[int] = [200, 250, 300]
        best = ""
        best_dpi: int | None = None
        last_error: BaseException | None = None

        for dpi in attempts:
            tmp_img = tmp_dir / f"p1_{dpi}.png"
            try:
                _render_pdf_first_page_png(pdf_path, tmp_img, dpi=dpi)
            except Exception as e:
                last_error = e
                continue

            # Apply pdf-ocr Skill preprocessing (deskew + shadow removal).
            preproc_img = _preprocess_png_for_ocr(tmp_img)

            try:
                raw = _tesseract_image_to_text(
                    preproc_img,
                    languages=languages,
                    tessdata_dir=tessdata_dir,
                    timeout_s=60,
                )
            except Exception as e:
                last_error = e
                continue

            text = _normalize_text(raw)
            if len(text) > len(best):
                best = text
                best_dpi = dpi

            if VENDOR_HINT_RE.search(best) and DATE_YMD_RE.search(best) and (
                AMOUNT_LABEL_RE.search(best)
                or re.search(r"[0-9]{1,3}(?:,[0-9]{3})+", best)
            ):
                break

        if not best:
            hints: list[str] = []
            if tessdata_dir is None:
                hints.append("tessdata_dir=not_found")
            else:
                # When --tessdata-dir is specified, all languages must exist there.
                if not (tessdata_dir / "jpn.traineddata").exists():
                    hints.append(f"jpn.traineddata=missing ({tessdata_dir})")
                if languages == "jpn+eng" and not (tessdata_dir / "eng.traineddata").exists():
                    hints.append(f"eng.traineddata=missing ({tessdata_dir})")
            if last_error is not None:
                hints.append(f"last_error={type(last_error).__name__}: {last_error}")
            hint = (" " + "; ".join(hints)) if hints else ""
            raise RuntimeError(f"tesseract OCR でテキスト抽出できませんでした。{hint}".strip())

        # Persist the chosen render for post-mortem debugging (best-effort).
        if best_dpi is not None:
            try:
                ocr_dir = run_dir / "ocr"
                _ensure_dir(ocr_dir)
                img_out = (
                    ocr_dir
                    / f"{_sanitize_filename(pdf_path.stem)}__tesseract_p1_{best_dpi}.png"
                )
                _render_pdf_first_page_png(pdf_path, img_out, dpi=best_dpi)
            except Exception:
                pass

        return best
    finally:
        try:
            tmp_obj.cleanup()
        except Exception:
            pass


_YOMITOKU_ANALYZERS: dict[tuple[bool, bool], Any] = {}


def _get_yomitoku_analyzer(*, use_gpu: bool = False, lite_mode: bool = True) -> Any:
    key = (use_gpu, lite_mode)
    cached = _YOMITOKU_ANALYZERS.get(key)
    if cached is not None:
        return cached

    if not _YOMITOKU_AVAILABLE:
        raise RuntimeError("yomitoku が利用できません。")
    try:
        from yomitoku import DocumentAnalyzer  # type: ignore
    except Exception as e:
        raise RuntimeError("yomitoku が利用できません。") from e
    device = "cuda" if use_gpu else "cpu"
    configs: dict[str, Any] = {}
    if lite_mode:
        configs = {
            "ocr": {"model_name": "lite"},
            "layout_analyzer": {"model_name": "lite"},
        }
    analyzer = DocumentAnalyzer(configs=configs, device=device)
    _YOMITOKU_ANALYZERS[key] = analyzer
    return analyzer


def _yomitoku_content_to_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        s = str(value)
    except Exception:
        return ""
    # YomiToku sometimes inserts '\n' between each character. Remove them.
    s = s.replace("\r", "").replace("\n", "")
    s = _normalize_text(s).strip()
    if not s or s == "None":
        return ""
    return s


def _extract_text_from_yomitoku_result(result: Any) -> str:
    parts: list[str] = []

    # paragraphs
    for p in getattr(result, "paragraphs", []) or []:
        txt = _yomitoku_content_to_text(getattr(p, "contents", None))
        if txt:
            parts.append(txt)

    # figures (some PDFs embed key info as figure-like blocks)
    for fig in getattr(result, "figures", []) or []:
        txt = _yomitoku_content_to_text(getattr(fig, "contents", None))
        if txt:
            parts.append(txt)
        cap = _yomitoku_content_to_text(getattr(fig, "caption", None))
        if cap:
            parts.append(cap)
        for p in getattr(fig, "paragraphs", []) or []:
            txt = _yomitoku_content_to_text(getattr(p, "contents", None))
            if txt:
                parts.append(txt)

    # tables (invoice key fields are often in tables)
    for tbl in getattr(result, "tables", []) or []:
        # Most common: tbl.cells -> list of cell objects with .contents
        cells = getattr(tbl, "cells", None)
        if cells:
            for cell in cells:
                txt = _yomitoku_content_to_text(getattr(cell, "contents", None))
                if txt:
                    parts.append(txt)
            continue

        # Fallback: tbl.data can be nested (rows/cols)
        data = getattr(tbl, "data", None)
        if isinstance(data, list):
            for row in data:
                if not isinstance(row, list):
                    continue
                for cell in row:
                    txt = _yomitoku_content_to_text(getattr(cell, "contents", None))
                    if txt:
                        parts.append(txt)

    return _normalize_text("\n".join([p for p in parts if p]))


def _extract_pdf_text_yomitoku_ocr(pdf_path: Path, *, run_dir: Path) -> str:
    if not (_YOMITOKU_AVAILABLE and _PYMUPDF_AVAILABLE):
        raise RuntimeError("yomitoku / PyMuPDF が利用できません。")
    try:
        from yomitoku.data.functions import load_image  # type: ignore
    except Exception as e:
        raise RuntimeError("yomitoku が利用できません。") from e
    # Use an ASCII temp directory to avoid native path issues on Windows.
    tmp_obj = tempfile.TemporaryDirectory(prefix="s12_13_ocr_")
    tmp_dir = Path(tmp_obj.name)
    try:
        attempts: list[tuple[int, bool]] = [
            (200, True),
            (300, True),
            (300, False),
        ]
        best = ""
        for dpi, lite_mode in attempts:
            tmp_img = tmp_dir / f"p1_{dpi}_{'lite' if lite_mode else 'full'}.png"
            _render_pdf_first_page_png(pdf_path, tmp_img, dpi=dpi)
            analyzer = _get_yomitoku_analyzer(use_gpu=False, lite_mode=lite_mode)
            img = load_image(str(tmp_img))
            # yomitoku 0.11.0: load_image() may return nested list
            while isinstance(img, (list, tuple)):
                if not img:
                    raise RuntimeError(f"load_image returned empty: {tmp_img}")
                img = img[0]
            result, _, _ = analyzer(img)
            text = _extract_text_from_yomitoku_result(result)
            if len(text) > len(best):
                best = text

            # If it already looks usable, stop early to save time.
            if VENDOR_HINT_RE.search(best) and DATE_YMD_RE.search(best) and (
                AMOUNT_LABEL_RE.search(best)
                or re.search(r"[0-9]{1,3}(?:,[0-9]{3})+", best)
            ):
                break

        return best
    finally:
        try:
            tmp_obj.cleanup()
        except Exception:
            pass


def _extract_invoice_fields_with_vision_ocr(
    pdf_path: Path,
    *,
    company_deny_regex: str | None,
    log_path: Path,
    vision_ocr_provider: str,
    sender: str | None = None,
    subject: str | None = None,
) -> ExtractedPdfFields | None:
    try:
        from vision_ocr import extract as _vision_extract
    except Exception as e:
        _append_log(log_path, f"[WARN] vision_ocr import failed: {pdf_path.name} error={e}")
        return None

    vr = _vision_extract(
        pdf_path,
        provider=vision_ocr_provider,
        max_pages=3,
        timeout_s=30.0,
        sender_hint=sender,
        subject_hint=subject,
    )
    if vr.error is not None:
        _append_log(log_path, f"[WARN] vision_ocr returned error: {pdf_path.name} error={vr.error}")
        return None

    parsed_from_text = None
    if vr.text:
        try:
            parsed_from_text = _extract_invoice_fields(
                text=vr.text,
                company_deny_regex=company_deny_regex,
            )
        except Exception as e:
            _append_log(log_path, f"[WARN] vision_ocr text parse failed: {pdf_path.name} error={e}")

    vendor = _clean_project_name(vr.vendor or (parsed_from_text.vendor if parsed_from_text else None))
    if vendor and company_deny_regex and re.search(company_deny_regex, vendor):
        vendor = None
    issue_date = vr.issue_date or (parsed_from_text.issue_date if parsed_from_text else None)
    amount = _parse_amount_token(str(vr.amount)) if vr.amount is not None else None
    if amount is None and parsed_from_text is not None:
        amount = parsed_from_text.amount
    invoice_no = vr.invoice_no or (parsed_from_text.invoice_no if parsed_from_text else None)
    project = parsed_from_text.project if parsed_from_text else None
    raw_ocr_text = vr.text or (parsed_from_text.raw_ocr_text if parsed_from_text else "")

    _append_log(
        log_path,
        "[INFO] vision_ocr result: "
        f"{pdf_path.name} provider={vr.provider} confidence={vr.confidence} "
        f"category={vr.vendor_category or '-'} manual={vr.requires_manual} "
        f"reasons={','.join(vr.review_reasons) if vr.review_reasons else '-'}",
    )

    if not vendor or not issue_date or amount is None:
        _append_log(
            log_path,
            "[INFO] vision_ocr incomplete: "
            f"{pdf_path.name} vendor={'y' if vendor else 'n'} "
            f"date={'y' if issue_date else 'n'} amount={'y' if amount is not None else 'n'}",
        )
        return None

    return ExtractedPdfFields(
        vendor=vendor,
        issue_date=issue_date,
        amount=amount,
        invoice_no=invoice_no,
        project=project,
        raw_ocr_text=raw_ocr_text,
        source="vision_ocr",
        vendor_category=vr.vendor_category,
        requires_manual=vr.requires_manual,
        review_reasons=tuple(vr.review_reasons),
    )


def _try_extract_text_method(
    extractor: Callable[[], str],
    *,
    company_deny_regex: str | None,
    pdf_path: Path,
    log_path: Path,
    warn_prefix: str,
) -> ExtractedPdfFields | None:
    """Run one text-extraction method and parse invoice fields. Logs WARN on failure."""
    try:
        text = extractor()
        fields = _extract_invoice_fields(text=text, company_deny_regex=company_deny_regex)
        if fields:
            return fields
    except Exception as e:
        _append_log(log_path, f"[WARN] {warn_prefix}: {pdf_path.name} error={e}")
    return None


def _extract_invoice_fields_from_pdf(
    pdf_path: Path,
    *,
    company_deny_regex: str | None,
    max_pages: int,
    run_dir: Path,
    log_path: Path,
    vision_ocr_enabled: bool = False,
    vision_ocr_provider: str = "openai",
    sender: str | None = None,
    subject: str | None = None,
) -> ExtractedPdfFields | None:
    # 1) pypdf
    if _PYPDF_AVAILABLE:
        fields = _try_extract_text_method(
            lambda: _extract_pdf_text_pypdf(pdf_path, max_pages=max_pages),
            company_deny_regex=company_deny_regex,
            pdf_path=pdf_path,
            log_path=log_path,
            warn_prefix="pypdf text extract failed",
        )
        if fields:
            return fields

    # 2) PyMuPDF
    if _PYMUPDF_AVAILABLE:
        fields = _try_extract_text_method(
            lambda: _extract_pdf_text_pymupdf(pdf_path, max_pages=max_pages),
            company_deny_regex=company_deny_regex,
            pdf_path=pdf_path,
            log_path=log_path,
            warn_prefix="pymupdf text extract failed",
        )
        if fields:
            return fields

    # 3) Vision API OCR (config-driven, prioritized cloud OCR)
    if vision_ocr_enabled:
        try:
            fields = _extract_invoice_fields_with_vision_ocr(
                pdf_path,
                company_deny_regex=company_deny_regex,
                log_path=log_path,
                vision_ocr_provider=vision_ocr_provider,
                sender=sender,
                subject=subject,
            )
            if fields:
                return fields
        except Exception as e:
            _append_log(log_path, f"[WARN] vision_ocr failed: {pdf_path.name} error={e}")

    # 4) tesseract OCR
    if _TESSERACT_AVAILABLE and _PYMUPDF_AVAILABLE:
        fields = _try_extract_text_method(
            lambda: _extract_pdf_text_tesseract_ocr(pdf_path, run_dir=run_dir),
            company_deny_regex=company_deny_regex,
            pdf_path=pdf_path,
            log_path=log_path,
            warn_prefix="tesseract ocr failed",
        )
        if fields:
            return fields

    # 5) EasyOCR (faster than YomiToku, good Japanese support)
    if _EASYOCR_AVAILABLE and _PYMUPDF_AVAILABLE:
        fields = _try_extract_text_method(
            lambda: _extract_pdf_text_easyocr_ocr(pdf_path, run_dir=run_dir),
            company_deny_regex=company_deny_regex,
            pdf_path=pdf_path,
            log_path=log_path,
            warn_prefix="easyocr ocr failed",
        )
        if fields:
            return fields

    # 6) YomiToku OCR (optional fallback)
    if _YOMITOKU_AVAILABLE and _PYMUPDF_AVAILABLE:
        fields = _try_extract_text_method(
            lambda: _extract_pdf_text_yomitoku_ocr(pdf_path, run_dir=run_dir),
            company_deny_regex=company_deny_regex,
            pdf_path=pdf_path,
            log_path=log_path,
            warn_prefix="yomitoku ocr failed",
        )
        if fields:
            return fields

    return None


def _expand_pdf_paths_for_debug(raw_paths: Sequence[str]) -> list[Path]:
    expanded: list[Path] = []
    for raw in raw_paths:
        p = Path(str(raw).strip())
        if p.is_dir():
            expanded.extend(sorted(p.rglob("*.pdf")))
        else:
            expanded.append(p)

    seen: set[str] = set()
    uniq: list[Path] = []
    for p in expanded:
        try:
            key = str(p.resolve())
        except Exception:
            key = str(p)
        if key in seen:
            continue
        seen.add(key)
        uniq.append(p)
    return uniq


def _write_debug_text(path: Path, text: str) -> None:
    try:
        path.write_text(text, encoding="utf-8")
    except Exception:
        # Best-effort: do not break debug flow on encoding issues.
        path.write_text(text, encoding="utf-8", errors="replace")


def _debug_extract_invoice_fields_from_pdf(
    pdf_path: Path,
    *,
    company_deny_regex: str | None,
    max_pages: int,
    debug_run_dir: Path,
    log_path: Path,
) -> ExtractedPdfFields | None:
    """
    Debug helper: extract invoice fields from a PDF and write intermediate
    extraction artifacts (texts + page1 image).

    This function does not move files and does not send emails.
    """

    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")
    if pdf_path.suffix.lower() != ".pdf":
        raise ValueError(f"Not a PDF: {pdf_path}")

    # Avoid clobbering debug outputs when multiple PDFs share the same file name.
    path_hash = hashlib.sha1(str(pdf_path).encode("utf-8", errors="ignore")).hexdigest()[:8]
    per_pdf_dir = debug_run_dir / f"{_sanitize_filename(pdf_path.stem)}__{path_hash}"
    _ensure_dir(per_pdf_dir)

    # Render the first page for manual inspection.
    if _PYMUPDF_AVAILABLE:
        try:
            _render_pdf_first_page_png(
                pdf_path, per_pdf_dir / "page1_200dpi.png", dpi=200
            )
        except Exception as e:
            _append_log(log_path, f"[WARN] render page1 failed: {pdf_path.name} error={e}")

    def _try_extract(method: str, extractor: Callable[[], str]) -> ExtractedPdfFields | None:
        try:
            text = extractor()
            _write_debug_text(per_pdf_dir / f"{method}.txt", text)
            fields = _extract_invoice_fields(text=text, company_deny_regex=company_deny_regex)
            _append_log(
                log_path,
                f"debug_extract {method}: pdf={pdf_path.name} text_len={len(text)} fields={'OK' if fields else 'NG'}",
            )
            if fields:
                _write_debug_text(per_pdf_dir / "picked_method.txt", method)
                return fields
        except Exception as e:
            _append_log(log_path, f"[WARN] debug_extract {method} failed: {pdf_path.name} error={e}")
        return None

    # 1) pypdf
    if _PYPDF_AVAILABLE:
        fields = _try_extract(
            "pypdf", lambda: _extract_pdf_text_pypdf(pdf_path, max_pages=max_pages)
        )
        if fields:
            return fields

    # 2) PyMuPDF
    if _PYMUPDF_AVAILABLE:
        fields = _try_extract(
            "pymupdf", lambda: _extract_pdf_text_pymupdf(pdf_path, max_pages=max_pages)
        )
        if fields:
            return fields

    # 3) tesseract OCR
    if _TESSERACT_AVAILABLE and _PYMUPDF_AVAILABLE:
        fields = _try_extract(
            "tesseract",
            lambda: _extract_pdf_text_tesseract_ocr(pdf_path, run_dir=per_pdf_dir),
        )
        if fields:
            return fields

    # 4) YomiToku OCR
    if _YOMITOKU_AVAILABLE and _PYMUPDF_AVAILABLE:
        fields = _try_extract(
            "yomitoku",
            lambda: _extract_pdf_text_yomitoku_ocr(pdf_path, run_dir=per_pdf_dir),
        )
        if fields:
            return fields

    return None


def _build_invoice_fields_from_pattern(
    vendor: str,
    date_str: str,
    amount_str: str,
    project: str | None,
    deny_re: re.Pattern | None,
) -> ExtractedPdfFields | None:
    """Deny-list check + amount parse → ExtractedPdfFields or None."""
    if not date_str:
        return None
    if deny_re and deny_re.search(vendor):
        return None
    try:
        return ExtractedPdfFields(
            vendor=vendor,
            issue_date=date_str,
            amount=int(amount_str),
            invoice_no=None,
            project=project,
        )
    except ValueError:
        return None


def _extract_invoice_fields_from_filename(
    filename: str, company_deny_regex: str | None
) -> ExtractedPdfFields | None:
    """
    Extract invoice fields from filename before PDF extraction.

    Supported patterns:
    1. {vendor}_{YYYYMMDD}_{amount}_{construction}.pdf (44.3%)
    2. {vendor}_{YYYYMMDD}_{amount}.pdf (22.1%)
    3. {vendor}-{YYYYMMDD}-{amount}-{construction}.pdf (dash variant)
    4. {vendor}_{YYYY.MM.DD}_{amount}_{construction}.pdf (dot-date variant)
    5. {YYYY},{MM},{DD}{vendor}{amount}.pdf (legacy comma format)

    Returns ExtractedPdfFields or None if no pattern matches.
    """
    # Normalize and remove extension
    normalized = normalize("NFKC", filename)
    stem = normalized.removesuffix(".pdf")

    deny_re = re.compile(company_deny_regex) if company_deny_regex else None
    project: str | None = None

    # Pattern 1 & 2: vendor_YYYYMMDD_amount[_construction]
    # Example: 名鉄協商_20251227_400.pdf or 名鉄協商_20251227_400_東海興業.pdf
    # Also handles 9-digit date typos (e.g. 202501225 → 20251225 by removing the extra digit).
    pattern_a = re.compile(
        r"^(?P<vendor>.+?)_(?P<date>\d{6,9})_(?P<amount>[\d,，]+)(?:[_（【([](?P<construction>.+))?$"
    )
    m = pattern_a.match(stem)
    if m:
        vendor = m.group("vendor")
        date_str = m.group("date")
        project = _clean_project_name(m.group("construction"))
        amount_str = m.group("amount").replace(",", "").replace("，", "")

        # Validate / correct date_str to exactly 8 digits (YYYYMMDD).
        if len(date_str) != 8:
            corrected = _normalize_yyyymmdd_candidate(date_str)
            if corrected is not None:
                date_str = corrected
            elif len(date_str) == 6:
                # YYMMDD: 年2桁省略。例: "260125" → "20260125"
                # ここ（日付スロット位置）でのみ解釈。グローバルには適用しない。
                try:
                    yy, mo, d = int(date_str[0:2]), int(date_str[2:4]), int(date_str[4:6])
                    y = 2000 + yy  # 2000-2099 と仮定
                    datetime(y, mo, d)
                    date_str = f"{y:04d}{mo:02d}{d:02d}"
                except (ValueError, IndexError):
                    date_str = ""
            else:
                date_str = ""  # Cannot recover a valid date; fall through.

        result = _build_invoice_fields_from_pattern(vendor, date_str, amount_str, project, deny_re)
        if result is not None:
            return result

    # Pattern 3: vendor-YYYYMMDD-amount[-construction] (dash separator)
    pattern_dash = re.compile(
        r"^(?P<vendor>.+?)-(?P<date>\d{8})-(?P<amount>[\d,，]+)(?:-(?P<construction>.+))?$"
    )
    m = pattern_dash.match(stem)
    if m:
        vendor = m.group("vendor")
        date_str = m.group("date")
        project = _clean_project_name(m.group("construction"))
        amount_str = m.group("amount").replace(",", "").replace("，", "")

        result = _build_invoice_fields_from_pattern(vendor, date_str, amount_str, project, deny_re)
        if result is not None:
            return result

    # Pattern 4: vendor_YYYY.MM.DD_amount[_construction] (dot-date)
    pattern_dot = re.compile(
        r"^(?P<vendor>.+?)_(?P<year>\d{4})\.(?P<month>\d{1,2})\.(?P<day>\d{1,2})_(?P<amount>[\d,，]+)(?:_(?P<construction>.+))?$"
    )
    m = pattern_dot.match(stem)
    if m:
        vendor = m.group("vendor")
        year = m.group("year")
        month = m.group("month").zfill(2)
        day = m.group("day").zfill(2)
        date_str = f"{year}{month}{day}"
        project = _clean_project_name(m.group("construction"))
        amount_str = m.group("amount").replace(",", "").replace("，", "")

        result = _build_invoice_fields_from_pattern(vendor, date_str, amount_str, project, deny_re)
        if result is not None:
            return result

    # Pattern 5: YYYY,MM,DDvendoramount.pdf (legacy comma format)
    # Example: 2025,12,27名鉄協商400.pdf
    pattern_legacy = re.compile(
        r"^(?P<year>\d{4}),(?P<month>\d{1,2}),(?P<day>\d{1,2})(?P<vendor>[^\d]+)(?P<amount>[\d,，]+)$"
    )
    m = pattern_legacy.match(stem)
    if m:
        project = None
        vendor = m.group("vendor")
        year = m.group("year")
        month = m.group("month").zfill(2)
        day = m.group("day").zfill(2)
        date_str = f"{year}{month}{day}"
        amount_str = m.group("amount").replace(",", "").replace("，", "")

        result = _build_invoice_fields_from_pattern(vendor, date_str, amount_str, project, deny_re)
        if result is not None:
            return result

    return None


def _correct_vendor_ocr(vendor: str) -> str:
    """Apply known OCR misrecognition corrections to vendor name."""
    for wrong, correct in VENDOR_OCR_CORRECTIONS.items():
        if wrong in vendor:
            vendor = vendor.replace(wrong, correct)
    return vendor


def _extract_vendor_from_text(text: str, company_deny_regex: str | None) -> str | None:
    deny_re = re.compile(company_deny_regex) if company_deny_regex else None
    vendor_candidates: list[str] = []
    for line in [ln.strip() for ln in text.splitlines() if ln.strip()]:
        if len(vendor_candidates) >= 30:
            break
        if ("様" in line) or ("御中" in line):
            continue

        candidate = line
        if deny_re and deny_re.search(candidate):
            candidate = deny_re.sub(" ", candidate).strip()
            if not candidate:
                continue
            if re.fullmatch(r"[（(](?:株|有|合)[）)]|株式会社|有限会社|合同会社", candidate):
                continue

        candidate = re.sub(
            r"^(発行元|会社名|請求元|取引先|支払先|相手先|販売元|納入元)\s*[:：]?\s*",
            "",
            candidate,
        ).strip()

        if not VENDOR_HINT_RE.search(candidate):
            continue
        if "〒" in candidate:
            candidate = candidate[: candidate.index("〒")].strip()
        if not candidate:
            continue
        vendor_candidates.append(candidate)
    return vendor_candidates[0] if vendor_candidates else None


def _extract_issue_date_from_text(text: str) -> str | None:
    issue_date: str | None = None
    for m in DATE_LABEL_RE.finditer(text):
        parsed = _parse_ymd(m.group(2))
        if parsed:
            issue_date = parsed
            break
    if not issue_date:
        dates = []
        for m in DATE_YMD_RE.finditer(text):
            parsed = _parse_ymd(m.group(0))
            if parsed:
                dates.append(parsed)
        uniq = sorted(set(dates))
        if len(uniq) == 1:
            issue_date = uniq[0]

    if not issue_date:
        _WAREKI_LABEL_RE = re.compile(
            r"(?:請求(?:年月)?日|発行(?:年月)?日|作成(?:年月)?日|日付)\s*[:：]?\s*"
            r"(?:令和\s*|R\s*)?([0-9]{1,2})\s*年\s*([0-9]{1,2})\s*月\s*([0-9]{1,2})\s*日"
        )
        for wm in _WAREKI_LABEL_RE.finditer(text):
            era_y, mo, d = int(wm.group(1)), int(wm.group(2)), int(wm.group(3))
            if 1 <= era_y <= 20 and 1 <= mo <= 12 and 1 <= d <= 31:
                western_y = 2018 + era_y
                issue_date = f"{western_y:04}{mo:02}{d:02}"
                break

    if not issue_date:
        _WAREKI_RE = re.compile(
            r"(?:令和\s*|R\s*)?([0-9]{1,2})\s*年\s*([0-9]{1,2})\s*月\s*([0-9]{1,2})\s*日"
        )
        for wm in _WAREKI_RE.finditer(text):
            after = text[wm.end() : wm.end() + 10].lstrip()
            if after and after[0] in "~〜迄":
                continue
            era_y, mo, d = int(wm.group(1)), int(wm.group(2)), int(wm.group(3))
            if 1 <= era_y <= 20 and 1 <= mo <= 12 and 1 <= d <= 31:
                western_y = 2018 + era_y
                issue_date = f"{western_y:04}{mo:02}{d:02}"
                break

    if not issue_date:
        _WAREKI_SPACED_RE = re.compile(
            r"([0-9]{1,2})\s*年\s*([0-9])\s+([0-9])\s+([0-9])\s+([0-9])\s*日"
        )
        for wm in _WAREKI_SPACED_RE.finditer(text):
            era_y = int(wm.group(1))
            mmdd = wm.group(2) + wm.group(3) + wm.group(4) + wm.group(5)
            mo, d = int(mmdd[:2]), int(mmdd[2:])
            if 1 <= era_y <= 20 and 1 <= mo <= 12 and 1 <= d <= 31:
                western_y = 2018 + era_y
                issue_date = f"{western_y:04}{mo:02}{d:02}"
                break

    return issue_date


def _extract_amount_from_text(text: str) -> int | None:
    def _amount_label_priority(label: str) -> int:
        if "支払決定" in label:
            return 0
        if ("支払" in label) or ("お支払" in label) or ("お支払い" in label):
            return 1
        if "請求" in label:
            return 2
        return 3

    candidates: list[tuple[int, int]] = []
    for m in AMOUNT_LABEL_RE.finditer(text):
        label = m.group(1)
        raw = m.group(2).replace(",", "").strip()
        if raw.isdigit() and int(raw) > 0:
            candidates.append((_amount_label_priority(label), int(raw)))

    if candidates:
        candidates.sort(key=lambda x: (x[0], -x[1]))
        return candidates[0][1]

    amounts: list[int] = []
    for m in re.finditer(r"[¥￥]\s*([0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)", text):
        raw = m.group(1).replace(",", "")
        if raw.isdigit():
            amounts.append(int(raw))
    for m in re.finditer(r"([0-9]{1,3}(?:,[0-9]{3})+)\s*円", text):
        raw = m.group(1).replace(",", "")
        if raw.isdigit():
            amounts.append(int(raw))
    if not amounts:
        for m in re.finditer(r"([0-9]{1,3}(?:,[0-9]{3})+)", text):
            raw = m.group(1).replace(",", "")
            if raw.isdigit():
                v = int(raw)
                if v >= 1000:
                    amounts.append(v)
    return max(amounts) if amounts else None


def _extract_invoice_no_from_text(text: str) -> str | None:
    m_no = INVOICE_NO_RE.search(text)
    if m_no:
        return m_no.group(2).strip() or None
    return None


def _extract_invoice_fields(text: str, company_deny_regex: str | None) -> ExtractedPdfFields | None:
    text = _normalize_text(text)
    # Normalize spaced-out CJK characters (PDF text extraction artifact).
    # Example: "東 海 イ ン プ ル 建 設" → "東海インプル建設"
    text = re.sub(r"(?<=[^\x00-\x7f]) (?=[^\x00-\x7f])", "", text)
    # Fix common OCR reversals (e.g. some OCR engines reverse "株式会社" to "会社株式").
    text = text.replace("会社株式", "株式会社")
    # Normalize spaced commas inside numbers: "1, 651,980" → "1,651,980"
    text = re.sub(r"(\d),\s+(\d)", r"\1,\2", text)

    vendor = _extract_vendor_from_text(text, company_deny_regex)
    issue_date = _extract_issue_date_from_text(text)
    amount = _extract_amount_from_text(text)
    invoice_no = _extract_invoice_no_from_text(text)

    if not vendor or not issue_date or amount is None:
        return None

    # Apply OCR corrections to vendor name (pdf-ocr Skill).
    vendor = _correct_vendor_ocr(vendor)

    return ExtractedPdfFields(vendor=vendor, issue_date=issue_date, amount=amount, invoice_no=invoice_no, raw_ocr_text=text)


def _render_template(template: str, values: dict[str, str]) -> str:
    try:
        return template.format(**values)
    except KeyError as e:
        raise ValueError(f"rename template has unknown key: {e}") from e


def _load_bool_safe(raw: dict, key: str, default: bool = False) -> bool:
    """JSON config から bool を安全に読み込む。
    文字列 "false" を bool(False) に変換し、意図しないデフォルト True を防ぐ。
    """
    if key not in raw:
        return default
    value = raw[key]
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized == "true":
            return True
        if normalized == "false":
            return False
    raise ValueError(f"config.{key} must be bool or 'true'/'false' string, got: {value!r}")


def _move_file(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(src), str(dst))


def _load_project_master(path) -> list:
    """project_master.xlsx から ProjectMasterEntry リストを読み込む"""
    import openpyxl
    from datetime import date

    if not Path(path).exists():
        print(f"[WARN] project_master.xlsx が見つかりません: {path}")
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(row, name):
        try:
            idx = headers.index(name)
            v = row[idx].value
            return str(v).strip() if v is not None else ""
        except (ValueError, IndexError):
            return ""

    def split_keywords(raw_value: str) -> tuple[str, ...]:
        prepared = re.sub(r"(?<=\d),(?=\d)", "", str(raw_value or ""))
        prepared = prepared.replace("、", ",").replace(";", ",")
        keywords: list[str] = []
        for part in prepared.split(","):
            kw = part.strip()
            if not kw:
                continue
            norm = _normalize_for_match(kw)
            if len(norm) < 3:
                continue
            if norm.isdigit() and len(norm) < 4:
                continue
            if kw not in keywords:
                keywords.append(kw)
        return tuple(keywords)

    entries = []
    today = date.today()
    for row in ws.iter_rows(min_row=2):
        kojimei = col(row, "工事名称")
        if not kojimei:
            continue
        status = col(row, "status") or "active"
        if status != "active":
            continue
        end_date_str = col(row, "end_date")
        if end_date_str:
            try:
                end_date = date.fromisoformat(end_date_str)
                if end_date < today:
                    continue
            except ValueError:
                pass
        kw_raw = col(row, "キーワード")
        kw_list = split_keywords(kw_raw) if kw_raw else ()
        entries.append(ProjectMasterEntry(
            kojiban=col(row, "工事番号"),
            kojimei=kojimei,
            busho=col(row, "部署") or "新築",
            keywords=kw_list,
            start_date=col(row, "start_date"),
            end_date=end_date_str,
            status=status,
        ))

    wb.close()
    print(f"[INFO] project_master.xlsx から {len(entries)} 件のルーティングルールを読み込みました: {path}")
    return entries


def _resolve_project_master_path(config_dir: Path) -> Path:
    versioned = sorted(config_dir.glob("project_master_v*.xlsx"))
    if versioned:
        return versioned[-1]
    fallback = config_dir / "project_master.xlsx"
    return fallback


def _load_web_download_config(raw_section: dict | None) -> Any:
    """Parse web_download config section into WebDownloadConfig or None."""
    if not raw_section:
        return None
    if not _WEB_DOWNLOAD_AVAILABLE:
        return None
    vendors = []
    for v in raw_section.get("vendors", []):
        vendors.append(VendorWebConfig(
            handler_id=str(v.get("handler_id", "")),
            target_name=str(v.get("target_name", "")),
            sender_pattern=str(v.get("sender_pattern", "")),
            url_pattern=str(v.get("url_pattern", "")),
            options=dict(v.get("options", {})),
        ))
    return WebDownloadConfig(
        enabled=bool(raw_section.get("enabled", False)),
        headless=bool(raw_section.get("headless", True)),
        download_timeout_ms=int(raw_section.get("download_timeout_ms", 30000)),
        vendors=tuple(vendors),
    )


def _resolve_and_load_project_master(
    config_path: Path, project_master_override: Path | None
) -> tuple[list, Path]:
    project_master_path = project_master_override or _resolve_project_master_path(config_path.parent)
    if project_master_override is not None:
        print(f"[INFO] project_master override を使用します: {project_master_path}")
    pm_entries = _load_project_master(project_master_path)
    if not pm_entries and project_master_path.exists():
        print(f"[WARN] project_master.xlsx は存在しますが有効なルールが0件です: {project_master_path}")
    return pm_entries, project_master_path


def _apply_fallback_subdir_timestamp(routing: RoutingConfig) -> RoutingConfig:
    if not routing.fallback_subdir:
        return routing
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return RoutingConfig(
        enabled=routing.enabled,
        fallback_subdir=f"{routing.fallback_subdir}_{ts}",
        rules=routing.rules,
        sender_rules=routing.sender_rules,
        ban_senders=routing.ban_senders,
        threshold=routing.threshold,
        tie_margin=routing.tie_margin,
        vision_ocr_enabled=routing.vision_ocr_enabled,
        vision_ocr_provider=routing.vision_ocr_provider,
    )


def _parse_keywords(raw: Any) -> list[str]:
    """Parse keywords value (list[str] | str | None) → list of non-empty stripped strings."""
    if isinstance(raw, list):
        return [str(kw).strip() for kw in raw if str(kw).strip()]
    if isinstance(raw, str):
        s = raw.strip()
        return [s] if s else []
    return []


def _load_config(path: Path, project_master_override: Path | None = None) -> ToolConfig:
    # PowerShell's Set-Content can emit UTF-8 BOM; accept it.
    raw = json.loads(path.read_text(encoding="utf-8-sig"))
    outlook_raw = raw.get("outlook") if isinstance(raw, dict) else None
    outlook = None
    if isinstance(outlook_raw, dict):
        outlook = OutlookSelection(
            folder_path=str(outlook_raw.get("folder_path", "")).strip(),
            profile_name=(
                str(outlook_raw["profile_name"]).strip()
                if outlook_raw.get("profile_name") is not None
                else None
            ),
            unread_only=bool(outlook_raw.get("unread_only", True)),
            max_messages=int(outlook_raw.get("max_messages", 200)),
            received_within_days=(
                int(outlook_raw["received_within_days"])
                if outlook_raw.get("received_within_days") is not None
                else None
            ),
            subject_allow_regex=(
                str(outlook_raw["subject_allow_regex"]).strip()
                if outlook_raw.get("subject_allow_regex")
                else None
            ),
            subject_deny_regex=(
                str(outlook_raw["subject_deny_regex"]).strip()
                if outlook_raw.get("subject_deny_regex")
                else None
            ),
            mark_as_read_on_success=bool(outlook_raw.get("mark_as_read_on_success", False)),
            use_flag_status=bool(outlook_raw.get("use_flag_status", False)),
            mark_flag_on_success=bool(outlook_raw.get("mark_flag_on_success", False)),
        )
        if not outlook.folder_path:
            raise ValueError("config.outlook.folder_path が空です。")

    merge_raw = raw.get("merge") if isinstance(raw, dict) else None
    merge = MergeConfig()
    if isinstance(merge_raw, dict):
        merge = MergeConfig(
            enabled=bool(merge_raw.get("enabled", True)),
            output_name=str(merge_raw.get("output_name", "merged.pdf")).strip()
            or "merged.pdf",
        )

    print_raw = raw.get("print") if isinstance(raw, dict) else None
    prn = PrintConfig()
    if isinstance(print_raw, dict):
        prn = PrintConfig(
            enabled=bool(print_raw.get("enabled", False)),
            printer_name=(
                str(print_raw["printer_name"]).strip()
                if print_raw.get("printer_name")
                else None
            ),
            method=(str(print_raw.get("method", "shell")).strip() or "shell").lower(),
            pdf_output_dir=(
                str(print_raw["pdf_output_dir"]).strip()
                if print_raw.get("pdf_output_dir")
                else None
            ),
        )

    mail_raw = raw.get("mail") if isinstance(raw, dict) else None
    mail = MailConfig()
    if isinstance(mail_raw, dict):
        mail = MailConfig(
            send_success=bool(mail_raw.get("send_success", True)),
            success_to=tuple(mail_raw.get("success_to") or ()),
            error_to=tuple(mail_raw.get("error_to") or ()),
            error_cc=tuple(mail_raw.get("error_cc") or ()),
        )

    rename_raw = raw.get("rename") if isinstance(raw, dict) else None
    ren = RenameConfig()
    if isinstance(rename_raw, dict):
        ren = RenameConfig(
            enabled=bool(rename_raw.get("enabled", True)),
            create_vendor_subdir=bool(rename_raw.get("create_vendor_subdir", True)),
            vendor_subdir_template=str(
                rename_raw.get("vendor_subdir_template", "{vendor_short}")
            ).strip()
            or "{vendor_short}",
            file_name_template=str(
                rename_raw.get(
                    "file_name_template", "{vendor}_{issue_date}_{amount_comma}_{project}.pdf"
                )
            ).strip()
            or "{vendor}_{issue_date}_{amount_comma}_{project}.pdf",
            company_deny_regex=(
                str(rename_raw["company_deny_regex"]).strip()
                if rename_raw.get("company_deny_regex")
                else None
            ),
            max_pages=int(rename_raw.get("max_pages", 2)),
            filename_first=bool(rename_raw.get("filename_first", True)),
        )

    routing_raw = raw.get("routing") if isinstance(raw, dict) else None
    routing = RoutingConfig()
    if isinstance(routing_raw, dict):
        parsed_rules: list[RoutingRule] = []
        rules_raw = routing_raw.get("rules")
        if isinstance(rules_raw, list):
            for rule_raw in rules_raw:
                if not isinstance(rule_raw, dict):
                    continue
                subdir = str(rule_raw.get("subdir", "")).strip()
                if not subdir:
                    continue
                kws: list[str] = _parse_keywords(rule_raw.get("keywords"))
                parsed_rules.append(RoutingRule(subdir=subdir, keywords=tuple(kws)))
        elif isinstance(rules_raw, dict):
            for subdir_raw, keywords_raw in rules_raw.items():
                subdir = str(subdir_raw).strip()
                if not subdir:
                    continue
                kws: list[str] = _parse_keywords(keywords_raw)
                parsed_rules.append(RoutingRule(subdir=subdir, keywords=tuple(kws)))

        # threshold: config指定があればそれを使う。なければデフォルト1
        cfg_threshold = routing_raw.get("threshold")
        threshold_val = int(cfg_threshold) if cfg_threshold is not None else 1
        # tie_margin: config指定があればfloat。なければNone（タイ判定しない）
        cfg_tie_margin = routing_raw.get("tie_margin")
        tie_margin_val = float(cfg_tie_margin) if cfg_tie_margin is not None else None

        # sender_rules: L1 sender-based routing rules
        parsed_sender_rules: list[SenderRule] = []
        sr_raw = routing_raw.get("sender_rules")
        if isinstance(sr_raw, list):
            for sr in sr_raw:
                if not isinstance(sr, dict):
                    continue
                s_sender = str(sr.get("sender", "")).lower().strip()
                s_subdir = str(sr.get("subdir", "")).strip()
                if not s_sender or not s_subdir:
                    continue
                s_kws: list[str] = _parse_keywords(sr.get("keywords"))
                parsed_sender_rules.append(SenderRule(sender=s_sender, subdir=s_subdir, keywords=tuple(s_kws)))

        # ban_senders: senders always routed to fallback
        ban_raw = routing_raw.get("ban_senders")
        parsed_ban: set[str] = set()
        if isinstance(ban_raw, list):
            for b in ban_raw:
                b_s = str(b).lower().strip()
                if b_s:
                    parsed_ban.add(b_s)

        routing = RoutingConfig(
            enabled=bool(routing_raw.get("enabled", False)),
            fallback_subdir=str(routing_raw.get("fallback_subdir", "")).strip(),
            rules=tuple(parsed_rules),
            sender_rules=tuple(parsed_sender_rules),
            ban_senders=frozenset(parsed_ban),
            threshold=threshold_val,
            tie_margin=tie_margin_val,
            vision_ocr_enabled=bool(routing_raw.get("vision_ocr_enabled", False)),
            vision_ocr_provider=str(routing_raw.get("vision_ocr_provider", "openai")),
        )

    pm_entries, project_master_path = _resolve_and_load_project_master(path, project_master_override)
    routing = _apply_fallback_subdir_timestamp(routing)

    cfg = ToolConfig(
        artifact_dir=str(raw.get("artifact_dir", DEFAULT_ARTIFACT_DIR)).strip()
        or DEFAULT_ARTIFACT_DIR,
        save_dir=(str(raw["save_dir"]).strip() if raw.get("save_dir") else None),
        outlook=outlook,
        merge=merge,
        print=prn,
        mail=mail,
        rename=ren,
        routing=routing,
        fail_on_url_only_mail=_load_bool_safe(raw, "fail_on_url_only_mail", default=False),
        enable_zip_processing=_load_bool_safe(raw, "enable_zip_processing", default=False),
        enable_pdf_decryption=_load_bool_safe(raw, "enable_pdf_decryption", default=False),
        project_master=tuple(pm_entries),
        project_master_path=str(project_master_path),
        enable_web_download=_load_bool_safe(raw, "enable_web_download", default=False),
        web_download=_load_web_download_config(raw.get("web_download")),
    )

    return cfg


def _is_unc_path(path: Path) -> bool:
    return str(path).startswith("\\\\")


def _save_dir_bracket_fallback_candidates(path: Path) -> list[Path]:
    """
    Build fallback candidates for common bracket variants:
    - 全角: 【】
    - 半角: []
    """

    raw = str(path)
    if ("【" in raw) or ("】" in raw):
        alt = raw.replace("【", "[").replace("】", "]")
    elif ("[" in raw) or ("]" in raw):
        alt = raw.replace("[", "【").replace("]", "】")
    else:
        return []
    if alt != raw:
        return [Path(alt)]
    return []


def _resolve_save_dir_with_fallback(path: Path, *, timeout_seconds: float, log_path: Path) -> Path:
    """
    If the configured save_dir is unreachable, try bracket-style fallbacks
    once and use the first reachable candidate.
    """

    primary = _test_path_with_timeout(path, timeout_seconds=timeout_seconds)
    if primary is True:
        return path

    for cand in _save_dir_bracket_fallback_candidates(path):
        exists = _test_path_with_timeout(cand, timeout_seconds=timeout_seconds)
        if exists is True:
            _append_log(log_path, f"[WARN] save_dir fallback used: {path} -> {cand}")
            return cand
    return path


def _ps_single_quote(value: str) -> str:
    # Escape single quotes for PowerShell single-quoted string literal.
    return "'" + value.replace("'", "''") + "'"


def _test_path_with_timeout(path: Path, timeout_seconds: float) -> bool | None:
    """
    Return True/False if the existence check succeeded, or None if it timed out/failed.
    Use PowerShell Test-Path for UNC paths to avoid long hangs.
    """

    if not _is_unc_path(path):
        try:
            return path.exists()
        except Exception:
            return None

    cmd = f"Test-Path -LiteralPath {_ps_single_quote(str(path))}"
    try:
        cp = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", cmd],
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
        )
    except subprocess.TimeoutExpired:
        return None
    except Exception:
        return None

    if cp.returncode != 0:
        return None
    out = (cp.stdout or "").strip().lower()
    if out == "true":
        return True
    if out == "false":
        return False
    return None


def _outlook_namespace(profile_name: str | None = None) -> Any:
    if not _OUTLOOK_AVAILABLE:
        raise RuntimeError("Outlook COM が利用できません (pywin32/win32com が必要です)。")
    _install_com_message_filter()
    is_new_instance = False
    try:
        # 既存の Outlook セッションを優先して接続（ログイン済み・正しいプロファイル）
        outlook = win32.GetActiveObject("Outlook.Application")
    except Exception:
        # Outlook が起動していない場合は新規起動
        outlook = win32.Dispatch("Outlook.Application")
        is_new_instance = True
    mapi = _com_retry(lambda: outlook.GetNamespace("MAPI"))
    if profile_name:
        profile = profile_name.strip()
        if profile:
            try:
                # ShowDialog=False to avoid UI; NewSession=True to ensure the profile is loaded.
                _com_retry(lambda: mapi.Logon(profile, "", False, True))
            except Exception as e:
                raise RuntimeError(
                    "Outlook プロファイルにログオンできませんでした。\n"
                    f"profile_name={profile_name}"
                ) from e
    elif is_new_instance:
        # 新規起動時: profile_name 未指定でもデフォルトプロファイルを確実に読み込む
        try:
            _com_retry(lambda: mapi.Logon("", "", False, True))
        except Exception:
            pass  # すでに Logon 済みの場合は無視
    return mapi


def _resolve_outlook_folder(mapi: Any, folder_path: str) -> Any:
    # folder_path example: "\\RK用の経理\\受信トレイ\\サブフォルダ"
    s = folder_path.strip()
    if not s.startswith("\\\\"):
        raise ValueError(
            f"outlook.folder_path は \\\\ から始まる必要があります: {folder_path}"
        )
    parts = [p for p in s.strip("\\").split("\\") if p]
    if len(parts) < 1:
        raise ValueError(f"outlook.folder_path の形式が不正です: {folder_path}")

    store_name = parts[0]
    store = None
    stores = _com_retry(lambda: mapi.Stores)
    count = _com_retry(lambda: int(stores.Count))
    for i in range(1, count + 1):
        s = _com_retry(lambda i=i: stores.Item(i))
        display_name = (s.DisplayName or "").strip()
        if display_name == store_name:
            store = _com_retry(lambda s=s: s.GetRootFolder())
            break
    if store is None:
        available = []
        for i in range(1, count + 1):
            try:
                available.append(str(_com_retry(lambda i=i: stores.Item(i).DisplayName)))
            except Exception:
                continue
        raise ValueError(
            "Outlookストアが見つかりません。\n"
            f"指定: {store_name}\n"
            f"利用可能: {available}"
        )

    if len(parts) == 1:
        return store

    cur = store
    for name in parts[1:]:
        try:
            cur = _com_retry(lambda name=name: cur.Folders.Item(name))
        except Exception as e:
            raise ValueError(
                f"Outlookフォルダが見つかりません: {folder_path}\n"
                f"見つからない要素: {name}"
            ) from e
    return cur


def _safe_sender(item: Any) -> str:
    # SenderEmailAddress is often an Exchange X500 string. Prefer SenderName when needed.
    try:
        addr = (item.SenderEmailAddress or "").strip()
    except Exception:
        addr = ""
    try:
        name = (item.SenderName or "").strip()
    except Exception:
        name = ""
    if addr and "@" in addr:
        return addr
    return name or addr or "(unknown)"


def _safe_received_time(item: Any) -> datetime | None:
    try:
        rt = item.ReceivedTime
        if isinstance(rt, datetime):
            if rt.tzinfo is not None:
                # Normalize to naive local time to avoid mixing aware/naive datetimes.
                return rt.astimezone().replace(tzinfo=None)
            return rt
    except Exception:
        return None
    return None


def _iter_mail_items(folder: Any, unread_only: bool, max_messages: int) -> list[Any]:
    items = _com_retry(lambda: folder.Items)
    try:
        _com_retry(lambda: items.Sort("[ReceivedTime]", True), retries=5)
    except Exception:
        pass

    res: list[Any] = []
    count = _com_retry(lambda: int(items.Count))
    for i in range(1, count + 1):
        if len(res) >= max_messages:
            break
        try:
            item = _com_retry(lambda i=i: items.Item(i), retries=10)
        except Exception:
            continue
        # MailItem Class=43
        try:
            if int(getattr(item, "Class", 0)) != 43:
                continue
        except Exception:
            continue
        if unread_only:
            try:
                if not bool(item.UnRead):
                    continue
            except Exception:
                continue
        res.append(item)
    return res


def _is_password_mail(subject: str) -> bool:
    s = subject.lower()
    return ("パスワード" in subject) or ("hennge" in s) or ("password" in s)


def _clean_password_token(token: str) -> str:
    t = _normalize_text(token or "").strip()
    # Strip common wrappers (quotes/brackets/punctuation) but keep symbols like '+' that are valid in passwords.
    t = t.strip(" \t\r\n\"'“”‘’()[]{}<>「」『』")
    t = t.strip(" \t\r\n,.;:：。．、")
    return t


def _is_plausible_password(token: str) -> bool:
    t = token.strip()
    if len(t) < 4:
        return False
    # Avoid obvious masks
    if set(t) <= {"*", "・", "●"}:
        return False
    # Require at least one alnum (PPAP passwords are typically mixed)
    if not re.search(r"[A-Za-z0-9]", t):
        return False
    return True


def _extract_password(body: str) -> str | None:
    body = _normalize_text(body or "")
    m = PASSWORD_RE.search(body)
    if m:
        cand = _clean_password_token(m.group(1))
        return cand if _is_plausible_password(cand) else None
    # fallback: find a plausible token on a line containing パスワード
    for line in body.splitlines():
        if "パスワード" not in line and "password" not in line.lower():
            continue
        tokens = re.findall(r"[^\s]{4,}", line)
        for tok in reversed(tokens):
            cand = _clean_password_token(tok)
            if _is_plausible_password(cand):
                return cand
    return None


def _extract_urls(text: str) -> list[str]:
    urls = []
    for m in URL_RE.finditer(text or ""):
        u = m.group(0).strip().rstrip(").,>")
        if u not in urls:
            urls.append(u)
    return urls


def _subject_allowed(subject: str, allow_re: re.Pattern[str] | None) -> bool:
    if allow_re is None:
        return True
    return bool(allow_re.search(subject))


def _subject_denied(subject: str, deny_re: re.Pattern[str] | None) -> bool:
    if deny_re is None:
        return False
    return bool(deny_re.search(subject))


def _unique_path(base: Path) -> Path:
    if not base.exists():
        return base
    stem = base.stem
    suffix = base.suffix
    for i in range(1, 999):
        cand = base.with_name(f"{stem}__dup{i:03d}{suffix}")
        if not cand.exists():
            return cand
    raise RuntimeError(f"同名ファイルが多すぎます: {base}")


def _is_pdf_encrypted(path: Path) -> bool:
    if not _PYPDF_AVAILABLE:
        return False
    try:
        r = PdfReader(str(path))
        return bool(getattr(r, "is_encrypted", False))
    except Exception:
        # treat as non-encrypted here; parse errors will be handled later
        return False


def _decrypt_pdf(src: Path, dst: Path, password: str) -> None:
    if not _PYPDF_AVAILABLE:
        raise RuntimeError("pypdf が利用できません。")
    reader = PdfReader(str(src))
    if not reader.is_encrypted:
        raise ValueError("PDFは暗号化されていません。")
    ok = reader.decrypt(password)
    if ok == 0:
        raise ValueError("PDF復号に失敗しました（パスワード不一致）。")

    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    with dst.open("wb") as f:
        writer.write(f)


def _extract_email_address(email_str: str) -> str:
    """Extract email address from various formats.

    Handles:
    - Plain: user@example.com
    - Display name: Name <user@example.com>
    - Edge cases: trailing/leading spaces, quotes

    Returns lowercase email or empty string if no @ found.
    """
    text = (email_str or "").strip()
    # Remove quotes
    text = text.strip('"\'')

    # Extract from "Name <email>" format
    if "<" in text and ">" in text:
        start = text.find("<")
        end = text.find(">", start)
        text = text[start + 1:end].strip()

    text = text.lower()
    return text if "@" in text else ""


def _extract_email_domain(email_str: str) -> str:
    """Extract domain from email address (e.g., 'user@example.com' -> 'example.com')"""
    addr = _extract_email_address(email_str)
    if not addr or "@" not in addr:
        return ""
    return addr.split("@", 1)[-1]


def _candidate_passwords_with_fallback(
    target_time: datetime,
    target_sender: str,
    notes: Sequence[PasswordNote],
    log_path: Path,
) -> list[str]:
    """Find password candidates using staged fallback strategy.

    Stage 1: Strict - 30 sec + exact sender match
    Stage 2: Relaxed - 3 min + domain match
    Stage 3: Last resort - 10 min + time only (max 5 candidates)

    Returns list of password candidates (empty if none found).
    """
    target_addr = _extract_email_address(target_sender)
    target_domain = _extract_email_domain(target_sender)

    # Stage 1: 30 seconds + exact sender match
    stage1_window = timedelta(seconds=30)
    stage1_candidates: list[tuple[float, str]] = []
    for n in notes:
        note_addr = _extract_email_address(n.sender)
        if target_addr and note_addr and target_addr == note_addr:
            dt = abs((n.received_time - target_time).total_seconds())
            if dt <= stage1_window.total_seconds():
                stage1_candidates.append((dt, n.password))

    if stage1_candidates:
        _append_log(log_path, f"[Password] Stage1 match: exact sender, {len(stage1_candidates)} candidates")
        return _deduplicate_passwords(stage1_candidates)

    # Stage 2: 3 minutes + domain match
    stage2_window = timedelta(minutes=3)
    stage2_candidates: list[tuple[float, str]] = []
    for n in notes:
        note_domain = _extract_email_domain(n.sender)
        if target_domain and note_domain and target_domain == note_domain:
            dt = abs((n.received_time - target_time).total_seconds())
            if dt <= stage2_window.total_seconds():
                stage2_candidates.append((dt, n.password))

    if stage2_candidates:
        _append_log(log_path, f"[Password] Stage2 match: domain only, {len(stage2_candidates)} candidates")
        return _deduplicate_passwords(stage2_candidates)

    # Stage 3: 10 minutes + time only (last resort)
    stage3_window = timedelta(minutes=10)
    stage3_candidates: list[tuple[float, str]] = []
    for n in notes:
        dt = abs((n.received_time - target_time).total_seconds())
        if dt <= stage3_window.total_seconds():
            stage3_candidates.append((dt, n.password))

    if stage3_candidates:
        _append_log(log_path, f"[Password] Stage3 match: time only, {len(stage3_candidates)} candidates")
        return _deduplicate_passwords(stage3_candidates)

    _append_log(log_path, "[Password] No candidates found in all stages")
    return []


def _deduplicate_passwords(scored: list[tuple[float, str]]) -> list[str]:
    """Deduplicate and return passwords sorted by time delta."""
    scored.sort(key=lambda x: (x[0], x[1]))
    uniq: list[str] = []
    seen: set[str] = set()
    for _, pw in scored:
        if not pw:
            continue
        if pw in seen:
            continue
        uniq.append(pw)
        seen.add(pw)
        if len(uniq) >= ZIP_PASSWORD_MAX_CANDIDATES:
            break
    return uniq


def _unique_dir(base: Path) -> Path:
    if not base.exists():
        return base
    for i in range(1, 999):
        cand = base.with_name(f"{base.name}__dup{i:03d}")
        if not cand.exists():
            return cand
    raise RuntimeError(f"同名ディレクトリが多すぎます: {base}")


def _sanitize_zip_member_relpath(member_name: str) -> Path:
    # Zip entries are usually POSIX-style paths even on Windows.
    raw = member_name.replace("\\", "/").strip()
    rel = Path(raw)
    if rel.is_absolute() or (len(rel.parts) > 0 and rel.parts[0].endswith(":")):
        raise ValueError(f"zip member has absolute path: {member_name}")
    parts: list[str] = []
    for p in rel.parts:
        if p in ("", "."):
            continue
        if p == "..":
            raise ValueError(f"zip member has parent traversal: {member_name}")
        parts.append(_sanitize_filename(p))
    if not parts:
        raise ValueError(f"zip member path is empty: {member_name}")
    return Path(*parts)


def _is_wrong_zip_password_error(exc: Exception) -> bool:
    msg = str(exc).lower()
    return ("password" in msg) or ("crc" in msg)


def _extract_pdf_members_from_zip(
    zip_path: Path,
    *,
    password: str | None,
    out_dir: Path,
    log_path: Path,
) -> list[Path]:
    pwd_bytes = password.encode("utf-8") if password else None
    extracted: list[Path] = []

    def _extract_with_open(zf: Any) -> list[Path]:
        infos = [i for i in zf.infolist() if not getattr(i, "is_dir", lambda: False)()]
        pdf_infos = [i for i in infos if str(i.filename).lower().endswith(".pdf")]
        if not pdf_infos:
            return []
        if len(pdf_infos) > ZIP_MAX_MEMBERS:
            raise ValueError(
                f"zip内PDFが多すぎます: {zip_path.name} pdf_members={len(pdf_infos)}"
            )
        total = 0
        for info in pdf_infos:
            size = int(getattr(info, "file_size", 0))
            if size <= 0:
                continue
            if size > ZIP_MAX_MEMBER_UNCOMPRESSED_BYTES:
                raise ValueError(
                    f"zip内PDFが大きすぎます: {zip_path.name} member={info.filename} size={size}"
                )
            total += size
            if total > ZIP_MAX_TOTAL_UNCOMPRESSED_BYTES:
                raise ValueError(
                    f"zip展開サイズが上限超過: {zip_path.name} total={total} limit={ZIP_MAX_TOTAL_UNCOMPRESSED_BYTES}"
                )

        for info in pdf_infos:
            rel = _sanitize_zip_member_relpath(str(info.filename))
            dest = out_dir / rel
            dest = _unique_path(dest)
            _ensure_dir(dest.parent)
            _append_log(log_path, f"extract zip member: {zip_path.name}::{info.filename} -> {dest}")
            # zipfile/pyzipper both accept pwd=...
            with zf.open(info, pwd=pwd_bytes) as src, dest.open("wb") as dst:
                shutil.copyfileobj(src, dst)
            extracted.append(dest)
        return extracted

    # Prefer pyzipper (AES対応) when available; fallback to stdlib zipfile.
    if _PYZIPPER_AVAILABLE:
        try:
            with pyzipper.AESZipFile(str(zip_path), "r") as zf:
                return _extract_with_open(zf)
        except Exception as e:
            # If pyzipper fails, fallback to stdlib as a second attempt for non-AES zips.
            _append_log(log_path, f"[WARN] pyzipper failed, fallback to zipfile: {zip_path.name} error={e}")

    with zipfile.ZipFile(str(zip_path), "r") as zf:
        return _extract_with_open(zf)


def _merge_pdfs(input_paths: Sequence[Path], output_path: Path) -> None:
    if not _PYPDF_AVAILABLE:
        raise RuntimeError("pypdf が利用できません。")
    writer = PdfWriter()
    for p in input_paths:
        reader = PdfReader(str(p))
        for page in reader.pages:
            writer.add_page(page)
    with output_path.open("wb") as f:
        writer.write(f)


def _default_printer() -> str | None:
    if not _WIN32_PRINT_AVAILABLE:
        return None
    try:
        return str(win32print.GetDefaultPrinter())
    except Exception:
        return None


def _print_via_shell(pdf_path: Path, printer_name: str | None) -> None:
    if not _WIN32_PRINT_AVAILABLE:
        raise RuntimeError("印刷には pywin32 (win32api/win32print) が必要です。")
    verb = "printto" if printer_name else "print"
    params = f"\"{printer_name}\"" if printer_name else None
    win32api.ShellExecute(0, verb, str(pdf_path), params, str(pdf_path.parent), 0)


def _print_to_pdf_copy(pdf_path: Path, output_dir: Path) -> Path:
    """
    MVP print mode:
    do not send to a physical printer; copy printable PDFs to output_dir.
    """
    _ensure_dir(output_dir)
    dst = _unique_path(output_dir / _sanitize_filename(pdf_path.name))
    shutil.copy2(pdf_path, dst)
    return dst


def _write_report_json(path: Path, report: RunReport) -> None:
    path.write_text(
        json.dumps(asdict(report), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _write_report_csv(path: Path, saved: Sequence[SavedAttachment]) -> None:
    fields = [
        "message_entry_id",
        "message_subject",
        "sender",
        "received_time",
        "attachment_name",
        "saved_path",
        "original_saved_path",
        "vendor",
        "issue_date",
        "amount",
        "invoice_no",
        "project",
        "sha256",
        "was_encrypted",
        "decrypted_path",
        "encrypted_original_path",
        "route_subdir",
        "route_reason",
        "processing_status",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in saved:
            w.writerow(asdict(r))


def _write_new_project_candidates_csv(path: Path, saved: Sequence[SavedAttachment]) -> bool:
    rows = [r for r in saved if r.route_reason == "new_project_candidate"]
    if not rows:
        return False
    fields = [
        "message_subject",
        "sender",
        "received_time",
        "attachment_name",
        "vendor",
        "issue_date",
        "amount",
        "project",
        "saved_path",
        "route_subdir",
        "route_reason",
        "processing_status",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow({k: asdict(r).get(k) for k in fields})
    return True


def _route_and_move_file(
    primary_path: Path,
    dest_dir: Path,
    final_name: str,
    log_path: Path,
    unresolved: list[str],
    log_prefix: str = "rename/move",
) -> Path:
    _ensure_dir(dest_dir)
    final_path = _unique_path(dest_dir / final_name)
    _append_log(log_path, f"{log_prefix}: {primary_path.name} -> {final_path}")
    try:
        _move_file(primary_path, final_path)
    except Exception:
        unresolved.append(f"failed to move file: {primary_path.name} -> {final_path}")
        final_path = primary_path
    return final_path


def _check_duplicate_attachment(
    attachment_name: str,
    sha: str,
    processed_index: dict,
    entry_id: str,
    subject: str,
    sender: str,
    received_time_s: str,
    pdf_path: Path,
    save_dir: Path,
    log_path: Path,
) -> "SavedAttachment | None":
    manifest_key = _processed_attachment_key(entry_id, attachment_name, sha)
    existing_record = processed_index.get(manifest_key)
    if not existing_record:
        return None
    existing_saved_path = str(existing_record.get("saved_path") or "").strip()
    existing_route_reason = str(existing_record.get("route_reason") or "").strip()
    existing_name = Path(existing_saved_path).name if existing_saved_path else ""
    can_skip_duplicate = (
        bool(existing_saved_path)
        and Path(existing_saved_path).exists()
        and _is_within_dir(Path(existing_saved_path), save_dir)
        and not existing_name.startswith("要確認_")
        and existing_route_reason not in {"fallback", "new_project_candidate"}
    )
    if not can_skip_duplicate:
        return None
    try:
        if pdf_path.exists():
            pdf_path.unlink()
    except Exception:
        pass
    _append_log(
        log_path,
        f"[SKIP] duplicate attachment: {attachment_name} -> {existing_saved_path}",
    )
    return SavedAttachment(
        message_entry_id=entry_id,
        message_subject=subject,
        sender=sender,
        received_time=received_time_s,
        attachment_name=attachment_name,
        saved_path=existing_saved_path,
        original_saved_path=str(pdf_path),
        vendor=existing_record.get("vendor"),
        issue_date=existing_record.get("issue_date"),
        amount=existing_record.get("amount"),
        invoice_no=existing_record.get("invoice_no"),
        project=existing_record.get("project"),
        sha256=sha,
        was_encrypted=bool(existing_record.get("was_encrypted", False)),
        decrypted_path=existing_record.get("decrypted_path"),
        encrypted_original_path=existing_record.get("encrypted_original_path"),
        route_subdir=existing_record.get("route_subdir"),
        route_reason=existing_record.get("route_reason") or "idempotency_manifest",
        processing_status="duplicate_skipped",
    )


def _decrypt_pdf_if_needed(
    pdf_path: Path,
    was_encrypted: bool,
    cfg: "ToolConfig",
    received_dt: datetime,
    sender: str,
    password_notes: "Sequence[PasswordNote]",
    log_path: Path,
    unresolved: list[str],
    subject: str = "",
) -> "tuple[Path, Path | None]":
    if not was_encrypted:
        return pdf_path, None
    if not cfg.enable_pdf_decryption:
        _append_log(
            log_path,
            f"[MVP-SKIP] encrypted pdf decryption is disabled: {pdf_path.name}",
        )
        return pdf_path, None
    passwords = _candidate_passwords_with_fallback(
        target_time=received_dt,
        target_sender=sender,
        notes=password_notes,
        log_path=log_path,
    )
    if not passwords:
        unresolved.append(
            f"encrypted pdf password not found: {pdf_path.name} (subject={subject})"
        )
        return pdf_path, None
    out_base = pdf_path.with_name(pdf_path.stem + "__decrypted.pdf")
    decrypted_path: Path | None = None
    for pw in passwords:
        out = _unique_path(out_base)
        try:
            _append_log(log_path, f"decrypt pdf: {pdf_path.name} -> {out.name}")
            _decrypt_pdf(pdf_path, out, pw)
            decrypted_path = out
            break
        except Exception as e:
            try:
                out.unlink()
            except Exception:
                pass
            _append_log(
                log_path,
                f"[WARN] decrypt failed (try next): {pdf_path.name} error={e}",
            )
    if decrypted_path is None:
        unresolved.append(
            f"failed to decrypt encrypted pdf: {pdf_path.name} (subject={subject})"
        )
        return pdf_path, None
    return decrypted_path, decrypted_path


def _process_saved_pdf_file(
    *,
    cfg: ToolConfig,
    pdf_path: Path,
    attachment_name: str,
    entry_id: str,
    subject: str,
    sender: str,
    body_snippet: str = "",
    received_dt: datetime,
    received_time_s: str,
    password_notes: Sequence[PasswordNote],
    save_dir: Path,
    run_dir: Path,
    log_path: Path,
    unresolved: list[str],
    processed_index: dict[str, dict[str, Any]],
    processed_manifest_path: Path,
) -> SavedAttachment:
    sha = _sha256_file(pdf_path)
    manifest_key = _processed_attachment_key(entry_id, attachment_name, sha)

    duplicate = _check_duplicate_attachment(
        attachment_name=attachment_name,
        sha=sha,
        processed_index=processed_index,
        entry_id=entry_id,
        subject=subject,
        sender=sender,
        received_time_s=received_time_s,
        pdf_path=pdf_path,
        save_dir=save_dir,
        log_path=log_path,
    )
    if duplicate is not None:
        return duplicate

    was_encrypted = _is_pdf_encrypted(pdf_path)
    encrypted_original_path: str | None = str(pdf_path) if was_encrypted else None

    primary_path, decrypted_path = _decrypt_pdf_if_needed(
        pdf_path=pdf_path,
        was_encrypted=was_encrypted,
        cfg=cfg,
        received_dt=received_dt,
        sender=sender,
        password_notes=password_notes,
        log_path=log_path,
        unresolved=unresolved,
        subject=subject,
    )

    final_path = primary_path
    vendor: str | None = None
    issue_date: str | None = None
    amount: int | None = None
    invoice_no: str | None = None
    project: str | None = None
    route_subdir: str | None = None
    route_reason: str | None = None

    partial_fields = _extract_partial_invoice_fields_from_filename(
        attachment_name, cfg.rename.company_deny_regex
    )
    review_partial = partial_fields
    review_reasons: tuple[str, ...] = ()
    review_raw_ocr_text = ""

    if cfg.rename.enabled:
        fields = None

        # 1) filename-first (fast and stable)
        if cfg.rename.filename_first:
            fields = _extract_invoice_fields_from_filename(
                attachment_name,
                company_deny_regex=cfg.rename.company_deny_regex,
            )
            if fields is not None:
                _append_log(
                    log_path,
                    f"[INFO] extracted from filename: vendor={fields.vendor}, date={fields.issue_date}, amount={fields.amount}",
                )

        # 2) fallback to PDF text extraction (only when decrypt succeeded if encrypted)
        can_read_pdf = (not was_encrypted) or (decrypted_path is not None)
        if fields is None and can_read_pdf:
            fields = _extract_invoice_fields_from_pdf(
                primary_path,
                company_deny_regex=cfg.rename.company_deny_regex,
                max_pages=cfg.rename.max_pages,
                run_dir=run_dir,
                log_path=log_path,
                vision_ocr_enabled=cfg.routing.vision_ocr_enabled,
                vision_ocr_provider=cfg.routing.vision_ocr_provider,
                sender=sender,
                subject=subject,
            )
            if fields is not None:
                _append_log(
                    log_path,
                    f"[INFO] extracted from pdf: vendor={fields.vendor}, date={fields.issue_date}, amount={fields.amount}",
                )

        if fields is not None:
            fields = _apply_invoice_guardrails_with_sender(fields, sender=sender)
            if fields.requires_manual:
                review_reasons = tuple(fields.review_reasons)
                if any(str(reason).startswith("non_invoice_signal:") for reason in review_reasons):
                    review_partial = partial_fields
                else:
                    review_partial = _partial_from_extracted_fields(fields, partial_fields)
                review_raw_ocr_text = fields.raw_ocr_text
                _append_log(
                    log_path,
                    "[INFO] manual review guard: "
                    f"source={fields.source or 'local'} vendor={fields.vendor} "
                    f"category={fields.vendor_category or '-'} "
                    f"reasons={','.join(fields.review_reasons) if fields.review_reasons else '-'}",
                )
                fields = None

        if fields is not None:
            vendor = fields.vendor
            issue_date = fields.issue_date
            amount = fields.amount
            invoice_no = fields.invoice_no
            project = fields.project
            values = {
                "vendor": vendor,
                "vendor_short": _vendor_short(vendor),
                "issue_date": issue_date,
                "amount": str(amount),
                "amount_comma": f"{amount:,}",
                "invoice_no": invoice_no or "",
                "project": project or "",
            }

            vendor_subdir = (
                _sanitize_filename(_render_template(cfg.rename.vendor_subdir_template, values))
                if cfg.rename.create_vendor_subdir
                else ""
            )
            _full_text = " ".join(filter(None, [
                subject,
                attachment_name,
                fields.raw_ocr_text if fields is not None else "",
            ]))
            route_subdir, route_reason = _determine_route_decision(
                cfg=cfg,
                save_dir=save_dir,
                sender=sender,
                attachment_name=attachment_name,
                subject=subject,
                project=project,
                full_text=_full_text,
            )
            filename = _sanitize_filename(
                _cleanup_rendered_filename(_render_template(cfg.rename.file_name_template, values))
            )
            dest_dir = save_dir
            if route_subdir:
                dest_dir = dest_dir / route_subdir
            if vendor_subdir:
                dest_dir = dest_dir / vendor_subdir
            final_path = _route_and_move_file(
                primary_path, dest_dir, filename, log_path, unresolved, log_prefix="rename/move"
            )
        else:
            # MVP rule: keep automation moving, but mark for manual review.
            vendor = review_partial.vendor
            issue_date = review_partial.issue_date
            amount = review_partial.amount
            project = review_partial.project

            _full_text = " ".join(filter(None, [
                subject,
                attachment_name,
                review_raw_ocr_text,
            ]))
            route_subdir, route_reason = _determine_route_decision(
                cfg=cfg,
                save_dir=save_dir,
                sender=sender,
                attachment_name=attachment_name,
                subject=subject,
                project=project,
                full_text=_full_text,
            )
            review_name = _build_review_filename(
                received_dt=received_dt,
                attachment_name=attachment_name,
                partial=review_partial,
                review_reasons=review_reasons,
                route_reason=route_reason,
            )
            # Option C routing: if project keyword matched uniquely, route to
            # {project_subdir}/要確認/ so the responsible PM finds it in their folder.
            # Otherwise fall back to central 要確認/review_required/ (unassigned).
            if route_subdir:
                dest_dir = save_dir / route_subdir / "要確認"
            else:
                dest_dir = save_dir / "要確認" / "review_required"
            final_path = _route_and_move_file(
                primary_path, dest_dir, review_name, log_path, unresolved, log_prefix="review move"
            )
    else:
        # Rename disabled: move the raw attachment name into save_dir.
        _full_text = " ".join(filter(None, [
            subject,
            attachment_name,
        ]))
        route_subdir, route_reason = _determine_route_decision(
            cfg=cfg,
            save_dir=save_dir,
            sender=sender,
            attachment_name=attachment_name,
            subject=subject,
            project=None,
            full_text=_full_text,
        )
        dest_dir = (save_dir / route_subdir) if route_subdir else save_dir
        final_path = _route_and_move_file(
            primary_path, dest_dir, primary_path.name, log_path, unresolved, log_prefix="move (rename disabled)"
        )

    # P0: routing_state for review_helper integration
    _routing_state = "auto_final"
    _review_reason_str: str | None = None
    if final_path and final_path.name.startswith("要確認"):
        _routing_state = "review_required"
        _review_reason_str = ",".join(review_reasons) if review_reasons else None

    if final_path.exists() and _is_within_dir(final_path, save_dir):
        record = {
            "key": manifest_key,
            "message_entry_id": entry_id,
            "attachment_name": attachment_name,
            "saved_path": str(final_path),
            "vendor": vendor,
            "issue_date": issue_date,
            "amount": amount,
            "invoice_no": invoice_no,
            "project": project,
            "sha256": sha,
            "was_encrypted": was_encrypted,
            "decrypted_path": str(final_path) if (was_encrypted and decrypted_path is not None) else None,
            "encrypted_original_path": encrypted_original_path,
            "route_subdir": route_subdir,
            "route_reason": route_reason,
            # --- P0: review_helper integration ---
            "sender": sender,
            "subject": subject,
            "body_snippet": body_snippet[:500] if body_snippet else None,
            "routing_state": _routing_state,
            "review_reason": _review_reason_str,
            "resolved_by": None,
            "resolved_at": None,
        }
        processed_index[manifest_key] = record
        _append_processed_attachment_record(processed_manifest_path, record)

    return SavedAttachment(
        message_entry_id=entry_id,
        message_subject=subject,
        sender=sender,
        received_time=received_time_s,
        attachment_name=attachment_name,
        saved_path=str(final_path),
        original_saved_path=str(pdf_path),
        vendor=vendor,
        issue_date=issue_date,
        amount=amount,
        invoice_no=invoice_no,
        project=project,
        sha256=sha,
        was_encrypted=was_encrypted,
        decrypted_path=str(final_path) if (was_encrypted and decrypted_path is not None) else None,
        encrypted_original_path=encrypted_original_path,
        route_subdir=route_subdir,
        route_reason=route_reason,
        processing_status="saved",
    )


def _validate_mail_config(cfg: ToolConfig, dry_run: bool, scan_only: bool) -> None:
    if scan_only or dry_run:
        return
    if not cfg.mail.error_to:
        raise ValueError("config.mail.error_to が空です（エラー通知先が必要です）。")
    if cfg.mail.send_success and (not cfg.mail.success_to):
        raise ValueError(
            "config.mail.success_to が空です（send_success=true のため通知先が必要です）。"
        )


def _validate_runtime_preflight(
    *,
    cfg: ToolConfig,
    folder_path: str,
    save_dir: Path | None,
    dry_run: bool,
    do_print: bool,
    log_path: Path,
) -> None:
    errors: list[str] = []
    warnings: list[str] = []

    if not folder_path.startswith("\\\\"):
        errors.append(f"outlook folder path must start with \\\\: {folder_path}")
    if cfg.outlook is None:
        errors.append("config.outlook is not set")
    if not dry_run and save_dir is None:
        errors.append("save_dir is required in execute mode")
    if cfg.routing.enabled and not cfg.routing.fallback_subdir:
        warnings.append("routing.fallback_subdir is empty")

    provider = str(cfg.routing.vision_ocr_provider or "").strip().lower()
    if cfg.routing.vision_ocr_enabled:
        env_requirements: dict[str, tuple[str, ...]] = {
            "claude": ("ANTHROPIC_API_KEY",),
            "openai": ("OPENAI_API_KEY",),
            "gemini": ("GOOGLE_API_KEY or GEMINI_API_KEY",),
            "azure_di": ("AZURE_DI_ENDPOINT", "AZURE_DI_KEY"),
            "azure_di_gpt4o": ("AZURE_DI_ENDPOINT", "AZURE_DI_KEY"),
            "hybrid_gpt4o": ("AZURE_DI_ENDPOINT", "AZURE_DI_KEY"),
            "auto": ("at least one OCR provider API key",),
        }
        if provider not in env_requirements:
            errors.append(f"unsupported vision_ocr_provider: {cfg.routing.vision_ocr_provider}")
        elif provider == "auto":
            provider_keys = (
                os.environ.get("ANTHROPIC_API_KEY"),
                os.environ.get("OPENAI_API_KEY"),
                os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY"),
                os.environ.get("AZURE_DI_KEY"),
            )
            if not any(provider_keys):
                errors.append("vision_ocr_enabled=true but no OCR provider API key is set")
        elif provider == "gemini":
            if not (os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY")):
                errors.append("vision_ocr_provider=gemini requires GOOGLE_API_KEY or GEMINI_API_KEY")
        else:
            missing = [name for name in env_requirements[provider] if not os.environ.get(name)]
            if missing:
                errors.append(
                    f"vision_ocr_provider={cfg.routing.vision_ocr_provider} requires env vars: {', '.join(missing)}"
                )

    if do_print:
        if cfg.print.method not in ("shell", "pdf_copy"):
            errors.append(f"unsupported print.method: {cfg.print.method}")
        if cfg.print.method == "shell" and not (cfg.print.printer_name or _default_printer()):
            errors.append("print requested but no printer_name/default printer is available")

    for warning in warnings:
        _append_log(log_path, f"[WARN] preflight: {warning}")
    if errors:
        raise ValueError("config preflight failed:\n- " + "\n- ".join(errors))
    _append_log(
        log_path,
        "preflight ok: "
        f"routing={cfg.routing.enabled} vision_ocr={cfg.routing.vision_ocr_enabled}/{cfg.routing.vision_ocr_provider} "
        f"print={do_print}",
    )


def _send_success_mail(cfg: ToolConfig, report: RunReport, run_dir: Path, dry_run: bool) -> None:
    if not cfg.mail.send_success:
        return
    if report.dry_run:
        return

    save_dir_html = (
        html_link_to_path(Path(report.save_dir)) if report.save_dir else "(未指定)"
    )
    merged_html = (
        html_link_to_path(Path(report.merged_pdf_path))
        if report.merged_pdf_path
        else "(なし)"
    )
    new_project_candidates = [a for a in report.saved_attachments if a.route_reason == "new_project_candidate"]
    new_project_candidates_html = (
        html_link_to_path(Path(report.new_project_candidates_path))
        if report.new_project_candidates_path
        else "(なし)"
    )
    paragraphs = [
        f"シナリオ12・13（受信メールPDF保存・一括印刷）が完了しました。",
        f"Run ID: {report.run_id}",
        f"ドライラン: {report.dry_run}",
        f"保存先: {save_dir_html}",
        f"Outlookフォルダ: {report.outlook_folder_path or '(未指定)'}",
        f"保存PDF数: {len(report.saved_attachments)}",
        f"URLのみタスク数: {len(report.url_only_tasks)}",
        f"新規案件候補: {len(new_project_candidates)}",
        f"新規案件候補CSV: {new_project_candidates_html}",
        f"結合PDF: {merged_html}",
        f"印刷ジョブ: {len(report.printed_paths)}",
        f"レポート: {html_link_to_path(run_dir)}",
    ]
    bullets: list[str] = []
    max_details = 30
    for a in report.saved_attachments[:max_details]:
        p = Path(a.saved_path)
        link = html_link_to_path(p, label=p.name)
        vendor = _vendor_short(a.vendor) if a.vendor else "(vendor不明)"
        issue = a.issue_date or "(日付不明)"
        amount = f"{a.amount:,}円" if a.amount is not None else "(金額不明)"
        # Avoid shadowing the imported `html` module below.
        inv = f" / {html.escape(a.invoice_no)}" if a.invoice_no else ""
        bullets.append(f"{link} / {vendor} / {issue} / {amount}{inv}")
    if len(report.saved_attachments) > max_details:
        bullets.append(f"... 他{len(report.saved_attachments) - max_details}件")
    if report.url_only_tasks:
        bullets.append("URLのみメールが検知されています（ダウンロード工程が未実装/未設定の場合は停止対象）")
    if new_project_candidates:
        bullets.append("新規案件候補あり: VPN接続後に軽技確認 -> フォルダ作成/マスタ更新 -> 再実行")

    html_body = build_simple_html(
        title="【完了】シナリオ12・13 受信メールPDF保存・一括印刷",
        paragraphs=paragraphs,
        bullets=bullets or None,
    )
    attachments = [run_dir / "report.csv", run_dir / "report.json"]
    if report.new_project_candidates_path:
        attachments.append(Path(report.new_project_candidates_path))
    send_outlook(
        OutlookEmail(
            to=cfg.mail.success_to,
            subject=f"【完了】受信メールPDF保存・一括印刷 Run={report.run_id}",
            html_body=html_body,
            attachments=attachments,
        ),
        dry_run=dry_run,
    )


def _send_error_mail(
    cfg: ToolConfig,
    run_id: str,
    run_dir: Path,
    summary: str,
    details: Sequence[str],
    dry_run: bool,
) -> None:
    paragraphs = [
        "シナリオ12・13（受信メールPDF保存・一括印刷）でエラー/未解決が発生しました。",
        f"Run ID: {run_id}",
        summary,
        f"レポート/ログ: {html_link_to_path(run_dir)}",
    ]
    base = build_simple_html(
        title="★エラー発生★ 受信メールPDF保存・一括印刷",
        paragraphs=paragraphs,
        bullets=list(details) if details else None,
    )
    # Preserve traceback formatting in HTML mail body.
    tb_text = "\n\n".join([d for d in details if d]) if details else ""
    tb_html = (
        f"<h4>Traceback</h4><pre>{html.escape(tb_text)}</pre>" if tb_text else ""
    )
    html_body = base.replace("</body>", f"{tb_html}\n  </body>")
    send_outlook(
        OutlookEmail(
            to=cfg.mail.error_to,
            cc=cfg.mail.error_cc,
            subject=f"★エラー発生★ 受信メールPDF保存・一括印刷 Run={run_id}",
            html_body=html_body,
            attachments=[p for p in [run_dir / "run.log", run_dir / "report.json"] if p.exists()],
        ),
        dry_run=dry_run,
    )


def _append_log(log_path: Path, line: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_path.write_text("", encoding="utf-8") if not log_path.exists() else None
    with log_path.open("a", encoding="utf-8") as f:
        f.write(f"[{ts}] {line}\n")


def _list_outlook_stores(mapi: Any) -> list[str]:
    stores: list[str] = []
    folders = _com_retry(lambda: mapi.Folders)
    count = _com_retry(lambda: int(folders.Count))
    for i in range(1, count + 1):
        try:
            stores.append(str(_com_retry(lambda i=i: folders.Item(i).Name)))
        except Exception:
            continue
    return stores


def _list_outlook_child_folders(mapi: Any, folder_path: str) -> list[str]:
    folder = _resolve_outlook_folder(mapi, folder_path)
    names: list[str] = []
    subfolders = _com_retry(lambda: folder.Folders)
    count = _com_retry(lambda: int(subfolders.Count))
    for i in range(1, count + 1):
        try:
            names.append(str(_com_retry(lambda i=i: subfolders.Item(i).Name)))
        except Exception:
            continue
    return names


def _finalize_and_notify(
    *,
    cfg: ToolConfig,
    run_id: str,
    started_at: datetime,
    run_dir: Path,
    dry_run: bool,
    mail_dry_run: bool,
    log_path: Path,
    unresolved: list,
    scan_only: bool,
    mail_enabled: bool,
    saved_rows: list,
    url_only_tasks: list,
    merged_pdf_path: str | None,
    printed_paths: list,
    folder_path: str,
    scanned: int,
    save_dir: Path | None,
) -> int:
    finished_at = datetime.now()
    report = RunReport(
        run_id=run_id,
        started_at=started_at.isoformat(timespec="seconds"),
        finished_at=finished_at.isoformat(timespec="seconds"),
        dry_run=dry_run,
        save_dir=str(save_dir) if save_dir else None,
        outlook_folder_path=folder_path,
        scanned_messages=scanned,
        saved_attachments=tuple(saved_rows),
        url_only_tasks=tuple(url_only_tasks),
        merged_pdf_path=merged_pdf_path,
        printed_paths=tuple(printed_paths),
        project_master_path=cfg.project_master_path,
        new_project_candidates_path=None,
        unresolved=tuple(unresolved),
    )
    new_project_candidates_path = run_dir / "new_project_candidates.csv"
    if _write_new_project_candidates_csv(new_project_candidates_path, saved_rows):
        report = RunReport(
            run_id=report.run_id,
            started_at=report.started_at,
            finished_at=report.finished_at,
            dry_run=report.dry_run,
            save_dir=report.save_dir,
            outlook_folder_path=report.outlook_folder_path,
            scanned_messages=report.scanned_messages,
            saved_attachments=report.saved_attachments,
            url_only_tasks=report.url_only_tasks,
            merged_pdf_path=report.merged_pdf_path,
            printed_paths=report.printed_paths,
            project_master_path=report.project_master_path,
            new_project_candidates_path=str(new_project_candidates_path),
            unresolved=report.unresolved,
        )
        _append_log(
            log_path,
            f"new project candidate report: {new_project_candidates_path}",
        )
    _write_report_json(run_dir / "report.json", report)
    _write_report_csv(run_dir / "report.csv", saved_rows)

    if unresolved:
        _append_log(log_path, f"unresolved_count={len(unresolved)}")
        if mail_enabled:
            _send_error_mail(
                cfg=cfg,
                run_id=run_id,
                run_dir=run_dir,
                summary="未解決項目があるため処理を停止しました。",
                details=unresolved,
                dry_run=mail_dry_run,
            )
        return 2

    if scan_only:
        _append_log(log_path, "scan_only done")
        return 0

    if mail_enabled:
        _send_success_mail(
            cfg=cfg, report=report, run_dir=run_dir, dry_run=mail_dry_run
        )
    _append_log(log_path, "success")
    return 0


def _run_debug_extract_mode(args: Any, project_master_override: Path | None, ap: Any) -> int:
    cfg = (
        _load_config(Path(args.config), project_master_override=project_master_override)
        if args.config
        else ToolConfig(project_master_path=str(project_master_override) if project_master_override else None)
    )
    run_id = _now_run_id()
    debug_run_dir = Path(cfg.artifact_dir) / f"debug_extract_{run_id}"
    _ensure_dir(debug_run_dir)
    log_path = debug_run_dir / "run.log"
    _append_log(
        log_path,
        f"start debug_extract run_id={run_id} config={args.config or '(none)'}",
    )

    pdf_paths = _expand_pdf_paths_for_debug(args.debug_extract_pdf)
    if not pdf_paths:
        ap.error("--debug-extract-pdf: no PDFs found in given paths")

    ok = 0
    ng = 0
    for pdf_path in pdf_paths:
        try:
            fields = _debug_extract_invoice_fields_from_pdf(
                pdf_path,
                company_deny_regex=cfg.rename.company_deny_regex,
                max_pages=int(cfg.rename.max_pages),
                debug_run_dir=debug_run_dir,
                log_path=log_path,
            )
            if fields:
                ok += 1
                invoice_no = fields.invoice_no or ""
                print(
                    f"OK\t{pdf_path}\tvendor={fields.vendor}\tissue_date={fields.issue_date}\tamount={fields.amount}\tinvoice_no={invoice_no}"
                )
            else:
                ng += 1
                print(f"NG\t{pdf_path}")
        except Exception as e:
            ng += 1
            _append_log(log_path, f"[ERROR] debug_extract failed: pdf={pdf_path} error={e}")
            print(f"NG\t{pdf_path}\terror={e}")

    print(f"summary: ok={ok} ng={ng} total={len(pdf_paths)}")
    print(f"debug artifacts: {debug_run_dir}")
    return 0 if ng == 0 else 2


def main(argv: Sequence[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", type=str, required=False, help="JSON config path")
    ap.add_argument(
        "--project-master-path",
        type=str,
        default=None,
        help="Override project_master path for this run only (xlsx/csv/tsv).",
    )
    ap.add_argument(
        "--scan-only",
        action="store_true",
        help="Scan candidate mails/attachments only (no save/print, no email send)",
    )
    ap.add_argument(
        "--outlook-profile",
        type=str,
        default=None,
        help="Outlook profile name to log on (optional). Example: RKM",
    )
    ap.add_argument(
        "--list-outlook-stores",
        action="store_true",
        help="List Outlook stores (top-level mailboxes) and exit",
    )
    ap.add_argument(
        "--list-outlook-folders",
        type=str,
        default=None,
        help="List child folders under the given Outlook folder path and exit (e.g. \\\\Store\\\\受信トレイ)",
    )
    ap.add_argument(
        "--debug-extract-pdf",
        type=str,
        nargs="+",
        default=None,
        help="Debug: extract invoice fields from given PDF file(s) or directory(ies) and exit (no Outlook).",
    )
    ap.add_argument(
        "--outlook-folder-path",
        type=str,
        default=None,
        help="Override config.outlook.folder_path (e.g. \\\\Store\\\\受信トレイ)",
    )
    ap.add_argument(
        "--include-read",
        action="store_true",
        help="Include read mails (override config.outlook.unread_only=false)",
    )
    ap.add_argument(
        "--execute",
        action="store_true",
        help="Actually save/decrypt (default: dry-run)",
    )
    ap.add_argument(
        "--print",
        action="store_true",
        help="Actually print (requires --execute and config.print.enabled=true)",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Force dry-run (ignores --execute)",
    )
    ap.add_argument(
        "--dry-run-mail",
        action="store_true",
        help="Do not send emails; print email content instead",
    )
    ap.add_argument(
        "--max-messages",
        type=int,
        default=None,
        help="Override config.outlook.max_messages",
    )
    ap.add_argument(
        "--save-dir",
        type=str,
        default=None,
        help="Override config.save_dir (save destination directory)",
    )
    ap.add_argument(
        "--payment-month",
        type=str,
        required=False,
        help="支払月サブフォルダ名（例: '2026.3末支払.RK'）。完成形をそのまま指定",
    )
    args = ap.parse_args(list(argv) if argv is not None else None)
    project_master_override = None
    if args.project_master_path:
        project_master_override = Path(args.project_master_path).expanduser()
        if not project_master_override.exists():
            ap.error(f"--project-master-path not found: {project_master_override}")

    if args.list_outlook_stores:
        mapi = _outlook_namespace(profile_name=args.outlook_profile)
        for s in _list_outlook_stores(mapi):
            print(s)
        return 0

    if args.list_outlook_folders:
        mapi = _outlook_namespace(profile_name=args.outlook_profile)
        parent = str(args.list_outlook_folders)
        for name in _list_outlook_child_folders(mapi, parent):
            # Print both name and the full path for copy/paste.
            print(f"{name}\t{parent.rstrip('\\\\')}\\{name}")
        return 0

    if args.debug_extract_pdf:
        return _run_debug_extract_mode(args, project_master_override, ap)

    if not args.config:
        _script_dir = Path(__file__).resolve().parent
        _env_path = _script_dir.parent / "config" / "env.txt"
        if _env_path.exists():
            _env = _env_path.read_text(encoding="utf-8").strip().upper()
            if _env == "PROD":
                args.config = str(_script_dir.parent / "config" / "tool_config_prod.json")
            else:
                args.config = str(_script_dir.parent / "config" / "tool_config_test.json")
        else:
            ap.error(
                "--config is required unless using --list-outlook-stores/--list-outlook-folders/--debug-extract-pdf"
            )

    cfg = _load_config(Path(args.config), project_master_override=project_master_override)
    scan_only = bool(args.scan_only)
    dry_run = True if scan_only else (bool(args.dry_run) or (not args.execute))
    mail_enabled = not scan_only
    # Two-key safety for printing: config.print.enabled AND CLI --print must be true.
    do_print = bool(args.print) and (not dry_run) and bool(cfg.print.enabled)
    mail_dry_run = True if scan_only else (bool(args.dry_run_mail) or dry_run)
    _validate_mail_config(cfg, dry_run=dry_run, scan_only=scan_only)

    run_id = _now_run_id()
    run_dir = Path(cfg.artifact_dir) / f"run_{run_id}"
    _ensure_dir(run_dir)
    log_path = run_dir / "run.log"
    _append_log(log_path, f"start run_id={run_id} dry_run={dry_run} config={args.config}")
    _append_log(
        log_path,
        f"project_master: {cfg.project_master_path or '(none)'} entries={len(cfg.project_master)}",
    )

    started_at = datetime.now()
    merged_pdf_path: str | None = None
    printed_paths: list[str] = []
    unresolved: list[str] = []

    try:
        save_dir: Path | None = None
        effective_save_dir = (str(args.save_dir).strip() if args.save_dir else None) or cfg.save_dir
        if effective_save_dir is None:
            if not dry_run:
                raise ValueError("config.save_dir が未設定です（保存先フォルダを指定してください）。")
        else:
            save_dir = Path(effective_save_dir)
            save_dir = _resolve_save_dir_with_fallback(
                save_dir,
                timeout_seconds=(2.0 if dry_run else 8.0),
                log_path=log_path,
            )
            # Existence check can hang on UNC paths; use timeout-based probing.
            if not dry_run:
                exists = _test_path_with_timeout(save_dir, timeout_seconds=8.0)
                if exists is not True:
                    raise ValueError(f"保存先フォルダが存在しません/到達できません: {save_dir}")
            else:
                # In dry-run/scan-only, do not block on network paths.
                if _is_unc_path(save_dir):
                    exists = _test_path_with_timeout(save_dir, timeout_seconds=2.0)
                    if exists is False:
                        _append_log(
                            log_path, f"[WARN] save_dir unreachable (dry-run): {save_dir}"
                        )
                    elif exists is None:
                        _append_log(
                            log_path,
                            f"[WARN] save_dir check timed out/failed (dry-run): {save_dir}",
                        )
                else:
                    if not save_dir.exists():
                        _append_log(
                            log_path, f"[WARN] save_dir does not exist (dry-run): {save_dir}"
                        )

        # --payment-month: append sub-folder to save_dir
        if args.payment_month and save_dir is not None:
            save_dir = save_dir / args.payment_month
            _ensure_dir(save_dir)
            _append_log(log_path, f"payment-month sub-folder: {save_dir}")

        if cfg.outlook is None:
            raise ValueError("config.outlook が未設定です。")

        outlook_cfg = cfg.outlook
        max_messages = args.max_messages if args.max_messages is not None else outlook_cfg.max_messages

        allow_re = re.compile(outlook_cfg.subject_allow_regex) if outlook_cfg.subject_allow_regex else None
        deny_re = re.compile(outlook_cfg.subject_deny_regex) if outlook_cfg.subject_deny_regex else None

        outlook_profile = (
            str(args.outlook_profile).strip()
            if args.outlook_profile
            else (outlook_cfg.profile_name or None)
        )
        folder_path = (
            str(args.outlook_folder_path).strip()
            if args.outlook_folder_path
            else outlook_cfg.folder_path
        )
        if outlook_profile:
            src = "cli" if args.outlook_profile else "config"
            _append_log(log_path, f"using outlook profile ({src}): {outlook_profile}")
        if args.outlook_folder_path:
            _append_log(log_path, f"using outlook folder_path (cli): {folder_path}")

        _validate_runtime_preflight(
            cfg=cfg,
            folder_path=folder_path,
            save_dir=save_dir,
            dry_run=dry_run,
            do_print=do_print,
            log_path=log_path,
        )

        mapi = _outlook_namespace(profile_name=outlook_profile)
        folder = _resolve_outlook_folder(mapi, folder_path)
        _append_log(log_path, f"resolved outlook folder: {folder_path}")

        effective_unread_only = False if bool(args.include_read) else bool(outlook_cfg.unread_only)
        if args.include_read:
            _append_log(log_path, "override: include_read=true -> unread_only=false")

        items = _iter_mail_items(folder, effective_unread_only, max_messages=max_messages)
        _append_log(log_path, f"loaded messages: {len(items)} (unread_only={effective_unread_only})")

        # Restrict by received time if configured.
        if outlook_cfg.received_within_days is not None:
            cutoff = datetime.now() - timedelta(days=outlook_cfg.received_within_days)
            kept = []
            for it in items:
                rt = _safe_received_time(it)
                if rt and rt >= cutoff:
                    kept.append(it)
            items = kept
            _append_log(log_path, f"filtered by received_within_days: {len(items)}")

        # Restrict by FlagStatus if configured.
        if outlook_cfg.use_flag_status:
            kept = []
            flag_read_errors = 0
            for it in items:
                try:
                    flag_status = int(it.FlagStatus)
                except Exception as e:
                    flag_status = -1  # Safe: skip this mail (don't accidentally re-process)
                    flag_read_errors += 1
                    _append_log(log_path, f"[WARN] FlagStatus read failed for message, skipping: {e}")
                if flag_status == 0:  # No flag = unprocessed
                    kept.append(it)
            items = kept
            _append_log(log_path, f"filtered by FlagStatus (unprocessed only): {len(items)}, skipped_due_to_read_error: {flag_read_errors}")

        password_notes: list[PasswordNote] = []
        candidates: list[Any] = []
        scanned = 0
        for it in items:
            scanned += 1
            try:
                subject = str(it.Subject or "")
            except Exception:
                subject = ""

            if not _subject_allowed(subject, allow_re) or _subject_denied(subject, deny_re):
                continue

            if _is_password_mail(subject):
                try:
                    body = str(it.Body or "")
                except Exception:
                    body = ""
                pw = _extract_password(body)
                rt = _safe_received_time(it)
                sender = _safe_sender(it)
                if pw and rt:
                    password_notes.append(
                        PasswordNote(received_time=rt, subject=subject, password=pw, sender=sender)
                    )
            candidates.append(it)

        _append_log(
            log_path,
            f"candidates={len(candidates)} password_notes={len(password_notes)}",
        )

        saved_rows: list[SavedAttachment] = []
        url_only_tasks: list[UrlOnlyTask] = []
        processed_manifest_path = Path(cfg.artifact_dir) / PROCESSED_MANIFEST_NAME
        processed_index = (
            _load_processed_attachment_index(processed_manifest_path) if not dry_run else {}
        )
        if not dry_run:
            _append_log(
                log_path,
                f"loaded processed manifest: {processed_manifest_path} records={len(processed_index)}",
            )

        for it in candidates:
            try:
                subject = str(it.Subject or "")
            except Exception:
                subject = ""

            sender = _safe_sender(it)
            rt_dt = _safe_received_time(it) or datetime.now()
            rt_s = rt_dt.strftime("%Y-%m-%d %H:%M:%S")

            try:
                entry_id = str(it.EntryID)
            except Exception:
                entry_id = f"(no_entry_id:{rt_s})"

            unresolved_before_mail = len(unresolved)

            def process_pdf(pdf_path: Path, att_name: str) -> SavedAttachment:
                return _process_saved_pdf_file(
                    cfg=cfg,
                    pdf_path=pdf_path,
                    attachment_name=att_name,
                    entry_id=entry_id,
                    subject=subject,
                    sender=sender,
                    body_snippet=body[:500] if body else "",
                    received_dt=rt_dt,
                    received_time_s=rt_s,
                    password_notes=password_notes,
                    save_dir=save_dir,
                    run_dir=run_dir,
                    log_path=log_path,
                    unresolved=unresolved,
                    processed_index=processed_index,
                    processed_manifest_path=processed_manifest_path,
                )

            def add_url_task(urls_list: list[str]) -> None:
                url_only_tasks.append(UrlOnlyTask(
                    message_entry_id=entry_id,
                    subject=subject,
                    sender=sender,
                    received_time=rt_s,
                    urls=tuple(_mask_url(u) for u in urls_list),
                ))

            # URL-only detection (body) for later; avoid logging sensitive bodies.
            try:
                body = str(it.Body or "")
            except Exception:
                body = ""
            urls = _extract_urls(body)

            # Save PDF attachments.
            attachments_saved = 0
            try:
                atts = it.Attachments
                att_count = int(atts.Count)
            except Exception:
                att_count = 0
                atts = None

            if att_count <= 0:
                if urls:
                    web_pdfs: list[Path] = []
                    if (
                        _WEB_DOWNLOAD_AVAILABLE
                        and cfg.enable_web_download
                        and cfg.web_download
                        and not scan_only
                    ):
                        web_dl_dir = run_dir / "web_downloads"
                        web_pdfs_result = dispatch_web_download(
                            cfg=cfg.web_download,
                            sender=sender,
                            urls=list(urls),
                            subject=subject,
                            received_dt=rt_dt,
                            download_dir=web_dl_dir,
                            log_path=log_path,
                            dry_run=dry_run,
                            scan_only=scan_only,
                        )
                        web_pdfs = web_pdfs_result.pdfs
                        for err_msg in web_pdfs_result.errors:
                            unresolved.append(err_msg)
                    if web_pdfs:
                        for wp in web_pdfs:
                            saved_rows.append(process_pdf(wp, wp.name))
                            attachments_saved += 1
                    else:
                        add_url_task(urls)
                continue

            for j in range(1, att_count + 1):
                try:
                    att = atts.Item(j)
                    att_name = str(att.FileName or "")
                except Exception:
                    continue
                att_name_l = att_name.lower().strip()
                is_pdf = att_name_l.endswith(".pdf")
                is_zip = att_name_l.endswith(".zip")
                if not (is_pdf or is_zip):
                    continue
                if is_zip and (not cfg.enable_zip_processing):
                    _append_log(
                        log_path,
                        f"[MVP-SKIP] zip processing is disabled: {att_name}",
                    )
                    continue

                received_ymd = rt_dt.strftime("%Y%m%d")
                subj_slug = _slug_subject(subject)
                att_slug = _sanitize_filename(att_name)
                tmp_name = f"{received_ymd}__{subj_slug}__{att_slug}"
                tmp_name = _sanitize_filename(tmp_name)

                if dry_run:
                    dest_preview = (save_dir / tmp_name) if save_dir else Path(tmp_name)
                    kind = "zip" if is_zip else "pdf"
                    _append_log(
                        log_path,
                        f"[DRY] save attachment({kind}): {att_name} -> {dest_preview}",
                    )
                    if is_pdf:
                        # fake sha for report
                        saved_rows.append(
                            SavedAttachment(
                                message_entry_id=entry_id,
                                message_subject=subject,
                                sender=sender,
                                received_time=rt_s,
                                attachment_name=att_name,
                                saved_path=str(dest_preview),
                                original_saved_path=str(dest_preview),
                                vendor=None,
                                issue_date=None,
                                amount=None,
                                invoice_no=None,
                                project=None,
                                sha256="(dry-run)",
                                was_encrypted=False,
                                decrypted_path=None,
                                encrypted_original_path=None,
                                route_subdir=None,
                                route_reason=None,
                                processing_status="dry_run",
                            )
                        )
                        attachments_saved += 1
                    continue

                if save_dir is None:
                    raise RuntimeError("save_dir must be set in execute mode")
                attachments_dir = run_dir / "attachments"
                _ensure_dir(attachments_dir)

                if is_pdf:
                    dest_path = _unique_path(attachments_dir / tmp_name)
                    _append_log(
                        log_path, f"save attachment(pdf): {att_name} -> {dest_path}"
                    )
                    att.SaveAsFile(str(dest_path))
                    saved_rows.append(process_pdf(dest_path, att_name))
                    attachments_saved += 1
                    continue

                # ZIP attachment: extract PDFs and process each.
                zip_path = _unique_path(attachments_dir / tmp_name)
                _append_log(log_path, f"save attachment(zip): {att_name} -> {zip_path}")
                att.SaveAsFile(str(zip_path))

                extracted_parent = attachments_dir / "extracted"
                _ensure_dir(extracted_parent)
                extracted_dir = _unique_dir(
                    extracted_parent / _sanitize_filename(zip_path.stem)
                )
                _ensure_dir(extracted_dir)

                pw_candidates = _candidate_passwords_with_fallback(
                    target_time=rt_dt,
                    target_sender=sender,
                    notes=password_notes,
                    log_path=log_path,
                )
                attempts: list[str | None] = [None] + list(pw_candidates)
                extracted_pdfs: list[Path] = []
                last_error: Exception | None = None
                for pw in attempts:
                    _recreate_dir(extracted_dir)
                    try:
                        extracted_pdfs = _extract_pdf_members_from_zip(
                            zip_path,
                            password=pw,
                            out_dir=extracted_dir,
                            log_path=log_path,
                        )
                        last_error = None
                        break
                    except Exception as e:
                        last_error = e
                        if _is_wrong_zip_password_error(e) and pw is not attempts[-1]:
                            _append_log(
                                log_path,
                                f"[WARN] zip extract failed (try next password): {zip_path.name} error={e}",
                            )
                            continue
                        break

                if last_error is not None:
                    hint = ""
                    if isinstance(last_error, NotImplementedError) and (not _PYZIPPER_AVAILABLE):
                        hint = " / AES暗号ZIPの可能性があります（pyzipperが必要）。"
                    unresolved.append(
                        f"ZIP展開に失敗: {zip_path.name} (subject={subject}) error={last_error}{hint}"
                    )
                    continue

                if not extracted_pdfs:
                    unresolved.append(
                        f"ZIP内にPDFが見つかりません: {zip_path.name} (subject={subject})"
                    )
                    continue

                for pdf_path in extracted_pdfs:
                    try:
                        rel = str(pdf_path.relative_to(extracted_dir))
                    except Exception:
                        rel = pdf_path.name
                    display_name = f"{att_name}::{rel}"
                    saved_rows.append(process_pdf(pdf_path, display_name))
                    attachments_saved += 1

            if attachments_saved == 0 and urls:
                add_url_task(urls)

            # Mimic the human workflow: reading/handling a mail marks it as processed.
            # This is intentionally best-effort to avoid blocking invoice collection.
            if (
                (not dry_run)
                and bool(outlook_cfg.mark_as_read_on_success)
                and attachments_saved > 0
                and len(unresolved) == unresolved_before_mail
            ):
                try:
                    it.UnRead = False
                    it.Save()
                    _append_log(log_path, f"marked as read: entry_id={entry_id}")
                except Exception as e:
                    _append_log(
                        log_path,
                        f"[WARN] failed to mark as read: entry_id={entry_id} error={e}",
                    )

            # Set FlagStatus = 2 (red flag) on successful processing.
            if (
                (not dry_run)
                and bool(outlook_cfg.mark_flag_on_success)
                and attachments_saved > 0
                and len(unresolved) == unresolved_before_mail
            ):
                try:
                    it.FlagStatus = 2
                    it.Save()
                    _append_log(log_path, f"marked with red flag: entry_id={entry_id}")
                except Exception as e:
                    _append_log(
                        log_path,
                        f"[WARN] failed to set flag: entry_id={entry_id} error={e}",
                    )

        # Cleanup web browser sessions
        if _WEB_DOWNLOAD_AVAILABLE and cfg.enable_web_download:
            cleanup_web_sessions()

        # Merge/Print targets (prefer decrypted if exists). Skip existence checks in dry-run/scan.
        printable: list[Path] = []
        if not dry_run:
            for r in saved_rows:
                p = Path(r.decrypted_path) if r.decrypted_path else Path(r.saved_path)
                if p.exists():
                    printable.append(p)

        if not dry_run and cfg.merge.enabled and printable:
            merged = run_dir / cfg.merge.output_name
            _append_log(log_path, f"merge pdfs: {len(printable)} -> {merged}")
            _merge_pdfs(printable, merged)
            merged_pdf_path = str(merged)

        # Print (jidoka: do not print in dry-run)
        if do_print and printable:
            targets = [Path(merged_pdf_path)] if merged_pdf_path else printable
            if cfg.print.method == "shell":
                printer = cfg.print.printer_name or _default_printer()
                for p in targets:
                    _append_log(log_path, f"print: {p} -> printer={printer or '(default)'}")
                    _print_via_shell(p, printer_name=printer)
                    printed_paths.append(str(p))
                    time.sleep(1.0)  # small gap to avoid overwhelming the print spooler
            elif cfg.print.method == "pdf_copy":
                out_dir = (
                    Path(cfg.print.pdf_output_dir)
                    if cfg.print.pdf_output_dir
                    else (run_dir / "printed_pdf")
                )
                for p in targets:
                    out_path = _print_to_pdf_copy(p, out_dir)
                    _append_log(log_path, f"print(pdf_copy): {p} -> {out_path}")
                    printed_paths.append(str(out_path))
            else:
                raise ValueError(f"unsupported print.method: {cfg.print.method}")

        # URL-only tasks are unresolved by default (unless fail_on_url_only_mail is false)
        if (not scan_only) and cfg.fail_on_url_only_mail and url_only_tasks:
            unresolved.append(
                f"URL-only mails detected: {len(url_only_tasks)} (web download is out of MVP scope)"
            )

        return _finalize_and_notify(
            cfg=cfg,
            run_id=run_id,
            started_at=started_at,
            run_dir=run_dir,
            dry_run=dry_run,
            mail_dry_run=mail_dry_run,
            log_path=log_path,
            unresolved=unresolved,
            scan_only=scan_only,
            mail_enabled=mail_enabled,
            saved_rows=saved_rows,
            url_only_tasks=url_only_tasks,
            merged_pdf_path=merged_pdf_path,
            printed_paths=printed_paths,
            folder_path=folder_path,
            scanned=scanned,
            save_dir=save_dir,
        )

    except Exception as e:
        tb = traceback.format_exc()
        _append_log(log_path, f"exception: {e}\n{tb}")
        if mail_enabled:
            try:
                _send_error_mail(
                    cfg=cfg,
                    run_id=run_id,
                    run_dir=run_dir,
                    summary=str(e),
                    details=[tb],
                    dry_run=mail_dry_run,
                )
            except Exception:
                # last resort: do not hide the original failure; still return non-zero.
                pass
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
