"""Build vendor_hints.json from golden dataset and vendor master."""

from __future__ import annotations

import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from training_data import (
    DEFAULT_BACKUP_PST,
    DEFAULT_TEACHER_PDF_DIR,
    build_teacher_pdf_records,
    project_keyword_candidates,
)
from vendor_matching import (
    DEFAULT_BLOCKED_VENDOR_CANDIDATES,
    DEFAULT_NON_VENDOR_PATTERNS,
    LEGACY_VENDOR_HINTS,
    clean_vendor_text,
    extract_sender_domain,
    is_non_vendor_candidate,
    resolve_preferred_canonical,
)

CONFIG_DIR = Path(__file__).resolve().parent.parent / "config"
MASTER_PATH = CONFIG_DIR / "vendor_delivery_master_v4.xlsx"
GOLDEN_PATH = CONFIG_DIR / "golden_dataset_v4.json"
OUTPUT_PATH = CONFIG_DIR / "vendor_hints.json"


def _collapse_category(current: str | None, incoming: str | None) -> str | None:
    priorities = {"variable": 3, "gray": 2, "fixed": 1}
    current_val = priorities.get((current or "").lower(), 0)
    incoming_val = priorities.get((incoming or "").lower(), 0)
    if incoming_val >= current_val:
        return incoming if incoming_val else current
    return current


def _is_service_row(company_name: str, delivery_type: str) -> bool:
    if delivery_type == "パスワード通知":
        return True
    lowered = clean_vendor_text(company_name).lower()
    return any(
        token in lowered
        for token in ("請求web", "パスワード通知", "f-filter", "mypage")
    )


def _merge_entry(entries: dict[str, dict], payload: dict) -> None:
    canonical = resolve_preferred_canonical(payload.get("canonical"))
    if not canonical:
        return

    entry = entries.setdefault(
        canonical,
        {
            "canonical": canonical,
            "aliases": set(),
            "sender_domains": set(),
            "web_domains": set(),
            "address_keywords": set(),
            "registration_numbers": set(),
            "project_keywords": set(),
            "deny_patterns": set(),
            "sources": set(),
            "category": None,
        },
    )
    for field_name in (
        "aliases",
        "sender_domains",
        "web_domains",
        "address_keywords",
        "registration_numbers",
        "project_keywords",
        "deny_patterns",
        "sources",
    ):
        for value in payload.get(field_name) or ():
            if value:
                entry[field_name].add(str(value).strip())

    category = str(payload.get("category") or "").strip().lower()
    if category and category != "regression":
        entry["category"] = _collapse_category(entry["category"], category)


def build_vendor_hints() -> dict:
    entries: dict[str, dict] = {}
    category_counter: Counter[str] = Counter()
    teacher_canonicals: set[str] = set()
    teacher_rows_with_vendor = 0
    teacher_rows_with_sender_context = 0
    teacher_sender_domain_count = 0
    teacher_project_keyword_count = 0

    golden_rows = json.loads(GOLDEN_PATH.read_text(encoding="utf-8"))
    for row in golden_rows:
        canonical = str(row.get("vendor") or "").strip()
        category = str(row.get("category") or "").strip().lower()
        is_noise, _ = is_non_vendor_candidate(canonical)
        if is_noise and canonical not in LEGACY_VENDOR_HINTS:
            continue
        if canonical:
            _merge_entry(
                entries,
                {
                    "canonical": canonical,
                    "category": category,
                    "sources": ("golden_dataset_v4",),
                },
            )
            if category and category != "regression":
                category_counter[category] += 1

    wb = load_workbook(MASTER_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else "" for v in next(rows)]
    for values in rows:
        row = dict(zip(headers, values))
        company_name = str(row.get("会社名") or "").strip()
        delivery_type = str(row.get("配送形式") or "").strip()
        if not company_name or _is_service_row(company_name, delivery_type):
            continue
        is_noise, _ = is_non_vendor_candidate(company_name)
        if is_noise and company_name not in LEGACY_VENDOR_HINTS:
            continue

        sender_domain = extract_sender_domain(str(row.get("送信者アドレス") or ""))
        web_domain = str(row.get("Webポータルドメイン") or "").strip().lower()
        display_name = clean_vendor_text(str(row.get("送信者表示名") or ""))

        aliases = set()
        if display_name and display_name != clean_vendor_text(company_name):
            aliases.add(display_name)
        if company_name in LEGACY_VENDOR_HINTS:
            aliases.update(LEGACY_VENDOR_HINTS[company_name])

        _merge_entry(
            entries,
            {
                "canonical": company_name,
                "aliases": sorted(aliases),
                "sender_domains": [sender_domain] if sender_domain else [],
                "web_domains": [web_domain] if web_domain else [],
                "sources": ("vendor_delivery_master_v4",),
            },
        )

    teacher_records = build_teacher_pdf_records(
        pdf_root=DEFAULT_TEACHER_PDF_DIR,
        pst_path=DEFAULT_BACKUP_PST,
    )
    for record in teacher_records:
        canonical = record.canonical_vendor or resolve_preferred_canonical(record.vendor)
        if not canonical:
            continue
        teacher_rows_with_vendor += 1
        teacher_canonicals.add(canonical)

        aliases: list[str] = []
        if record.vendor and clean_vendor_text(record.vendor) != clean_vendor_text(canonical):
            aliases.append(record.vendor)
        project_keywords = list(project_keyword_candidates(record.project))
        sender_domains = list(record.sender_domains)
        if sender_domains:
            teacher_rows_with_sender_context += 1
            teacher_sender_domain_count += len(sender_domains)
        teacher_project_keyword_count += len(project_keywords)

        sources = ["teacher_pdf"]
        if sender_domains:
            sources.append("backup_pst_context")

        _merge_entry(
            entries,
            {
                "canonical": canonical,
                "aliases": aliases,
                "sender_domains": sender_domains,
                "project_keywords": project_keywords,
                "sources": tuple(sources),
            },
        )

    vendors = []
    for canonical in sorted(entries):
        entry = entries[canonical]
        vendors.append(
            {
                "canonical": canonical,
                "aliases": sorted(v for v in entry["aliases"] if v and v != canonical),
                "sender_domains": sorted(v for v in entry["sender_domains"] if v),
                "web_domains": sorted(v for v in entry["web_domains"] if v),
                "address_keywords": sorted(v for v in entry["address_keywords"] if v),
                "registration_numbers": sorted(v for v in entry["registration_numbers"] if v),
                "project_keywords": sorted(v for v in entry["project_keywords"] if v),
                "deny_patterns": sorted(v for v in entry["deny_patterns"] if v),
                "category": entry["category"],
                "sources": sorted(entry["sources"]),
            }
        )

    return {
        "version": 1,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "rules": {
            "blocked_vendor_candidates": list(DEFAULT_BLOCKED_VENDOR_CANDIDATES),
            "non_vendor_patterns": list(DEFAULT_NON_VENDOR_PATTERNS),
        },
        "meta": {
            "golden_rows": len(golden_rows),
            "teacher_pdf_rows": len(teacher_records),
            "teacher_pdf_rows_with_vendor": teacher_rows_with_vendor,
            "teacher_pdf_rows_with_sender_context": teacher_rows_with_sender_context,
            "teacher_pdf_unique_canonicals": len(teacher_canonicals),
            "teacher_pdf_sender_domains": teacher_sender_domain_count,
            "teacher_pdf_project_keywords": teacher_project_keyword_count,
            "vendors": len(vendors),
            "category_counts": dict(category_counter),
        },
        "vendors": vendors,
    }


def main() -> None:
    payload = build_vendor_hints()
    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"wrote {OUTPUT_PATH} ({len(payload['vendors'])} vendors)")


if __name__ == "__main__":
    main()
