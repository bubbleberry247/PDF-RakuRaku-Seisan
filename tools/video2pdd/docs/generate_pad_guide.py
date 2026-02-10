"""
Generate client-facing PAD (Power Automate Desktop) procedure guide in Excel.

Usage:
    python -m tools.video2pdd.docs.generate_pad_guide [--output PATH]

Output: Excel file with step-by-step instructions for PAD recording and
Robin script extraction, formatted for client delivery.
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------------ #
#  Style constants
# ------------------------------------------------------------------ #

_TITLE_FONT = Font(name="Yu Gothic UI", bold=True, size=16)
_SECTION_FONT = Font(name="Yu Gothic UI", bold=True, size=13, color="1F4E79")
_SUBSECTION_FONT = Font(name="Yu Gothic UI", bold=True, size=11)
_BODY_FONT = Font(name="Yu Gothic UI", size=10)
_NOTE_FONT = Font(name="Yu Gothic UI", size=10, color="CC0000")
_HINT_FONT = Font(name="Yu Gothic UI", size=10, italic=True, color="666666")
_STEP_NUM_FONT = Font(name="Yu Gothic UI", bold=True, size=11, color="FFFFFF")

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_STEP_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
_ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
_WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_NOTE_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

_WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


# ------------------------------------------------------------------ #
#  Content data
# ------------------------------------------------------------------ #

DOCUMENT_TITLE = "Power Automate Desktop 操作録画・データ送付手順書"
DOCUMENT_SUBTITLE = "RPA業務プロセス設計書(PDD)作成用"

SECTIONS = [
    {
        "title": "1. Power Automate Desktop のインストール",
        "intro": "Power Automate Desktop（PAD）は、Microsoft社が無料で提供するRPAツールです。\n"
                 "Windows 11には標準搭載されています。\n"
                 "Windows 10の場合は、事前にお渡しするインストーラー(.exe)でインストールしてください。",
        "steps": [
            {
                "num": "1-1",
                "action": "インストーラーを実行",
                "detail": "お渡しした「Setup.Microsoft.PowerAutomate.exe」を\n"
                          "ダブルクリックして実行します。\n\n"
                          "※ Windows 11 の場合は標準搭載のため、\n"
                          "  スタートメニューで「Power Automate」を検索してください。\n"
                          "  見つかればこの手順はスキップできます。",
                "note": "",
            },
            {
                "num": "1-2",
                "action": "インストール画面の操作",
                "detail": "画面の指示に従い「次へ」をクリックして進みます。\n"
                          "特別な設定変更は不要です。\n"
                          "「インストール」をクリックするとインストールが開始されます。",
                "note": "※ 管理者権限の確認画面が出た場合は「はい」をクリック",
            },
            {
                "num": "1-3",
                "action": "ブラウザ拡張機能のインストール（任意）",
                "detail": "インストール中に「ブラウザ拡張機能をインストールしますか」\n"
                          "と表示される場合があります。\n\n"
                          "ブラウザ操作の録画精度が向上するため、\n"
                          "お使いのブラウザ（Edge/Chrome）の拡張機能を\n"
                          "インストールすることを推奨します。",
                "note": "",
            },
            {
                "num": "1-4",
                "action": "PADを起動してサインイン",
                "detail": "スタートメニューから「Power Automate」を起動します。\n\n"
                          "初回起動時にMicrosoftアカウントでサインインします。\n"
                          "会社のMicrosoft 365アカウントがある場合はそちらを使用してください。\n"
                          "個人のMicrosoftアカウント（無料）でも利用可能です。",
                "note": "※ サインインしないと利用できません",
            },
            {
                "num": "1-5",
                "action": "確認：ホーム画面が表示される",
                "detail": "「マイ フロー」画面が表示されれば準備完了です。",
                "note": "",
            },
        ],
        "tips": [
            "Windows 11をお使いの場合、スタートメニューに「Power Automate」が最初からあります。インストール不要です。",
            "Microsoft Storeからもインストールできますが、社内ポリシーでStoreが制限されている場合は.exeをご利用ください。",
        ],
    },
    {
        "title": "2. 業務操作の録画",
        "intro": "PADの「レコーダー」機能を使って、普段の業務操作（クリック・入力・画面遷移）を\n"
                 "自動的に記録します。録画中は普段どおり操作するだけでOKです。",
        "steps": [
            {
                "num": "2-1",
                "action": "新しいフローを作成",
                "detail": "PADホーム画面の「＋ 新しいフロー」をクリックし、\n"
                          "フロー名を入力します（例：「経費精算 申請処理」）。\n"
                          "「作成」をクリックしてフローエディタを開きます。",
                "note": "",
            },
            {
                "num": "2-2",
                "action": "レコーダーを起動",
                "detail": "フローエディタ上部の「レコーダー」ボタン（●）をクリックします。\n"
                          "レコーダーのコントロールパネルが表示されます。",
                "note": "",
            },
            {
                "num": "2-3",
                "action": "録画を開始",
                "detail": "レコーダーの「記録」ボタン（赤い●）をクリックして録画を開始します。\n\n"
                          "ここから、普段どおりに業務操作を行ってください：\n"
                          "  ・ブラウザを開く\n"
                          "  ・ログインする\n"
                          "  ・ボタンやメニューをクリックする\n"
                          "  ・テキストを入力する\n"
                          "  ・ファイルを開く/保存する\n"
                          "  など",
                "note": "※ 操作はゆっくり・確実に行ってください（早すぎると取りこぼすことがあります）",
            },
            {
                "num": "2-4",
                "action": "録画を停止",
                "detail": "業務操作が終わったら、レコーダーの「停止」ボタン（■）をクリック、\n"
                          "または「完了」をクリックします。\n\n"
                          "フローエディタに録画した操作がアクション一覧として表示されます。",
                "note": "",
            },
            {
                "num": "2-5",
                "action": "フローを保存",
                "detail": "フローエディタの左上「保存」アイコン（💾）をクリック、\n"
                          "または Ctrl+S で保存します。",
                "note": "※ 必ず保存してください",
            },
        ],
        "tips": [
            "録画中にパスワードを入力する場面がある場合、実際のパスワードが記録されます。"
            "送付前に確認し、パスワード部分は「****」等に置き換えてください。",
            "一連の業務を最初から最後まで通しで録画してください。途中で止めると手順が不完全になります。",
            "録画が長い場合（30分以上）は、業務の区切りで分けて複数フローにしてもOKです。",
        ],
    },
    {
        "title": "3. 録画データの取得（コピー）",
        "intro": "録画したフローのデータをテキスト形式で取り出します。\n"
                 "この操作により、業務手順が構造化されたテキストデータとして保存できます。\n\n"
                 "操作手順の動画はこちら：\n"
                 "https://drive.google.com/file/d/1apqnCYkZEsK1aruHi2TBQkVxqd1A6NO6/view?usp=sharing",
        "steps": [
            {
                "num": "3-1",
                "action": "フローを編集画面で開く",
                "detail": "PADホーム画面から、録画したフローをダブルクリックして\n"
                          "フローエディタを開きます。",
                "note": "",
            },
            {
                "num": "3-2",
                "action": "全アクションを選択",
                "detail": "フローエディタ内のアクション一覧をクリックした後、\n"
                          "キーボードで Ctrl+A を押して全アクションを選択します。\n"
                          "（全行がハイライトされます）",
                "note": "",
            },
            {
                "num": "3-3",
                "action": "コピー",
                "detail": "Ctrl+C を押してコピーします。\n"
                          "（見た目には何も変わりませんが、クリップボードにコピーされています）",
                "note": "",
            },
            {
                "num": "3-4",
                "action": "メモ帳に貼り付け",
                "detail": "「メモ帳」（Notepad）を新規で開きます。\n"
                          "（スタートメニューで「メモ帳」と検索）\n\n"
                          "Ctrl+V で貼り付けます。\n"
                          "テキストが表示されれば成功です。",
                "note": "",
            },
            {
                "num": "3-5",
                "action": "ファイルとして保存",
                "detail": "メモ帳で「ファイル」→「名前を付けて保存」を選択。\n\n"
                          "  ファイル名：業務名.robin\n"
                          "     （例：経費精算_申請処理.robin）\n\n"
                          "  エンコード：UTF-8\n"
                          "     （デフォルトでUTF-8ですが念のため確認）\n\n"
                          "  保存先：デスクトップなど分かりやすい場所",
                "note": "※ 拡張子は .robin としてください（.txt でも処理可能ですが .robin 推奨）",
            },
        ],
        "tips": [
            "コピーされるデータは「Robin言語」と呼ばれるPAD内部のスクリプト形式です。",
            "中身を読む必要はありません。そのまま送付していただければこちらで処理します。",
        ],
    },
    {
        "title": "4. データの送付",
        "intro": "保存した .robin ファイルをメールまたはチャットで送付してください。",
        "steps": [
            {
                "num": "4-1",
                "action": "送付前の確認",
                "detail": "以下を確認してください：\n\n"
                          "  □ .robin ファイルが保存されている\n"
                          "  □ ファイルを開くとテキストが表示される（空でない）\n"
                          "  □ パスワード等の機密情報が含まれていないか確認\n"
                          "     （含まれている場合は「****」に置換）\n"
                          "  □ ファイル名に業務名が含まれている",
                "note": "",
            },
            {
                "num": "4-2",
                "action": "メール / チャットで送付",
                "detail": ".robin ファイルを添付して送付してください。\n\n"
                          "メール本文に以下の情報を添えていただけると、\n"
                          "より正確なPDD（業務プロセス設計書）を作成できます：\n\n"
                          "  ・業務名（例：経費精算の申請処理）\n"
                          "  ・使用しているアプリケーション名\n"
                          "  ・業務の頻度（毎日/週次/月次/随時）\n"
                          "  ・特記事項（例外処理、判断が必要な箇所など）",
                "note": "",
            },
            {
                "num": "4-3",
                "action": "（任意）操作動画の送付",
                "detail": "録画中の画面を動画で撮影したもの（MP4等）があると、\n"
                          "より詳細なPDDを作成できます。\n\n"
                          "動画は必須ではありませんが、あると精度が向上します。\n"
                          "OBS Studio 等のスクリーンレコーダーで録画できます。",
                "note": "",
            },
        ],
        "tips": [
            "複数の業務がある場合は、業務ごとに別々の .robin ファイルを作成してください。",
            "ファイルサイズは通常数KB〜数十KBと小さいため、メール添付で問題ありません。",
        ],
    },
]

FAQ_DATA = [
    ("PADは有料ですか？", "いいえ。Power Automate Desktopは無料でご利用いただけます。"),
    ("録画中にミスした場合は？", "最初からやり直す必要はありません。\n"
     "録画を停止した後、フローエディタでアクションの削除・追加・並べ替えが可能です。\n"
     "ただし、最初から通しで録画し直すのが最も確実です。"),
    ("どの操作が録画されますか？", "マウスクリック、キーボード入力、ウィンドウ切替、\n"
     "ブラウザ操作（URL遷移、フォーム入力）などが自動的に記録されます。"),
    ("録画できない操作はありますか？", "右クリックメニュー、ドラッグ＆ドロップ、\n"
     "一部の特殊なアプリケーション（管理者権限が必要なもの）は\n"
     "録画できない場合があります。"),
    ("複数モニターでも録画できますか？", "はい。複数モニターに跨る操作も録画可能です。"),
    (".robin ファイルの中身を\n見る必要がありますか？", "いいえ。中身は自動生成されたスクリプトなので、\n"
     "確認や編集の必要はありません。そのまま送付してください。"),
]


# ------------------------------------------------------------------ #
#  Helper functions
# ------------------------------------------------------------------ #

def _apply_border(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        ws.cell(row=row, column=col).border = _THIN_BORDER


def _write_title_area(ws, row: int) -> int:
    """Write document title area. Returns next row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=DOCUMENT_TITLE)
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=DOCUMENT_SUBTITLE)
    cell.font = Font(name="Yu Gothic UI", size=11, color="666666")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(
        row=row, column=1,
        value=f"作成日: {datetime.now().strftime('%Y年%m月%d日')}",
    )
    cell.font = _HINT_FONT
    row += 2  # blank row

    return row


def _write_section(ws, row: int, section: dict) -> int:
    """Write a section with steps. Returns next row."""
    # Section title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=section["title"])
    cell.font = _SECTION_FONT
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = _SECTION_FILL
    row += 1

    # Section intro
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=section["intro"])
    cell.font = _BODY_FONT
    cell.alignment = _WRAP_ALIGN
    row += 1

    # Steps header
    headers = ["No.", "操作", "詳細", "注意事項"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(name="Yu Gothic UI", bold=True, size=10, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN
        cell.border = _THIN_BORDER
    row += 1

    # Steps
    for i, step in enumerate(section["steps"]):
        is_alt = i % 2 == 1

        ws.cell(row=row, column=1, value=step["num"]).font = Font(
            name="Yu Gothic UI", bold=True, size=10,
        )
        ws.cell(row=row, column=1).alignment = _CENTER_ALIGN

        ws.cell(row=row, column=2, value=step["action"]).font = Font(
            name="Yu Gothic UI", bold=True, size=10,
        )
        ws.cell(row=row, column=2).alignment = _WRAP_ALIGN

        ws.cell(row=row, column=3, value=step["detail"]).font = _BODY_FONT
        ws.cell(row=row, column=3).alignment = _WRAP_ALIGN

        note_cell = ws.cell(row=row, column=4, value=step["note"])
        if step["note"]:
            note_cell.font = _NOTE_FONT
        else:
            note_cell.font = _BODY_FONT
        note_cell.alignment = _WRAP_ALIGN

        if is_alt:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = _ALT_FILL

        _apply_border(ws, row, 4)
        row += 1

    # Tips
    if section.get("tips"):
        row += 1  # blank
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value="ポイント・注意事項")
        cell.font = _SUBSECTION_FONT
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = _WARN_FILL
        row += 1

        for tip in section["tips"]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = ws.cell(row=row, column=1, value=f"  ・{tip}")
            cell.font = _BODY_FONT
            cell.alignment = _WRAP_ALIGN
            row += 1

    row += 1  # blank between sections
    return row


def _write_faq(ws, row: int) -> int:
    """Write FAQ section. Returns next row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="5. よくあるご質問（FAQ）")
    cell.font = _SECTION_FONT
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = _SECTION_FILL
    row += 1

    # FAQ header
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws.cell(row=row, column=1, value="質問")
    cell.font = Font(name="Yu Gothic UI", bold=True, size=10, color="FFFFFF")
    cell.fill = _HEADER_FILL
    cell.alignment = _CENTER_ALIGN
    cell.border = _THIN_BORDER
    ws.cell(row=row, column=2).fill = _HEADER_FILL
    ws.cell(row=row, column=2).border = _THIN_BORDER

    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=3, value="回答")
    cell.font = Font(name="Yu Gothic UI", bold=True, size=10, color="FFFFFF")
    cell.fill = _HEADER_FILL
    cell.alignment = _CENTER_ALIGN
    cell.border = _THIN_BORDER
    ws.cell(row=row, column=4).fill = _HEADER_FILL
    ws.cell(row=row, column=4).border = _THIN_BORDER
    row += 1

    for i, (q, a) in enumerate(FAQ_DATA):
        is_alt = i % 2 == 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value=q)
        cell.font = Font(name="Yu Gothic UI", bold=True, size=10)
        cell.alignment = _WRAP_ALIGN
        cell.border = _THIN_BORDER
        ws.cell(row=row, column=2).border = _THIN_BORDER

        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=3, value=a)
        cell.font = _BODY_FONT
        cell.alignment = _WRAP_ALIGN
        cell.border = _THIN_BORDER
        ws.cell(row=row, column=4).border = _THIN_BORDER

        if is_alt:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = _ALT_FILL

        row += 1

    return row


# ------------------------------------------------------------------ #
#  Main generator
# ------------------------------------------------------------------ #

def generate_pad_guide(output_path: str) -> str:
    """Generate PAD procedure guide Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PAD操作手順書"

    # Page setup
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 30

    # Write content
    row = 1
    row = _write_title_area(ws, row)

    for section in SECTIONS:
        row = _write_section(ws, row, section)

    row = _write_faq(ws, row)

    # Footer
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(
        row=row, column=1,
        value="ご不明な点がございましたら、お気軽にお問い合わせください。",
    )
    cell.font = _HINT_FONT
    cell.alignment = Alignment(horizontal="center")

    # Print settings
    ws.print_area = f"A1:D{row}"
    ws.freeze_panes = "A5"

    # Save
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate PAD procedure guide Excel",
    )
    parser.add_argument(
        "--output", "-o",
        default=os.path.join(
            os.path.dirname(__file__),
            f"PAD操作手順書_{datetime.now().strftime('%Y%m%d')}.xlsx",
        ),
        help="Output Excel file path",
    )
    args = parser.parse_args()

    path = generate_pad_guide(args.output)
    print(f"Generated: {path}")


if __name__ == "__main__":
    main()
