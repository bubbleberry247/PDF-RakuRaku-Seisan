"""
phase3_excel.py - Generate PDD Excel workbook from event_log.

Phase 3 of Video2PDD v2 pipeline:
4 sheets:
  1. 概要フロー (Flow Summary) - High-level phase grouping
  2. 詳細手順 (Detailed Steps / PDD Table B) - Step-by-step actions
  3. 未解決項目 (Unresolved Items) - Items needing human review
  4. 実行メタデータ (Run Metadata) - Pipeline run information
"""
from __future__ import annotations

import os
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .event_log import (
    PHASE_EXCEL_GEN,
    is_phase_completed,
    is_video_pipeline,
    mark_phase_completed,
    mark_phase_failed,
    mark_phase_started,
)


# ------------------------------------------------------------------ #
#  Style constants
# ------------------------------------------------------------------ #

_HEADER_FONT = Font(name="Yu Gothic UI", bold=True, size=10)
_HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_BODY_FONT = Font(name="Yu Gothic UI", size=10)
_BODY_ALIGN = Alignment(vertical="top", wrap_text=True)

_PROVISIONAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_LOW_CONF_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_MERGED_FONT = Font(name="Yu Gothic UI", size=10, color="999999")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ------------------------------------------------------------------ #
#  Helpers
# ------------------------------------------------------------------ #

def _apply_header_style(ws, row: int, col_count: int) -> None:
    """Apply header style to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _apply_body_style(ws, row: int, col_count: int) -> None:
    """Apply body style to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _BODY_FONT
        cell.alignment = _BODY_ALIGN
        cell.border = _THIN_BORDER


def _auto_column_width(ws, col: int, min_width: int = 8, max_width: int = 60) -> None:
    """Auto-adjust column width based on content."""
    max_len = min_width
    col_letter = get_column_letter(col)
    for row in ws.iter_rows(min_col=col, max_col=col):
        for cell in row:
            if cell.value:
                # Approximate: CJK characters count as ~2 width
                val = str(cell.value)
                char_len = sum(2 if ord(c) > 0x7F else 1 for c in val)
                lines = val.split("\n")
                longest_line = max(
                    (sum(2 if ord(c) > 0x7F else 1 for c in line) for line in lines),
                    default=0,
                )
                max_len = max(max_len, min(longest_line, max_width))
    ws.column_dimensions[col_letter].width = max_len + 2


# ------------------------------------------------------------------ #
#  Sheet 1: Overview (Flow Summary)
# ------------------------------------------------------------------ #

def _write_overview_sheet(ws, event_log: dict) -> None:
    """Write Sheet 1: 概要フロー."""
    ws.title = "概要フロー"

    headers = ["フェーズ", "概要", "ステップ範囲", "対象ウィンドウ"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    phases = event_log.get("flow_phases", [])
    for i, phase in enumerate(phases):
        row = i + 2
        sr = phase.get("step_range", [0, 0])
        step_range_str = f"Step {sr[0]}" if sr[0] == sr[1] else f"Step {sr[0]}-{sr[1]}"

        ws.cell(row=row, column=1, value=phase.get("phase_id", i + 1))
        ws.cell(row=row, column=2, value=phase.get("summary", ""))
        ws.cell(row=row, column=3, value=step_range_str)
        ws.cell(row=row, column=4, value=phase.get("window_context", ""))
        _apply_body_style(ws, row, len(headers))

    # Column widths
    widths = [8, 40, 14, 30]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ------------------------------------------------------------------ #
#  Sheet 2: Detailed Steps (PDD Table B)
# ------------------------------------------------------------------ #

def _write_steps_sheet(ws, event_log: dict) -> None:
    """Write Sheet 2: 詳細手順."""
    ws.title = "詳細手順"

    is_video = is_video_pipeline(event_log)

    if is_video:
        headers = [
            "No", "時刻", "操作内容", "対象ウィンドウ", "対象要素",
            "入力値", "URL", "信頼度", "備考", "ステータス",
        ]
        widths = [5, 8, 50, 25, 25, 20, 35, 8, 35, 12]
    else:
        headers = [
            "No", "操作内容", "対象ウィンドウ", "対象要素",
            "入力値", "URL", "スクリーンショット", "備考", "ステータス",
        ]
        widths = [5, 50, 25, 25, 20, 35, 15, 35, 12]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    steps = event_log.get("steps", [])
    for i, step in enumerate(steps):
        row = i + 2
        is_merged = bool(step.get("merged_into"))
        is_provisional = step.get("verification_status") == "provisional"
        is_pending = step.get("verification_status") == "pending_review"
        confidence = step.get("confidence")

        # Build notes
        notes_parts: list[str] = []
        if step.get("duplicate_count", 0) > 1:
            notes_parts.append(f"同一操作 x{step['duplicate_count']}")
        if is_merged:
            notes_parts.append(f"-> Step {step['merged_into']}に統合")
        if step.get("gap_notes"):
            notes_parts.append(step["gap_notes"])
        notes = "; ".join(notes_parts)

        if is_video:
            ws.cell(row=row, column=1, value=step["num"])
            ws.cell(row=row, column=2, value=step.get("timestamp", ""))
            ws.cell(row=row, column=3, value=step.get("description", ""))
            ws.cell(row=row, column=4, value=step.get("window_alias", ""))
            ws.cell(row=row, column=5, value=step.get("element_alias", ""))
            ws.cell(row=row, column=6, value=step.get("input_value") or "")
            ws.cell(row=row, column=7, value=step.get("url") or "")
            ws.cell(row=row, column=8, value=f"{confidence:.2f}" if confidence else "")
            ws.cell(row=row, column=9, value=notes)
            ws.cell(row=row, column=10, value=step.get("verification_status", ""))
        else:
            ws.cell(row=row, column=1, value=step["num"])
            ws.cell(row=row, column=2, value=step.get("description", ""))
            ws.cell(row=row, column=3, value=step.get("window_alias", ""))
            ws.cell(row=row, column=4, value=step.get("element_alias", ""))
            ws.cell(row=row, column=5, value=step.get("input_value") or "")
            ws.cell(row=row, column=6, value=step.get("url") or "")
            ws.cell(row=row, column=7, value="")  # Screenshot placeholder
            ws.cell(row=row, column=8, value=notes)
            ws.cell(row=row, column=9, value=step.get("verification_status", ""))

        _apply_body_style(ws, row, len(headers))

        # Highlight based on confidence/status
        if is_video and confidence is not None and confidence < 0.7 and not is_merged:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = _LOW_CONF_FILL
        elif (is_provisional or is_pending) and not is_merged:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = _PROVISIONAL_FILL

        # Gray out merged rows
        if is_merged:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).font = _MERGED_FONT

    # Column widths
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"


# ------------------------------------------------------------------ #
#  Sheet 3: Unresolved Items
# ------------------------------------------------------------------ #

def _write_unresolved_sheet(ws, event_log: dict) -> None:
    """Write Sheet 3: 未解決項目."""
    ws.title = "未解決項目"

    headers = [
        "ID", "ステップ", "カテゴリ", "説明",
        "暫定値", "作成日時", "解決日時", "解決方法",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    items = event_log.get("unresolved_items", [])
    for i, item in enumerate(items):
        row = i + 2
        ws.cell(row=row, column=1, value=item.get("item_id", ""))
        ws.cell(row=row, column=2, value=item.get("step_num", ""))
        ws.cell(row=row, column=3, value=item.get("category", ""))
        ws.cell(row=row, column=4, value=item.get("description", ""))
        ws.cell(row=row, column=5, value=item.get("provisional_value") or "")
        ws.cell(row=row, column=6, value=item.get("created_at", ""))
        ws.cell(row=row, column=7, value=item.get("resolved_at") or "")
        ws.cell(row=row, column=8, value=item.get("resolution") or "")

        _apply_body_style(ws, row, len(headers))

        # Highlight unresolved (no resolved_at)
        if not item.get("resolved_at"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = _PROVISIONAL_FILL

    # Column widths
    widths = [10, 8, 18, 40, 30, 22, 22, 30]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"


# ------------------------------------------------------------------ #
#  Sheet 4: Run Metadata
# ------------------------------------------------------------------ #

def _write_metadata_sheet(ws, event_log: dict) -> None:
    """Write Sheet 4: 実行メタデータ."""
    ws.title = "実行メタデータ"

    headers = ["項目", "値"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    rm = event_log.get("run_metadata", {})
    inp = rm.get("input_files", {})
    ps = event_log.get("phase_status", {})
    rs = event_log.get("review_status", {})

    rows_data: list[tuple[str, str]] = [
        ("実行ID", rm.get("run_id", "")),
        ("作成日時", rm.get("created_at", "")),
        ("ツールバージョン", rm.get("tool_version", "")),
        ("実行ユーザー", rm.get("user", "")),
        ("ホスト名", rm.get("hostname", "")),
        ("出力ディレクトリ", rm.get("output_dir", "")),
        ("", ""),
        ("--- 入力ファイル ---", ""),
        ("動画ファイル", inp.get("video_file") or "なし"),
        ("動画ファイル SHA-256", inp.get("video_file_sha256") or ""),
        ("音声あり", str(inp.get("with_audio", False))),
        ("Robin Script", inp.get("robin_script") or "なし"),
        ("Robin Script SHA-256", inp.get("robin_script_sha256") or ""),
        ("ControlRepository", inp.get("control_repository") or "なし"),
        ("OBS動画", inp.get("obs_video") or "なし"),
        ("", ""),
        ("--- パイプライン状態 ---", ""),
    ]

    for phase_name, status in ps.items():
        completed = "完了" if status.get("completed") else "未完了"
        at = status.get("completed_at") or ""
        error = status.get("error") or ""
        value = completed
        if at:
            value += f" ({at})"
        if error:
            value += f" [ERROR: {error}]"
        rows_data.append((phase_name, value))

    rows_data.extend([
        ("", ""),
        ("--- レビュー状態 ---", ""),
        ("状態", rs.get("state", "")),
        ("未解決件数", str(rs.get("unresolved_count", 0))),
        ("通知日時", rs.get("notified_at") or "未通知"),
        ("", ""),
        ("--- 統計 ---", ""),
        ("総ステップ数", str(len(event_log.get("steps", [])))),
        ("確定ステップ", str(sum(
            1 for s in event_log.get("steps", [])
            if s.get("verification_status") == "confirmed"
        ))),
        ("暫定ステップ", str(sum(
            1 for s in event_log.get("steps", [])
            if s.get("verification_status") == "provisional"
        ))),
        ("重複統合", str(sum(
            1 for s in event_log.get("steps", [])
            if s.get("merged_into")
        ))),
        ("フロー フェーズ数", str(len(event_log.get("flow_phases", [])))),
    ])

    for i, (key, value) in enumerate(rows_data):
        row = i + 2
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)

        if key.startswith("---"):
            ws.cell(row=row, column=1).font = Font(
                name="Yu Gothic UI", bold=True, size=10,
            )
        else:
            _apply_body_style(ws, row, 2)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 70


# ------------------------------------------------------------------ #
#  Main Excel writer
# ------------------------------------------------------------------ #

def write_excel(event_log: dict, output_dir: str) -> str:
    """Generate PDD Excel workbook from event_log.

    Returns:
        Path to the generated Excel file.
    """
    wb = Workbook()

    # Sheet 1: Overview
    ws1 = wb.active
    _write_overview_sheet(ws1, event_log)

    # Sheet 2: Detailed Steps
    ws2 = wb.create_sheet()
    _write_steps_sheet(ws2, event_log)

    # Sheet 3: Unresolved Items
    ws3 = wb.create_sheet()
    _write_unresolved_sheet(ws3, event_log)

    # Sheet 4: Metadata
    ws4 = wb.create_sheet()
    _write_metadata_sheet(ws4, event_log)

    # Save
    os.makedirs(output_dir, exist_ok=True)
    run_id = event_log.get("run_metadata", {}).get("run_id", "output")
    filename = f"PDD_{run_id}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# ------------------------------------------------------------------ #
#  Phase runner
# ------------------------------------------------------------------ #

def run_phase3(event_log: dict, output_dir: str) -> dict[str, Any]:
    """Execute Phase 3: Excel generation.

    Returns:
        Summary dict with output path and sheet info.
    """
    if is_phase_completed(event_log, PHASE_EXCEL_GEN):
        return {"excel_path": "", "sheets": 4}

    mark_phase_started(event_log, PHASE_EXCEL_GEN)

    try:
        excel_path = write_excel(event_log, output_dir)
        mark_phase_completed(event_log, PHASE_EXCEL_GEN)

        steps = event_log.get("steps", [])
        unresolved = [
            u for u in event_log.get("unresolved_items", [])
            if u.get("resolved_at") is None
        ]

        return {
            "excel_path": excel_path,
            "sheets": 4,
            "total_steps": len(steps),
            "unresolved_count": len(unresolved),
        }

    except Exception as e:
        mark_phase_failed(event_log, PHASE_EXCEL_GEN, str(e))
        raise
