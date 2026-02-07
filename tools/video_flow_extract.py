"""
video_flow_extract.py - Screen-recording video (.mp4) -> business flow draft (Markdown + Excel).

Design goals (TPS):
- Pokayoke: validate inputs and fail early with clear errors.
- Jidoka: if OCR cannot determine key fields, mark steps as provisional + add unresolved items.

Notes:
- These videos often have no meaningful audio. This tool is OCR-based.
- Output is intended as a draft. Human review is required for 100% accuracy.

Usage:
  python -m tools.video_flow_extract --video-file "C:\\path\\to\\video.mp4"
  python -m tools.video_flow_extract --video-file "C:\\path\\to\\video.mp4" --output-dir "C:\\path\\to\\out"
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import secrets
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import numpy as np
from PIL import Image


TOOL_VERSION = "video-flow-extract-v1.0.0"


# --------------------------------------------------------------------------- #
#  Types
# --------------------------------------------------------------------------- #


@dataclass(frozen=True)
class VideoInfo:
    duration_sec: float
    size_bytes: int
    width: int
    height: int
    fps: float
    has_audio: bool


@dataclass(frozen=True)
class KeyFrame:
    index: int
    time_sec: float
    path: Path


# --------------------------------------------------------------------------- #
#  Time / IDs
# --------------------------------------------------------------------------- #


def _now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat()


def _generate_run_id() -> str:
    now = datetime.now()
    rand = secrets.token_hex(3)
    return f"{now.strftime('%Y%m%d_%H%M%S')}_{rand}"


def _format_ts(t: float) -> str:
    t = max(0.0, t)
    hh = int(t // 3600)
    mm = int((t % 3600) // 60)
    ss = int(t % 60)
    return f"{hh:02d}:{mm:02d}:{ss:02d}"


# --------------------------------------------------------------------------- #
#  IO helpers
# --------------------------------------------------------------------------- #


def _safe_slug(text: str, max_len: int = 80) -> str:
    # Keep Japanese characters; replace only Windows-forbidden characters.
    text = re.sub(r'[\\\\/:*?"<>|]+', "_", text)
    text = re.sub(r"\s+", "_", text).strip("_")
    if len(text) > max_len:
        return text[:max_len].rstrip("_")
    return text


def _sha256_file(path: Path, max_bytes: int = 10 * 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        h.update(f.read(max_bytes))
    return h.hexdigest()


def _run(cmd: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )


# --------------------------------------------------------------------------- #
#  FFprobe / FFmpeg
# --------------------------------------------------------------------------- #


def probe_video(video_file: Path) -> VideoInfo:
    cmd = [
        "ffprobe",
        "-hide_banner",
        "-v",
        "error",
        "-show_entries",
        "format=duration,size:stream=index,codec_type,codec_name,width,height,avg_frame_rate,channels,sample_rate",
        "-of",
        "json",
        str(video_file),
    ]
    out = _run(cmd).stdout
    data = json.loads(out)

    duration_sec = float(data["format"]["duration"])
    size_bytes = int(data["format"]["size"])

    width = 0
    height = 0
    fps = 0.0
    has_audio = False

    for s in data.get("streams", []):
        if s.get("codec_type") == "video":
            width = int(s.get("width") or 0)
            height = int(s.get("height") or 0)
            avg = (s.get("avg_frame_rate") or "0/0").split("/", 1)
            if len(avg) == 2 and avg[1] not in ("0", "0.0"):
                fps = float(avg[0]) / float(avg[1])
        if s.get("codec_type") == "audio":
            has_audio = True

    return VideoInfo(
        duration_sec=duration_sec,
        size_bytes=size_bytes,
        width=width,
        height=height,
        fps=fps,
        has_audio=has_audio,
    )


def extract_frames(
    video_file: Path,
    frames_dir: Path,
    interval_sec: float,
    scale_width: int,
) -> None:
    frames_dir.mkdir(parents=True, exist_ok=True)
    out_pattern = str(frames_dir / "frame_%06d.jpg")

    # fps=1/interval, scale to a fixed width (keeps aspect ratio).
    vf = f"fps=1/{interval_sec},scale={scale_width}:-1"
    cmd = [
        "ffmpeg",
        "-hide_banner",
        "-loglevel",
        "error",
        "-y",
        "-i",
        str(video_file),
        "-vf",
        vf,
        "-q:v",
        "4",
        out_pattern,
    ]
    _run(cmd)


# --------------------------------------------------------------------------- #
#  Hashing (for keyframe selection)
# --------------------------------------------------------------------------- #


def _ahash(img: Image.Image, hash_size: int = 8) -> int:
    g = img.convert("L").resize((hash_size, hash_size), Image.BILINEAR)
    pixels = np.asarray(g, dtype=np.float32)
    avg = float(pixels.mean())
    bits = (pixels > avg).astype(np.uint8).flatten()
    h = 0
    for bit in bits:
        h = (h << 1) | int(bit)
    return h


def _hamming(a: int, b: int) -> int:
    return int((a ^ b).bit_count())


def _crop_box(w: int, h: int, kind: str) -> tuple[int, int, int, int]:
    # Relative crops tuned for typical desktop recordings.
    if kind == "top":
        return (0, int(h * 0.05), int(w * 0.80), int(h * 0.25))
    if kind == "center":
        return (int(w * 0.25), int(h * 0.20), int(w * 0.75), int(h * 0.80))
    if kind == "bottom":
        return (0, int(h * 0.78), w, int(h * 0.97))
    raise ValueError(f"Unknown crop kind: {kind}")


def _signature(img: Image.Image) -> tuple[int, int, int]:
    w, h = img.size
    top = _ahash(img.crop(_crop_box(w, h, "top")))
    center = _ahash(img.crop(_crop_box(w, h, "center")))
    bottom = _ahash(img.crop(_crop_box(w, h, "bottom")))
    return (top, center, bottom)


def _signature_distance(a: tuple[int, int, int], b: tuple[int, int, int]) -> int:
    return _hamming(a[0], b[0]) + _hamming(a[1], b[1]) + _hamming(a[2], b[2])


def pick_keyframes(
    frame_paths: list[Path],
    interval_sec: float,
    dist_threshold: int,
    min_step_gap_sec: float,
) -> list[KeyFrame]:
    keyframes: list[KeyFrame] = []

    last_sig: tuple[int, int, int] | None = None
    last_kept_t: float = -1e9

    for i, fp in enumerate(frame_paths, start=1):
        t = (i - 1) * interval_sec
        img = Image.open(fp).convert("RGB")
        sig = _signature(img)

        if last_sig is None:
            keyframes.append(KeyFrame(index=i, time_sec=t, path=fp))
            last_sig = sig
            last_kept_t = t
            continue

        dist = _signature_distance(last_sig, sig)
        if dist >= dist_threshold and (t - last_kept_t) >= min_step_gap_sec:
            keyframes.append(KeyFrame(index=i, time_sec=t, path=fp))
            last_sig = sig
            last_kept_t = t

    # Always keep the last frame for end-state.
    if frame_paths:
        last_idx = len(frame_paths)
        last_t = (last_idx - 1) * interval_sec
        last_fp = frame_paths[-1]
        if not keyframes or keyframes[-1].path != last_fp:
            keyframes.append(KeyFrame(index=last_idx, time_sec=last_t, path=last_fp))

    return keyframes


# --------------------------------------------------------------------------- #
#  OCR + parsing
# --------------------------------------------------------------------------- #


def _init_easyocr_reader() -> Any:
    import easyocr  # lazy import (heavy)

    # Japanese + English for mixed UI.
    return easyocr.Reader(["ja", "en"], gpu=False)


def _ocr_lines(reader: Any, img: Image.Image) -> list[str]:
    arr = np.asarray(img.convert("RGB"))
    lines = reader.readtext(arr, detail=0, paragraph=False)
    # Normalize whitespace
    out: list[str] = []
    for s in lines:
        s2 = re.sub(r"\s+", " ", str(s)).strip()
        if s2:
            out.append(s2)
    return out


_RE_COUNT = re.compile(r"(\d{1,6})\s*件中\s*(\d{1,6})\s*(件|行)?")
_RE_DATE = re.compile(r"(20\d{2})[/-](\d{1,2})[/-](\d{1,2})")
_RE_YEN = re.compile(r"(\d{1,3}(?:,\d{3})+|\d+)\s*円")


def _extract_selected_count(lines: Iterable[str]) -> tuple[int, int] | None:
    text = " ".join(lines)
    m = _RE_COUNT.search(text)
    if not m:
        return None
    total = int(m.group(1))
    selected = int(m.group(2))
    return (total, selected)


def _extract_first_date(lines: Iterable[str]) -> str | None:
    text = " ".join(lines)
    m = _RE_DATE.search(text)
    if not m:
        return None
    yyyy, mm, dd = m.group(1), int(m.group(2)), int(m.group(3))
    return f"{yyyy}/{mm:02d}/{dd:02d}"


def _extract_first_amount_yen(lines: Iterable[str]) -> str | None:
    text = " ".join(lines)
    m = _RE_YEN.search(text)
    if not m:
        return None
    return f"{m.group(1)}円"


def _extract_title(lines: list[str]) -> str | None:
    # Prefer known keywords first.
    keywords = [
        "支払確定",
        "ログイン",
        "法定調書",
        "請求書",
        "集計",
        "楽楽精算",
    ]
    best: tuple[int, str] | None = None
    for s in lines:
        score = 0
        for kw in keywords:
            if kw in s:
                score += 10
        if "(" in s or "（" in s:
            score += 3
        # Penalize too-long lines (URLs etc)
        if len(s) > 60:
            score -= 8
        # Prefer lines with Japanese characters
        if re.search(r"[\u3040-\u30FF\u4E00-\u9FFF]", s):
            score += 2
        if best is None or score > best[0]:
            best = (score, s)
    if best and best[0] >= 5:
        return best[1]
    # Fallback: first Japanese-ish line under 40 chars
    for s in lines:
        if len(s) <= 40 and re.search(r"[\u3040-\u30FF\u4E00-\u9FFF]", s):
            return s
    return None


def _contains_any(lines: Iterable[str], needles: list[str]) -> bool:
    t = " ".join(lines)
    return any(n in t for n in needles)


def _guess_window_alias(title: str | None, modal: bool) -> str:
    if modal:
        return "確認ダイアログ"
    if title:
        return title
    return "(要確認: 画面名不明)"


def _pick_bank_line(lines: list[str]) -> str | None:
    for s in lines:
        if "銀行" in s:
            return s
    # Some OCR results miss '銀行' but include UFJ/三菱 etc
    for s in lines:
        if any(k in s for k in ["UFJ", "三菱", "みずほ", "三井住友"]):
            return s
    return None


def _build_description(
    window_alias: str,
    prev_window_alias: str | None,
    selected_count: tuple[int, int] | None,
    prev_selected_count: tuple[int, int] | None,
    payment_date: str | None,
    prev_payment_date: str | None,
    amount_yen: str | None,
    bank_line: str | None,
    modal: bool,
) -> tuple[str, str, str | None]:
    """Return (description, element_alias, input_value)."""
    prefix = ""
    if prev_window_alias and window_alias != prev_window_alias and not modal:
        prefix = f"{window_alias}を開き、"

    if modal:
        parts: list[str] = []
        if bank_line:
            parts.append(f"振込元={bank_line}")
        if payment_date:
            parts.append(f"支払日={payment_date}")
        if amount_yen:
            parts.append(f"合計={amount_yen}")
        detail = " / ".join(parts)
        if detail:
            return (
                f"確認ダイアログで内容を確認し、OKを押す（{detail}）",
                "OKボタン",
                None,
            )
        return ("確認ダイアログで内容を確認し、OKを押す", "OKボタン", None)

    if selected_count and prev_selected_count and selected_count != prev_selected_count:
        total, sel = selected_count
        prev_sel = prev_selected_count[1]
        return (
            f"{prefix}支払対象のチェックを変更し、選択件数を {prev_sel} → {sel} 件にする（全{total}件）",
            "チェックボックス",
            None,
        )
    if selected_count and not prev_selected_count:
        total, sel = selected_count
        if "支払確定" in window_alias:
            return (
                f"{prefix}支払確定一覧を表示し、支払対象を選択する（選択{sel}件/全{total}件）",
                "チェックボックス",
                None,
            )
        return (f"{prefix}一覧を表示する（選択{sel}件/全{total}件）", "チェックボックス", None)

    if payment_date and (prev_payment_date is None or payment_date != prev_payment_date):
        return (f"{prefix}支払日を {payment_date} に設定する", "支払日", payment_date)

    return (f"{prefix}画面上で操作を進める", "", None)


def _normalize_window_alias(alias: str, evidence_lines: list[str]) -> str:
    """Fix common OCR noise for known pages."""
    if alias.startswith("(要確認"):
        if _contains_any(evidence_lines, ["ログインID", "パスワード", "ログイン"]):
            return "楽楽精算ログイン"
    if alias:
        alias = alias.replace("黒楽精算", "楽楽精算")
    return alias


# --------------------------------------------------------------------------- #
#  Output
# --------------------------------------------------------------------------- #


def _write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _write_markdown(
    path: Path,
    video_file: Path,
    info: VideoInfo,
    settings: dict[str, Any],
    phases: list[dict[str, Any]],
    steps: list[dict[str, Any]],
    unresolved_items: list[dict[str, Any]],
) -> None:
    lines: list[str] = []
    lines.append("# 業務フロー起こし（動画）\n")
    lines.append(f"- 動画: `{video_file}`")
    lines.append(f"- 長さ: {info.duration_sec:.1f} sec")
    lines.append(f"- 解像度: {info.width}x{info.height} / fps={info.fps:.2f}")
    lines.append(f"- 音声: {'あり' if info.has_audio else 'なし/無音の可能性'}")
    lines.append("")
    lines.append("## 抽出設定")
    for k, v in settings.items():
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("## 概要フロー")
    for ph in phases:
        sr = ph.get("step_range", [0, 0])
        sr_s = f"Step {sr[0]}" if sr[0] == sr[1] else f"Step {sr[0]}-{sr[1]}"
        lines.append(f"- Phase {ph.get('phase_id')}: {ph.get('summary')} ({sr_s})")
    lines.append("")
    lines.append("## 詳細手順")
    for s in steps:
        t0 = s.get("time_start", "")
        t1 = s.get("time_end", "")
        status = s.get("verification_status", "")
        lines.append(f"### Step {s.get('num')} ({t0} - {t1}) [{status}]")
        lines.append(f"- 操作: {s.get('description','')}")
        if s.get("window_alias"):
            lines.append(f"- 画面: {s.get('window_alias')}")
        if s.get("element_alias"):
            lines.append(f"- 対象: {s.get('element_alias')}")
        if s.get("input_value"):
            lines.append(f"- 入力/値: {s.get('input_value')}")
        if s.get("evidence"):
            ev = s["evidence"]
            if isinstance(ev, list) and ev:
                lines.append("- 根拠（OCR）:")
                for e in ev[:12]:
                    lines.append(f"  - {e}")
        if s.get("screenshot_path"):
            lines.append(f"- 画像: `{s.get('screenshot_path')}`")
        if s.get("gap_notes"):
            lines.append(f"- 要確認: {s.get('gap_notes')}")
        lines.append("")
    lines.append("## 未解決項目（要レビュー）")
    if not unresolved_items:
        lines.append("- なし")
    else:
        for u in unresolved_items:
            lines.append(
                f"- {u.get('item_id')} / Step {u.get('step_num')}: {u.get('description')}"
            )
    lines.append("")
    _write_text(path, "\n".join(lines))


def _write_excel(path: Path, event_log: dict[str, Any]) -> None:
    # Minimal Excel writer (Video2PDD-like format, with time columns).
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Yu Gothic UI", bold=True, size=10)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="Yu Gothic UI", size=10)
    body_align = Alignment(vertical="top", wrap_text=True)
    provisional_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def apply_header(ws, row: int, cols: int) -> None:
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def apply_body(ws, row: int, cols: int) -> None:
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border

    wb = Workbook()

    # Sheet 1: 概要フロー
    ws1 = wb.active
    ws1.title = "概要フロー"
    headers1 = ["フェーズ", "概要", "ステップ範囲", "対象ウィンドウ"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    apply_header(ws1, 1, len(headers1))
    for i, ph in enumerate(event_log.get("flow_phases", []), start=1):
        row = i + 1
        sr = ph.get("step_range", [0, 0])
        sr_s = f"Step {sr[0]}" if sr[0] == sr[1] else f"Step {sr[0]}-{sr[1]}"
        ws1.cell(row=row, column=1, value=ph.get("phase_id", i))
        ws1.cell(row=row, column=2, value=ph.get("summary", ""))
        ws1.cell(row=row, column=3, value=sr_s)
        ws1.cell(row=row, column=4, value=ph.get("window_context", ""))
        apply_body(ws1, row, len(headers1))
    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 40
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 30

    # Sheet 2: 詳細手順
    ws2 = wb.create_sheet(title="詳細手順")
    headers2 = [
        "No",
        "開始",
        "終了",
        "操作内容",
        "対象ウィンドウ",
        "対象要素",
        "入力値",
        "根拠（OCR抜粋）",
        "画像パス",
        "ステータス",
    ]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    apply_header(ws2, 1, len(headers2))
    for i, st in enumerate(event_log.get("steps", []), start=1):
        row = i + 1
        ws2.cell(row=row, column=1, value=st.get("num", str(i).zfill(2)))
        ws2.cell(row=row, column=2, value=st.get("time_start", ""))
        ws2.cell(row=row, column=3, value=st.get("time_end", ""))
        ws2.cell(row=row, column=4, value=st.get("description", ""))
        ws2.cell(row=row, column=5, value=st.get("window_alias", ""))
        ws2.cell(row=row, column=6, value=st.get("element_alias", ""))
        ws2.cell(row=row, column=7, value=st.get("input_value") or "")
        evidence = st.get("evidence") or []
        if isinstance(evidence, list):
            ws2.cell(row=row, column=8, value="\n".join(evidence[:8]))
        else:
            ws2.cell(row=row, column=8, value=str(evidence))
        ws2.cell(row=row, column=9, value=st.get("screenshot_path") or "")
        ws2.cell(row=row, column=10, value=st.get("verification_status", ""))
        apply_body(ws2, row, len(headers2))

        if st.get("verification_status") == "provisional":
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=row, column=c).fill = provisional_fill
    widths2 = [5, 10, 10, 50, 28, 18, 16, 35, 35, 12]
    for col, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"

    # Sheet 3: 未解決項目
    ws3 = wb.create_sheet(title="未解決項目")
    headers3 = ["ID", "ステップ", "カテゴリ", "説明", "暫定値", "作成日時"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    apply_header(ws3, 1, len(headers3))
    items = event_log.get("unresolved_items", [])
    for i, it in enumerate(items, start=1):
        row = i + 1
        ws3.cell(row=row, column=1, value=it.get("item_id", ""))
        ws3.cell(row=row, column=2, value=it.get("step_num", ""))
        ws3.cell(row=row, column=3, value=it.get("category", ""))
        ws3.cell(row=row, column=4, value=it.get("description", ""))
        ws3.cell(row=row, column=5, value=it.get("provisional_value") or "")
        ws3.cell(row=row, column=6, value=it.get("created_at", ""))
        apply_body(ws3, row, len(headers3))
        for c in range(1, len(headers3) + 1):
            ws3.cell(row=row, column=c).fill = provisional_fill
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 8
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 50
    ws3.column_dimensions["E"].width = 30
    ws3.column_dimensions["F"].width = 22
    ws3.freeze_panes = "A2"

    # Sheet 4: 実行メタデータ
    ws4 = wb.create_sheet(title="実行メタデータ")
    headers4 = ["項目", "値"]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    apply_header(ws4, 1, len(headers4))
    rm = event_log.get("run_metadata", {})
    inp = rm.get("input_files", {})
    rows = [
        ("実行ID", rm.get("run_id", "")),
        ("作成日時", rm.get("created_at", "")),
        ("ツールバージョン", rm.get("tool_version", "")),
        ("実行ユーザー", rm.get("user", "")),
        ("ホスト名", rm.get("hostname", "")),
        ("出力ディレクトリ", rm.get("output_dir", "")),
        ("", ""),
        ("--- 入力 ---", ""),
        ("動画", inp.get("video_file", "")),
        ("動画 SHA-256", inp.get("video_file_sha256", "")),
        ("", ""),
        ("--- 統計 ---", ""),
        ("総ステップ数", str(len(event_log.get("steps", [])))),
        (
            "暫定ステップ",
            str(
                sum(
                    1
                    for s in event_log.get("steps", [])
                    if s.get("verification_status") == "provisional"
                )
            ),
        ),
        ("フロー フェーズ数", str(len(event_log.get("flow_phases", [])))),
        ("未解決件数", str(len(event_log.get("unresolved_items", [])))),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        row = i + 1
        ws4.cell(row=row, column=1, value=k)
        ws4.cell(row=row, column=2, value=v)
        apply_body(ws4, row, 2)
        if k.startswith("---"):
            ws4.cell(row=row, column=1).font = Font(name="Yu Gothic UI", bold=True, size=10)
    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 80

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# --------------------------------------------------------------------------- #
#  Main pipeline
# --------------------------------------------------------------------------- #


def _list_frames(frames_dir: Path) -> list[Path]:
    return sorted(frames_dir.glob("frame_*.jpg"))


def build_event_log_from_video(
    video_file: Path,
    output_dir: Path,
    interval_sec: float,
    scale_width: int,
    dist_threshold: int,
    min_step_gap_sec: float,
) -> dict[str, Any]:
    info = probe_video(video_file)

    frames_dir = output_dir / "frames_sampled"
    extract_frames(video_file, frames_dir, interval_sec=interval_sec, scale_width=scale_width)
    frame_paths = _list_frames(frames_dir)
    if not frame_paths:
        raise RuntimeError("No frames were extracted. Check ffmpeg availability and input video.")

    keyframes = pick_keyframes(
        frame_paths,
        interval_sec=interval_sec,
        dist_threshold=dist_threshold,
        min_step_gap_sec=min_step_gap_sec,
    )

    steps_dir = output_dir / "frames_steps"
    steps_dir.mkdir(parents=True, exist_ok=True)

    reader = _init_easyocr_reader()

    steps: list[dict[str, Any]] = []
    unresolved_items: list[dict[str, Any]] = []

    prev_window: str | None = None
    prev_selected: tuple[int, int] | None = None
    prev_pay_date: str | None = None

    for si, kf in enumerate(keyframes, start=1):
        img = Image.open(kf.path).convert("RGB")
        w, h = img.size
        top_crop = img.crop(_crop_box(w, h, "top"))
        center_crop = img.crop(_crop_box(w, h, "center"))
        bottom_crop = img.crop(_crop_box(w, h, "bottom"))

        top_lines = _ocr_lines(reader, top_crop)
        center_lines = _ocr_lines(reader, center_crop)
        bottom_lines = _ocr_lines(reader, bottom_crop)

        modal = _contains_any(center_lines, ["確認", "よろしい", "キャンセル", "OK"])
        title = _extract_title(top_lines)
        window_alias = _guess_window_alias(title, modal=modal)
        window_alias = _normalize_window_alias(window_alias, top_lines + center_lines + bottom_lines)

        selected = _extract_selected_count(top_lines)
        pay_date = _extract_first_date(bottom_lines) or _extract_first_date(center_lines)
        amount = _extract_first_amount_yen(center_lines)
        bank_line = _pick_bank_line(center_lines)

        desc, element_alias, input_value = _build_description(
            window_alias=window_alias,
            prev_window_alias=prev_window,
            selected_count=selected,
            prev_selected_count=prev_selected,
            payment_date=pay_date,
            prev_payment_date=prev_pay_date,
            amount_yen=amount,
            bank_line=bank_line,
            modal=modal,
        )

        # Evidence (small, to keep outputs readable)
        evidence: list[str] = []
        evidence.extend([l for l in top_lines if len(l) <= 40][:6])
        evidence.extend([l for l in bottom_lines if len(l) <= 40][:4])
        evidence.extend([l for l in center_lines if len(l) <= 40][:6])

        verification = "confirmed"
        gap_notes = ""
        if window_alias.startswith("(要確認") or title is None:
            verification = "provisional"
            gap_notes = "画面名のOCR抽出が不十分。該当時刻の画面を目視で確認してください。"
        if modal and amount is None:
            verification = "provisional"
            gap_notes = (gap_notes + " / " if gap_notes else "") + "確認ダイアログの合計金額が読み取れない可能性。"

        step_num = str(si).zfill(2)
        shot_name = f"step_{step_num}_{_format_ts(kf.time_sec).replace(':','-')}.jpg"
        shot_path = steps_dir / shot_name
        if not shot_path.exists():
            # Copy (not hardlink) for portability.
            shot_path.write_bytes(kf.path.read_bytes())

        steps.append(
            {
                "num": step_num,
                "time_start": _format_ts(kf.time_sec),
                "time_end": "",  # filled later
                "raw_frame_index": kf.index,
                "raw_frame_path": str(kf.path),
                "description": desc,
                "window_alias": window_alias,
                "element_alias": element_alias,
                "input_value": input_value,
                "verification_status": verification,
                "gap_notes": gap_notes,
                "evidence": evidence,
                "screenshot_path": str(shot_path),
                "extracted": {
                    "title": title,
                    "selected_count": selected,
                    "payment_date": pay_date,
                    "amount_yen": amount,
                    "bank": bank_line,
                    "modal": modal,
                },
            }
        )

        if verification == "provisional":
            unresolved_items.append(
                {
                    "item_id": f"U{secrets.token_hex(3).upper()}",
                    "step_num": step_num,
                    "category": "ocr_insufficient",
                    "description": gap_notes or "OCR抽出が不十分なため要レビュー。",
                    "provisional_value": None,
                    "created_at": _now_iso(),
                }
            )

        prev_window = window_alias
        prev_selected = selected or prev_selected
        prev_pay_date = pay_date or prev_pay_date

    # Fill time_end
    for i in range(len(steps)):
        if i + 1 < len(steps):
            steps[i]["time_end"] = steps[i + 1]["time_start"]
        else:
            steps[i]["time_end"] = _format_ts(info.duration_sec)

    # Build phases by window_alias grouping
    flow_phases: list[dict[str, Any]] = []
    if steps:
        start = 0
        phase_id = 1
        while start < len(steps):
            win = steps[start].get("window_alias", "")
            end = start
            while end + 1 < len(steps) and steps[end + 1].get("window_alias") == win:
                end += 1

            # Summary is rule-based.
            summary = win
            if "ログイン" in win:
                summary = "楽楽精算にログインする"
            elif "支払確定" in win:
                summary = "支払確定で支払対象を選択し、確定する"
            elif "確認ダイアログ" in win:
                summary = "確認ダイアログで内容を確認して確定する"

            flow_phases.append(
                {
                    "phase_id": phase_id,
                    "phase_name": f"phase_{phase_id}",
                    "step_range": [int(steps[start]["num"]), int(steps[end]["num"])],
                    "summary": summary,
                    "summary_source": "rule",
                    "window_context": win,
                    "unresolved": any(
                        s.get("verification_status") == "provisional"
                        for s in steps[start : end + 1]
                    ),
                }
            )
            phase_id += 1
            start = end + 1

    run_id = _generate_run_id()
    run_metadata = {
        "run_id": run_id,
        "created_at": _now_iso(),
        "tool_version": TOOL_VERSION,
        "input_files": {
            "video_file": str(video_file.resolve()),
            "video_file_sha256": _sha256_file(video_file),
        },
        "output_dir": str(output_dir.resolve()),
        "user": os.environ.get("USERNAME", os.environ.get("USER", "unknown")),
        "hostname": os.environ.get("COMPUTERNAME", ""),
        "video_info": {
            "duration_sec": info.duration_sec,
            "size_bytes": info.size_bytes,
            "width": info.width,
            "height": info.height,
            "fps": info.fps,
            "has_audio": info.has_audio,
        },
        "extract_settings": {
            "interval_sec": interval_sec,
            "scale_width": scale_width,
            "dist_threshold": dist_threshold,
            "min_step_gap_sec": min_step_gap_sec,
        },
    }

    return {
        "run_metadata": run_metadata,
        "steps": steps,
        "flow_phases": flow_phases,
        "unresolved_items": unresolved_items,
        "review_status": {
            "state": "pending_review" if unresolved_items else "reviewed",
            "unresolved_count": len(unresolved_items),
            "notified_at": None,
        },
    }


# --------------------------------------------------------------------------- #
#  CLI
# --------------------------------------------------------------------------- #


def _default_output_dir(video_file: Path) -> Path:
    repo_root = Path(__file__).resolve().parents[1]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = _safe_slug(video_file.stem)
    return repo_root / "_wip" / f"video_flow_extract_{ts}" / base


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

    p = argparse.ArgumentParser(description="Extract business flow draft from screen-recording video.")
    p.add_argument("--video-file", required=True, help="Path to .mp4 screen recording")
    p.add_argument("--output-dir", help="Output directory (default: repo/_wip/video_flow_extract_*/<video>)")
    p.add_argument("--interval-sec", type=float, default=1.0, help="Frame sampling interval in seconds")
    p.add_argument("--scale-width", type=int, default=1280, help="Extracted frame width")
    p.add_argument("--dist-threshold", type=int, default=18, help="AHash distance threshold for keyframe selection")
    p.add_argument("--min-step-gap-sec", type=float, default=3.0, help="Minimum seconds between keyframes")

    args = p.parse_args()

    video_file = Path(args.video_file)
    if not video_file.exists():
        raise SystemExit(f"ERROR: video not found: {video_file}")
    if video_file.stat().st_size == 0:
        raise SystemExit(f"ERROR: video is empty: {video_file}")

    output_dir = Path(args.output_dir) if args.output_dir else _default_output_dir(video_file)
    output_dir.mkdir(parents=True, exist_ok=True)

    info = probe_video(video_file)
    print("Video:")
    print(f"  path: {video_file}")
    print(f"  duration: {info.duration_sec:.1f} sec")
    print(f"  size: {info.size_bytes} bytes")
    print(f"  resolution: {info.width}x{info.height} fps={info.fps:.2f}")
    print(f"  audio: {'yes' if info.has_audio else 'no'}")
    print("Output:")
    print(f"  dir: {output_dir}")
    print()

    event_log = build_event_log_from_video(
        video_file=video_file,
        output_dir=output_dir,
        interval_sec=float(args.interval_sec),
        scale_width=int(args.scale_width),
        dist_threshold=int(args.dist_threshold),
        min_step_gap_sec=float(args.min_step_gap_sec),
    )

    # Save outputs
    event_log_path = output_dir / "event_log.json"
    md_path = output_dir / "flow.md"
    xlsx_path = output_dir / f"PDD_{event_log['run_metadata']['run_id']}.xlsx"

    _write_json(event_log_path, event_log)
    _write_markdown(
        md_path,
        video_file=video_file,
        info=info,
        settings=event_log["run_metadata"]["extract_settings"],
        phases=event_log.get("flow_phases", []),
        steps=event_log.get("steps", []),
        unresolved_items=event_log.get("unresolved_items", []),
    )
    _write_excel(xlsx_path, event_log)

    print("Done:")
    print(f"  event_log: {event_log_path}")
    print(f"  markdown:  {md_path}")
    print(f"  excel:     {xlsx_path}")
    print(
        f"  steps: {len(event_log.get('steps', []))}, unresolved: {len(event_log.get('unresolved_items', []))}"
    )


if __name__ == "__main__":
    main()
