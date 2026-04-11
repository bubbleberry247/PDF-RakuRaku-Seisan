from __future__ import annotations

import argparse
import csv
import re
import unicodedata
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable

import fitz
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


WARD_ORDER = (
    "千種区",
    "東区",
    "北区",
    "西区",
    "中村区",
    "中区",
    "昭和区",
    "瑞穂区",
    "熱田区",
    "中川区",
    "港区",
    "南区",
    "守山区",
    "緑区",
    "名東区",
    "天白区",
)

WARD_INDEX = {ward: index for index, ward in enumerate(WARD_ORDER)}

HTML_SOURCE_URL = "https://www.city.nagoya.jp/somu/page/0000137450.html"
WAITLIST_SOURCES = (
    ("SRC-2025-05", "2025-05-01", "令和7年5月1日現在", "https://www.kaigo-wel.city.nagoya.jp/_files/00140829/R0705_taiki.pdf"),
    ("SRC-2025-10", "2025-10-01", "令和7年10月1日現在", "https://www.kaigo-wel.city.nagoya.jp/_files/00147590/R710_taiki.pdf"),
)

DEFAULT_WORK_DIR = Path("work") / "nagoya_tokuyo"
DEFAULT_RAW_DIR = DEFAULT_WORK_DIR / "raw"
DEFAULT_XLSX_PATH = DEFAULT_WORK_DIR / "nagoya_tokuyo_ward_dashboard.xlsx"
DEFAULT_CSV_PATH = DEFAULT_WORK_DIR / "nagoya_tokuyo_ward_trend.csv"

HEADER_FILL = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
SUMMARY_FILL = PatternFill(fill_type="solid", start_color="EAF4E2", end_color="EAF4E2")
NOTE_FILL = PatternFill(fill_type="solid", start_color="F8F5D6", end_color="F8F5D6")
HEADER_FONT = Font(name="Meiryo", bold=True, size=10)
BODY_FONT = Font(name="Meiryo", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

HTML_TABLE_CAPTION = "特別養護老人ホーム(介護老人福祉施設及び地域密着型介護老人福祉施設)の待機者数等"
FULL_BED_RATE_NOTE = "今回利用した公表資料には満床率の列がなく、公開実績値は未収録"
PRESSURE_NOTE = "参考ひっ迫率 = 待機者数 ÷ 定員"


@dataclass(frozen=True)
class SourceRef:
    source_id: str
    category: str
    title: str
    url: str
    published_date: str
    snapshot_date: str
    note: str


@dataclass(frozen=True)
class WardSnapshot:
    snapshot_date: str
    snapshot_label: str
    ward: str
    facility_count: int
    capacity: int
    applicants: int
    source_id: str

    @property
    def reference_pressure_ratio(self) -> float:
        return self.applicants / self.capacity


@dataclass(frozen=True)
class SnapshotTotal:
    snapshot_date: str
    snapshot_label: str
    facility_count: int
    capacity: int
    applicants: int
    source_id: str

    @property
    def reference_pressure_ratio(self) -> float:
        return self.applicants / self.capacity


class CaptionedTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[dict[str, object]] = []
        self._in_table = False
        self._in_caption = False
        self._in_row = False
        self._in_cell = False
        self._current_caption: list[str] = []
        self._current_rows: list[list[str]] = []
        self._current_row: list[str] = []
        self._current_cell: list[str] = []

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag == "table":
            self._in_table = True
            self._current_caption = []
            self._current_rows = []
        elif self._in_table and tag == "caption":
            self._in_caption = True
        elif self._in_table and tag == "tr":
            self._in_row = True
            self._current_row = []
        elif self._in_row and tag in {"td", "th"}:
            self._in_cell = True
            self._current_cell = []
        elif self._in_cell and tag == "br":
            self._current_cell.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag == "caption" and self._in_caption:
            self._in_caption = False
        elif tag in {"td", "th"} and self._in_cell:
            self._in_cell = False
            self._current_row.append(clean_text("".join(self._current_cell)))
        elif tag == "tr" and self._in_row:
            self._in_row = False
            if any(cell for cell in self._current_row):
                self._current_rows.append(self._current_row)
        elif tag == "table" and self._in_table:
            self._in_table = False
            self.tables.append(
                {
                    "caption": clean_text("".join(self._current_caption)),
                    "rows": self._current_rows,
                }
            )

    def handle_data(self, data: str) -> None:
        if self._in_caption:
            self._current_caption.append(data)
        elif self._in_cell:
            self._current_cell.append(data)


def clean_text(value: str | None) -> str:
    if not value:
        return ""
    normalized = unicodedata.normalize("NFKC", value).replace("\u3000", " ")
    normalized = normalized.replace("\r", "\n")
    normalized = re.sub(r"[ \t]+", " ", normalized)
    normalized = re.sub(r"\n+", "\n", normalized)
    return normalized.strip()


def parse_number(value: str | None) -> int:
    cleaned = clean_text(value).replace(",", "").replace(" ", "")
    if not cleaned:
        raise ValueError("Number cell is empty")
    return int(cleaned)


def add_table(ws, display_name: str) -> None:
    table = Table(displayName=display_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_header(ws, row_number: int, end_column: int, fill: PatternFill = HEADER_FILL) -> None:
    for column_index in range(1, end_column + 1):
        cell = ws.cell(row=row_number, column=column_index)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER


def style_body_row(ws, row_number: int, end_column: int) -> None:
    for column_index in range(1, end_column + 1):
        cell = ws.cell(row=row_number, column=column_index)
        cell.font = BODY_FONT
        cell.alignment = LEFT


def set_widths(ws, widths: dict[int, float]) -> None:
    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width


def download_text(url: str) -> str:
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read().decode("utf-8", errors="ignore")


def download_file(url: str, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=60) as response:
        output_path.write_bytes(response.read())
    return output_path


def parse_2020_html_snapshot() -> tuple[list[WardSnapshot], SnapshotTotal]:
    parser = CaptionedTableParser()
    parser.feed(download_text(HTML_SOURCE_URL))
    target_table = None
    for table in parser.tables:
        caption = str(table["caption"])
        if caption.startswith(HTML_TABLE_CAPTION):
            target_table = table
            break
    if target_table is None:
        raise RuntimeError("Failed to find Nagoya 2020 ward summary table")

    rows = target_table["rows"]
    ward_rows: list[WardSnapshot] = []
    total_row: SnapshotTotal | None = None

    for row in rows[1:]:
        if len(row) < 4:
            continue
        ward = clean_text(row[0])
        facility_count = parse_number(row[1])
        capacity = parse_number(row[2])
        applicants = parse_number(row[3])
        if ward == "名古屋市合計":
            total_row = SnapshotTotal(
                snapshot_date="2020-10-01",
                snapshot_label="2020年10月1日",
                facility_count=facility_count,
                capacity=capacity,
                applicants=applicants,
                source_id="SRC-2020-HTML",
            )
            continue
        ward_rows.append(
            WardSnapshot(
                snapshot_date="2020-10-01",
                snapshot_label="2020年10月1日",
                ward=ward,
                facility_count=facility_count,
                capacity=capacity,
                applicants=applicants,
                source_id="SRC-2020-HTML",
            )
        )

    if total_row is None:
        raise RuntimeError("Failed to find total row in Nagoya 2020 ward summary table")

    validate_ward_rows(ward_rows, total_row, "SRC-2020-HTML")
    return sort_ward_rows(ward_rows), total_row


def parse_waitlist_pdf(pdf_path: Path, source_id: str, snapshot_date: str, snapshot_label: str) -> tuple[list[WardSnapshot], SnapshotTotal]:
    if not pdf_path.exists():
        raise FileNotFoundError(pdf_path)

    aggregate: dict[str, dict[str, int]] = defaultdict(lambda: {"facility_count": 0, "capacity": 0, "applicants": 0})
    explicit_total: SnapshotTotal | None = None

    document = fitz.open(pdf_path)
    try:
        for page in document:
            tables = page.find_tables().tables
            if not tables:
                raise RuntimeError(f"No table found on page {page.number + 1} of {pdf_path.name}")
            rows = tables[0].extract()
            for row in rows[1:]:
                if len(row) < 4:
                    continue
                ward = clean_text(row[0])
                facility_name = clean_text(row[1])
                capacity_text = clean_text(row[2])
                applicants_text = clean_text(row[3])
                if not capacity_text or not applicants_text:
                    continue
                capacity = parse_number(capacity_text)
                applicants = parse_number(applicants_text)
                normalized_ward = ward.replace(" ", "")
                if normalized_ward == "合計":
                    explicit_total = SnapshotTotal(
                        snapshot_date=snapshot_date,
                        snapshot_label=snapshot_label,
                        facility_count=sum(item["facility_count"] for item in aggregate.values()),
                        capacity=capacity,
                        applicants=applicants,
                        source_id=source_id,
                    )
                    continue
                if not facility_name:
                    continue
                item = aggregate[ward]
                item["facility_count"] += 1
                item["capacity"] += capacity
                item["applicants"] += applicants
    finally:
        document.close()

    if explicit_total is None:
        raise RuntimeError(f"Failed to find total row in {pdf_path.name}")

    ward_rows = [
        WardSnapshot(
            snapshot_date=snapshot_date,
            snapshot_label=snapshot_label,
            ward=ward,
            facility_count=item["facility_count"],
            capacity=item["capacity"],
            applicants=item["applicants"],
            source_id=source_id,
        )
        for ward, item in aggregate.items()
    ]
    validate_ward_rows(ward_rows, explicit_total, source_id)
    return sort_ward_rows(ward_rows), explicit_total


def sort_ward_rows(rows: Iterable[WardSnapshot]) -> list[WardSnapshot]:
    return sorted(rows, key=lambda row: WARD_INDEX.get(row.ward, len(WARD_INDEX)))


def validate_ward_rows(ward_rows: list[WardSnapshot], total_row: SnapshotTotal, source_id: str) -> None:
    observed_wards = {row.ward for row in ward_rows}
    missing = [ward for ward in WARD_ORDER if ward not in observed_wards]
    unexpected = sorted(observed_wards - set(WARD_ORDER))
    if missing or unexpected:
        raise RuntimeError(f"{source_id} ward coverage mismatch: missing={missing}, unexpected={unexpected}")

    summed_facility_count = sum(row.facility_count for row in ward_rows)
    summed_capacity = sum(row.capacity for row in ward_rows)
    summed_applicants = sum(row.applicants for row in ward_rows)
    if summed_facility_count != total_row.facility_count:
        raise RuntimeError(
            f"{source_id} facility total mismatch: wards={summed_facility_count}, total={total_row.facility_count}"
        )
    if summed_capacity != total_row.capacity:
        raise RuntimeError(f"{source_id} capacity total mismatch: wards={summed_capacity}, total={total_row.capacity}")
    if summed_applicants != total_row.applicants:
        raise RuntimeError(f"{source_id} applicant total mismatch: wards={summed_applicants}, total={total_row.applicants}")


def build_sources() -> tuple[SourceRef, ...]:
    return (
        SourceRef(
            source_id="SRC-2020-HTML",
            category="html",
            title="介護サービス事業所数等（NAGOYAライフ）",
            url=HTML_SOURCE_URL,
            published_date="2025-10-17",
            snapshot_date="2020-10-01",
            note="ページ内の『特別養護老人ホームの待機者数等（令和2年10月1日現在）』区別表を使用",
        ),
        SourceRef(
            source_id="SRC-2025-05",
            category="pdf",
            title="特別養護老人ホーム待機者数（令和7年5月1日現在）",
            url="https://www.kaigo-wel.city.nagoya.jp/_files/00140829/R0705_taiki.pdf",
            published_date="2025-05",
            snapshot_date="2025-05-01",
            note="施設別PDFを区ごとに再集計",
        ),
        SourceRef(
            source_id="SRC-2025-10",
            category="pdf",
            title="特別養護老人ホーム待機者数（令和7年10月1日現在）",
            url="https://www.kaigo-wel.city.nagoya.jp/_files/00147590/R710_taiki.pdf",
            published_date="2025-10",
            snapshot_date="2025-10-01",
            note="施設別PDFを区ごとに再集計",
        ),
    )


def collect_snapshots(raw_dir: Path, skip_download: bool) -> tuple[list[WardSnapshot], list[SnapshotTotal]]:
    all_rows: list[WardSnapshot] = []
    totals: list[SnapshotTotal] = []

    html_rows, html_total = parse_2020_html_snapshot()
    all_rows.extend(html_rows)
    totals.append(html_total)

    raw_dir.mkdir(parents=True, exist_ok=True)
    for source_id, snapshot_date, snapshot_label, url in WAITLIST_SOURCES:
        pdf_path = raw_dir / Path(url).name
        if not skip_download or not pdf_path.exists():
            download_file(url, pdf_path)
        ward_rows, total_row = parse_waitlist_pdf(pdf_path, source_id, snapshot_date, snapshot_label)
        all_rows.extend(ward_rows)
        totals.append(total_row)

    return sorted(all_rows, key=lambda row: (row.snapshot_date, WARD_INDEX[row.ward])), sorted(
        totals, key=lambda row: row.snapshot_date
    )


def trend_records(rows: list[WardSnapshot]) -> list[dict[str, object]]:
    return [
        {
            "snapshot_date": row.snapshot_date,
            "snapshot_label": row.snapshot_label,
            "ward": row.ward,
            "facility_count": row.facility_count,
            "capacity": row.capacity,
            "applicants": row.applicants,
            "public_full_bed_rate": None,
            "public_full_bed_rate_note": FULL_BED_RATE_NOTE,
            "reference_pressure_ratio": row.reference_pressure_ratio,
            "reference_pressure_ratio_note": PRESSURE_NOTE,
            "source_id": row.source_id,
        }
        for row in rows
    ]


def total_records(totals: list[SnapshotTotal]) -> list[dict[str, object]]:
    return [
        {
            "snapshot_date": row.snapshot_date,
            "snapshot_label": row.snapshot_label,
            "facility_count": row.facility_count,
            "capacity": row.capacity,
            "applicants": row.applicants,
            "public_full_bed_rate": None,
            "public_full_bed_rate_note": FULL_BED_RATE_NOTE,
            "reference_pressure_ratio": row.reference_pressure_ratio,
            "reference_pressure_ratio_note": PRESSURE_NOTE,
            "source_id": row.source_id,
        }
        for row in totals
    ]


def write_trend_sheet(wb: Workbook, rows: list[dict[str, object]]) -> None:
    ws = wb.active
    ws.title = "区別推移"
    headers = [
        "時点",
        "時点ラベル",
        "区",
        "施設数",
        "定員",
        "待機者数",
        "公開満床率",
        "満床率注記",
        "参考ひっ迫率",
        "参考ひっ迫率注記",
        "source_id",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for row in rows:
        ws.append(
            [
                row["snapshot_date"],
                row["snapshot_label"],
                row["ward"],
                row["facility_count"],
                row["capacity"],
                row["applicants"],
                row["public_full_bed_rate"],
                row["public_full_bed_rate_note"],
                row["reference_pressure_ratio"],
                row["reference_pressure_ratio_note"],
                row["source_id"],
            ]
        )
        style_body_row(ws, ws.max_row, len(headers))
        ws.cell(row=ws.max_row, column=7).number_format = "0.0%"
        ws.cell(row=ws.max_row, column=9).number_format = "0.0%"

    ws.freeze_panes = "A2"
    set_widths(
        ws,
        {
            1: 14,
            2: 18,
            3: 10,
            4: 10,
            5: 10,
            6: 12,
            7: 12,
            8: 28,
            9: 14,
            10: 24,
            11: 14,
        },
    )
    add_table(ws, "NagoyaWardTrend")


def write_latest_sheet(wb: Workbook, rows: list[dict[str, object]]) -> None:
    latest_date = max(str(row["snapshot_date"]) for row in rows)
    latest_rows = [row for row in rows if row["snapshot_date"] == latest_date]
    ws = wb.create_sheet("最新区別一覧")
    headers = [
        "区",
        "施設数",
        "定員",
        "待機者数",
        "公開満床率",
        "参考ひっ迫率",
        "時点",
        "source_id",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers), SUMMARY_FILL)

    for row in latest_rows:
        ws.append(
            [
                row["ward"],
                row["facility_count"],
                row["capacity"],
                row["applicants"],
                row["public_full_bed_rate"],
                row["reference_pressure_ratio"],
                row["snapshot_label"],
                row["source_id"],
            ]
        )
        style_body_row(ws, ws.max_row, len(headers))
        ws.cell(row=ws.max_row, column=5).number_format = "0.0%"
        ws.cell(row=ws.max_row, column=6).number_format = "0.0%"

    ws.freeze_panes = "A2"
    set_widths(ws, {1: 10, 2: 10, 3: 10, 4: 12, 5: 12, 6: 14, 7: 18, 8: 14})
    add_table(ws, "NagoyaWardLatest")


def write_totals_sheet(wb: Workbook, totals: list[dict[str, object]]) -> None:
    ws = wb.create_sheet("市合計推移")
    headers = [
        "時点",
        "時点ラベル",
        "施設数",
        "定員",
        "待機者数",
        "公開満床率",
        "満床率注記",
        "参考ひっ迫率",
        "参考ひっ迫率注記",
        "source_id",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers), SUMMARY_FILL)

    for row in totals:
        ws.append(
            [
                row["snapshot_date"],
                row["snapshot_label"],
                row["facility_count"],
                row["capacity"],
                row["applicants"],
                row["public_full_bed_rate"],
                row["public_full_bed_rate_note"],
                row["reference_pressure_ratio"],
                row["reference_pressure_ratio_note"],
                row["source_id"],
            ]
        )
        style_body_row(ws, ws.max_row, len(headers))
        ws.cell(row=ws.max_row, column=6).number_format = "0.0%"
        ws.cell(row=ws.max_row, column=8).number_format = "0.0%"

    ws.freeze_panes = "A2"
    set_widths(ws, {1: 14, 2: 18, 3: 10, 4: 10, 5: 12, 6: 12, 7: 28, 8: 14, 9: 24, 10: 14})
    add_table(ws, "NagoyaCityTotals")


def write_charts_sheet(wb: Workbook) -> None:
    latest_ws = wb["最新区別一覧"]
    totals_ws = wb["市合計推移"]
    ws = wb.create_sheet("グラフ")

    latest_max_row = latest_ws.max_row
    totals_max_row = totals_ws.max_row

    applicants_chart = BarChart()
    applicants_chart.type = "bar"
    applicants_chart.title = "区別 待機者数（最新時点）"
    applicants_chart.y_axis.title = "区"
    applicants_chart.x_axis.title = "待機者数"
    applicants_chart.height = 9
    applicants_chart.width = 18
    applicants_chart.add_data(Reference(latest_ws, min_col=4, min_row=1, max_row=latest_max_row), titles_from_data=True)
    applicants_chart.set_categories(Reference(latest_ws, min_col=1, min_row=2, max_row=latest_max_row))
    ws.add_chart(applicants_chart, "A1")

    capacity_chart = BarChart()
    capacity_chart.type = "bar"
    capacity_chart.title = "区別 定員（最新時点）"
    capacity_chart.y_axis.title = "区"
    capacity_chart.x_axis.title = "定員"
    capacity_chart.height = 9
    capacity_chart.width = 18
    capacity_chart.add_data(Reference(latest_ws, min_col=3, min_row=1, max_row=latest_max_row), titles_from_data=True)
    capacity_chart.set_categories(Reference(latest_ws, min_col=1, min_row=2, max_row=latest_max_row))
    ws.add_chart(capacity_chart, "J1")

    pressure_chart = BarChart()
    pressure_chart.type = "bar"
    pressure_chart.title = "区別 参考ひっ迫率（最新時点）"
    pressure_chart.y_axis.title = "区"
    pressure_chart.x_axis.title = "比率"
    pressure_chart.height = 9
    pressure_chart.width = 18
    pressure_chart.add_data(Reference(latest_ws, min_col=6, min_row=1, max_row=latest_max_row), titles_from_data=True)
    pressure_chart.set_categories(Reference(latest_ws, min_col=1, min_row=2, max_row=latest_max_row))
    pressure_chart.x_axis.numFmt = "0.0%"
    pressure_chart.series[0].graphicalProperties.solidFill = "4F81BD"
    ws.add_chart(pressure_chart, "A20")

    total_chart = LineChart()
    total_chart.title = "名古屋市合計の推移"
    total_chart.y_axis.title = "人数"
    total_chart.x_axis.title = "時点"
    total_chart.height = 8
    total_chart.width = 18
    total_chart.add_data(Reference(totals_ws, min_col=4, min_row=1, max_col=5, max_row=totals_max_row), titles_from_data=True)
    total_chart.set_categories(Reference(totals_ws, min_col=2, min_row=2, max_row=totals_max_row))
    ws.add_chart(total_chart, "J20")

    set_widths(ws, {1: 2, 10: 2})


def write_sources_sheet(wb: Workbook, sources: tuple[SourceRef, ...]) -> None:
    ws = wb.create_sheet("ソース情報")
    headers = ["source_id", "category", "title", "url", "published_date", "snapshot_date", "note"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for source in sources:
        ws.append(
            [
                source.source_id,
                source.category,
                source.title,
                source.url,
                source.published_date,
                source.snapshot_date,
                source.note,
            ]
        )
        style_body_row(ws, ws.max_row, len(headers))
    ws.freeze_panes = "A2"
    set_widths(ws, {1: 14, 2: 10, 3: 38, 4: 72, 5: 14, 6: 14, 7: 28})
    add_table(ws, "NagoyaWardSources")


def write_notes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("注意事項")
    notes = (
        "このブックは、名古屋市の公開資料に掲載された『定員』と『待機者数』を区別に整理したものです。",
        "公開資料に満床率の列がないため、『公開満床率』は空欄です。",
        "代わりに、需要圧の参考値として『参考ひっ迫率 = 待機者数 ÷ 定員』を掲載しています。",
        "2020年10月1日は市公式HTMLの区別集計表、2025年5月1日・10月1日は施設別PDFを区ごとに再集計しています。",
        "PDF/HTMLの構造変更や総計不一致が起きた場合は、スクリプトが例外停止するようにしています。",
    )
    ws.append(["項目", "内容"])
    style_header(ws, 1, 2, NOTE_FILL)
    for index, note in enumerate(notes, start=1):
        ws.append([f"注意{index}", note])
        style_body_row(ws, ws.max_row, 2)
    ws.freeze_panes = "A2"
    set_widths(ws, {1: 12, 2: 90})
    add_table(ws, "NagoyaWardNotes")


def save_csv(csv_path: Path, rows: list[dict[str, object]]) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "snapshot_date",
        "snapshot_label",
        "ward",
        "facility_count",
        "capacity",
        "applicants",
        "public_full_bed_rate",
        "public_full_bed_rate_note",
        "reference_pressure_ratio",
        "reference_pressure_ratio_note",
        "source_id",
    ]
    with csv_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def build_workbook(xlsx_path: Path, trend_rows: list[dict[str, object]], totals: list[dict[str, object]], sources: tuple[SourceRef, ...]) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    write_trend_sheet(wb, trend_rows)
    write_latest_sheet(wb, trend_rows)
    write_totals_sheet(wb, totals)
    write_charts_sheet(wb)
    write_sources_sheet(wb, sources)
    write_notes_sheet(wb)
    wb.save(xlsx_path)


def verify_outputs(xlsx_path: Path, csv_path: Path, expected_row_count: int, expected_total_count: int) -> None:
    wb = load_workbook(xlsx_path, data_only=False)
    try:
        expected_sheets = {"区別推移", "最新区別一覧", "市合計推移", "グラフ", "ソース情報", "注意事項"}
        if set(wb.sheetnames) != expected_sheets:
            raise RuntimeError(f"Unexpected workbook sheets: {wb.sheetnames}")

        trend_ws = wb["区別推移"]
        latest_ws = wb["最新区別一覧"]
        totals_ws = wb["市合計推移"]
        if trend_ws.max_row - 1 != expected_row_count:
            raise RuntimeError(f"Unexpected trend row count: {trend_ws.max_row - 1}")
        if latest_ws.max_row - 1 != len(WARD_ORDER):
            raise RuntimeError(f"Unexpected latest row count: {latest_ws.max_row - 1}")
        if totals_ws.max_row - 1 != expected_total_count:
            raise RuntimeError(f"Unexpected totals row count: {totals_ws.max_row - 1}")
    finally:
        wb.close()

    with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        rows = list(reader)
    if len(rows) != expected_row_count:
        raise RuntimeError(f"Unexpected CSV row count: {len(rows)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build Nagoya ward-level Tokuyo trend workbook and CSV.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX_PATH, help="Output workbook path")
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV_PATH, help="Output CSV path")
    parser.add_argument("--raw-dir", type=Path, default=DEFAULT_RAW_DIR, help="Directory used for downloaded PDFs")
    parser.add_argument(
        "--skip-download",
        action="store_true",
        help="Reuse existing raw PDFs if present",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    sources = build_sources()
    ward_rows, total_rows = collect_snapshots(args.raw_dir, args.skip_download)
    trend_rows = trend_records(ward_rows)
    totals = total_records(total_rows)
    save_csv(args.csv, trend_rows)
    build_workbook(args.xlsx, trend_rows, totals, sources)
    verify_outputs(args.xlsx, args.csv, expected_row_count=len(trend_rows), expected_total_count=len(totals))
    print(args.xlsx.resolve())
    print(args.csv.resolve())


if __name__ == "__main__":
    main()
