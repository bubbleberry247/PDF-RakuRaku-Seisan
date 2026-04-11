from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


WORK_DIR = Path("work") / "aichi_tokuyo"
DEFAULT_XLSX_PATH = WORK_DIR / "aichi_tokuyo_beds_applicants_pressure.xlsx"
DEFAULT_CSV_PATH = WORK_DIR / "aichi_tokuyo_beds_applicants_pressure.csv"

HEADER_FILL = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
NOTE_FILL = PatternFill(fill_type="solid", start_color="F4F6D0", end_color="F4F6D0")
HEADER_FONT = Font(name="Meiryo", bold=True, size=10)
BODY_FONT = Font(name="Meiryo", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


@dataclass(frozen=True)
class SourceRef:
    source_id: str
    category: str
    title: str
    url: str
    published_date: str
    pdf_page: str
    note: str


@dataclass(frozen=True)
class CapacityPoint:
    year: int
    wide_area_capacity: int
    community_capacity: int
    total_capacity: int
    source_id: str


@dataclass(frozen=True)
class ApplicantPoint:
    year: int
    survey_date: str
    applicants: int
    target: str
    scope: str
    source_id: str


SOURCES = (
    SourceRef(
        source_id="CAP-2020",
        category="capacity",
        title="第8期愛知県高齢者福祉保健医療計画",
        url="https://www.pref.aichi.jp/uploaded/life/324349_1267188_misc.pdf",
        published_date="2021-03",
        pdf_page="65",
        note="介護老人福祉施設（特別養護老人ホーム）2020年度目標（必要入所定員総数）。",
    ),
    SourceRef(
        source_id="CAP-2021-2023",
        category="capacity",
        title="第8期愛知県高齢者福祉保健医療計画",
        url="https://www.pref.aichi.jp/uploaded/life/324349_1267188_misc.pdf",
        published_date="2021-03",
        pdf_page="71",
        note="介護老人福祉施設（特別養護老人ホーム）2021-2023年度の必要入所定員総数。",
    ),
    SourceRef(
        source_id="CAP-2024-2026",
        category="capacity",
        title="第9期愛知県高齢者福祉保健医療計画",
        url="https://www.pref.aichi.jp/uploaded/attachment/519749.pdf",
        published_date="2024-03",
        pdf_page="70",
        note="介護老人福祉施設（特別養護老人ホーム）2024-2026年度の必要入所定員総数。",
    ),
    SourceRef(
        source_id="APP-2020",
        category="applicant",
        title="特別養護老人ホームへの入所申込者（待機者）の調査結果",
        url="https://www.pref.aichi.jp/uploaded/life/299162_1116575_misc.pdf",
        published_date="2020-07-31",
        pdf_page="1",
        note="2020年4月1日時点。早期に入所を希望する要介護3-5の待機者数。",
    ),
    SourceRef(
        source_id="APP-2023",
        category="applicant",
        title="特別養護老人ホームへの入所申込者（待機者）の調査結果",
        url="https://www.pref.aichi.jp/uploaded/attachment/468968.pdf",
        published_date="2023-07-25",
        pdf_page="1",
        note="2023年4月1日時点。早期に入所を希望する要介護3-5の待機者数。",
    ),
    SourceRef(
        source_id="NOTE-2020-METHOD",
        category="note",
        title="特別養護老人ホームへの入所申込者（待機者）の調査結果",
        url="https://www.pref.aichi.jp/uploaded/life/299162_1116575_misc.pdf",
        published_date="2020-07-31",
        pdf_page="1",
        note="2020年度調査で死亡者・他施設入所済み者等を除外する精査方法に変更されたため、2017年度以前とは単純比較しない。",
    ),
)


CAPACITY_POINTS = (
    CapacityPoint(2020, 26281, 3890, 30171, "CAP-2020"),
    CapacityPoint(2021, 26026, 3794, 29820, "CAP-2021-2023"),
    CapacityPoint(2022, 26416, 3910, 30326, "CAP-2021-2023"),
    CapacityPoint(2023, 26656, 3968, 30624, "CAP-2021-2023"),
    CapacityPoint(2024, 26490, 3823, 30313, "CAP-2024-2026"),
    CapacityPoint(2025, 26576, 3968, 30544, "CAP-2024-2026"),
    CapacityPoint(2026, 26576, 3997, 30573, "CAP-2024-2026"),
)


APPLICANT_POINTS = (
    ApplicantPoint(
        year=2020,
        survey_date="2020-04-01",
        applicants=4467,
        target="要介護3-5",
        scope="早期入所希望の待機者数",
        source_id="APP-2020",
    ),
    ApplicantPoint(
        year=2023,
        survey_date="2023-04-01",
        applicants=3502,
        target="要介護3-5",
        scope="早期入所希望の待機者数",
        source_id="APP-2023",
    ),
)


NOTES = (
    "このデータセットは、前タスクに合わせて特別養護老人ホームを対象に整理している。",
    "特養は医療機関の病床ではないため、『病床数』は特養の定員相当として県計画の『必要入所定員総数』を採用している。",
    "病床ひっ迫率は県の公表指標ではなく、参考算出値。算式は『希望者数 ÷ 必要入所定員総数』。",
    "希望者数は3年ごとの待機者調査のみ使用。2020年度調査で精査方法が変更されたため、2017年度以前の数値は連続比較から除外した。",
    "希望者数の基準日と定員の基準年度は完全一致ではないため、ひっ迫率は厳密な同日比率ではなく、計画定員に対する需給圧力の近似値として扱う。",
    "参考予想は公式値ではなく、2020年と2023年の希望者数の差を年平均に直して2024-2026年度へ線形外挿した簡易シナリオ。",
)


FORECAST_METHOD = "2020-2023希望者数の年平均増減を線形外挿した参考予想"


def set_widths(ws, widths: dict[int, float]) -> None:
    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width


def style_header(ws, row_number: int, end_column: int, fill: PatternFill = HEADER_FILL) -> None:
    for column_index in range(1, end_column + 1):
        cell = ws.cell(row=row_number, column=column_index)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER


def style_body_row(ws, row_number: int, end_column: int) -> None:
    for column_index in range(1, end_column + 1):
        cell = ws.cell(row=row_number, column=column_index)
        cell.font = BODY_FONT
        cell.alignment = LEFT


def add_table(ws, name: str) -> None:
    end_column = get_column_letter(ws.max_column)
    table = Table(displayName=name, ref=f"A1:{end_column}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def build_trend_rows() -> list[dict[str, object]]:
    applicants_by_year = {point.year: point for point in APPLICANT_POINTS}
    rows: list[dict[str, object]] = []

    for capacity in CAPACITY_POINTS:
        applicant = applicants_by_year.get(capacity.year)
        pressure_ratio = None
        applicant_source_id = ""
        applicant_scope = ""
        applicant_count = None
        survey_date = ""
        if applicant is not None:
            applicant_count = applicant.applicants
            applicant_scope = f"{applicant.scope}（{applicant.target}）"
            applicant_source_id = applicant.source_id
            survey_date = applicant.survey_date
            pressure_ratio = applicant.applicants / capacity.total_capacity

        rows.append(
            {
                "year": capacity.year,
                "wide_area_capacity": capacity.wide_area_capacity,
                "community_capacity": capacity.community_capacity,
                "total_capacity": capacity.total_capacity,
                "capacity_definition": "必要入所定員総数（計画値）",
                "applicant_survey_date": survey_date,
                "applicants": applicant_count,
                "applicant_definition": applicant_scope,
                "pressure_ratio": pressure_ratio,
                "pressure_definition": "参考算出値（希望者数 ÷ 必要入所定員総数）",
                "forecast_applicants_linear": None,
                "forecast_pressure_ratio_linear": None,
                "forecast_method": "",
                "capacity_source_id": capacity.source_id,
                "applicant_source_id": applicant_source_id,
            }
        )

    observed_points = sorted(APPLICANT_POINTS, key=lambda point: point.year)
    start_point = observed_points[0]
    end_point = observed_points[-1]
    slope = (end_point.applicants - start_point.applicants) / (end_point.year - start_point.year)

    for row in rows:
        if row["year"] <= end_point.year:
            continue
        forecast_applicants = max(
            0,
            int(round(end_point.applicants + slope * (row["year"] - end_point.year))),
        )
        row["forecast_applicants_linear"] = forecast_applicants
        row["forecast_pressure_ratio_linear"] = forecast_applicants / row["total_capacity"]
        row["forecast_method"] = FORECAST_METHOD

    return rows


def write_trend_sheet(wb: Workbook, rows: list[dict[str, object]]) -> None:
    ws = wb.active
    ws.title = "定員・希望者推移"
    headers = [
        "年度",
        "広域型定員",
        "地域密着型定員",
        "定員合計",
        "定員定義",
        "希望者調査基準日",
        "希望者数",
        "希望者定義",
        "病床ひっ迫率",
        "ひっ迫率定義",
        "参考予想希望者数",
        "参考予想ひっ迫率",
        "参考予想ロジック",
        "定員ソースID",
        "希望者ソースID",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for row in rows:
        ws.append(
            [
                row["year"],
                row["wide_area_capacity"],
                row["community_capacity"],
                row["total_capacity"],
                row["capacity_definition"],
                row["applicant_survey_date"],
                row["applicants"],
                row["applicant_definition"],
                row["pressure_ratio"],
                row["pressure_definition"],
                row["forecast_applicants_linear"],
                row["forecast_pressure_ratio_linear"],
                row["forecast_method"],
                row["capacity_source_id"],
                row["applicant_source_id"],
            ]
        )
        style_body_row(ws, ws.max_row, len(headers))

    for row in range(2, ws.max_row + 1):
        for column in (2, 3, 4, 7, 11):
            cell = ws.cell(row=row, column=column)
            if cell.value is not None:
                cell.number_format = "#,##0"
                cell.alignment = CENTER
        ratio_cell = ws.cell(row=row, column=9)
        if ratio_cell.value is not None:
            ratio_cell.number_format = "0.0%"
            ratio_cell.alignment = CENTER
        forecast_ratio_cell = ws.cell(row=row, column=12)
        if forecast_ratio_cell.value is not None:
            forecast_ratio_cell.number_format = "0.0%"
            forecast_ratio_cell.alignment = CENTER
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=6).alignment = CENTER
        ws.cell(row=row, column=14).alignment = CENTER
        ws.cell(row=row, column=15).alignment = CENTER

    ws.freeze_panes = "A2"
    set_widths(
        ws,
        {
            1: 9,
            2: 12,
            3: 14,
            4: 12,
            5: 28,
            6: 14,
            7: 11,
            8: 24,
            9: 12,
            10: 30,
            11: 14,
            12: 14,
            13: 34,
            14: 14,
            15: 14,
        },
    )
    add_table(ws, "AichiTokuyoTrend")


def write_applicant_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("希望者調査")
    headers = ["年度", "基準日", "希望者数", "対象", "定義", "ソースID"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for point in APPLICANT_POINTS:
        ws.append(
            [
                point.year,
                point.survey_date,
                point.applicants,
                point.target,
                point.scope,
                point.source_id,
            ]
        )
        style_body_row(ws, ws.max_row, len(headers))

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=3).number_format = "#,##0"
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=6).alignment = CENTER

    ws.freeze_panes = "A2"
    set_widths(ws, {1: 9, 2: 14, 3: 11, 4: 10, 5: 24, 6: 14})
    add_table(ws, "AichiTokuyoApplicants")


def write_source_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("ソース情報")
    headers = ["ソースID", "種別", "資料名", "公表日", "PDFページ", "URL", "メモ"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for source in SOURCES:
        ws.append(
            [
                source.source_id,
                source.category,
                source.title,
                source.published_date,
                source.pdf_page,
                source.url,
                source.note,
            ]
        )
        style_body_row(ws, ws.max_row, len(headers))

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=5).alignment = CENTER

    ws.freeze_panes = "A2"
    set_widths(ws, {1: 14, 2: 10, 3: 36, 4: 12, 5: 10, 6: 62, 7: 54})
    add_table(ws, "AichiTokuyoSources")


def write_note_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("注意事項")
    ws.append(["項目", "内容"])
    style_header(ws, 1, 2, NOTE_FILL)

    for index, note in enumerate(NOTES, start=1):
        ws.append([f"注意{index}", note])
        style_body_row(ws, ws.max_row, 2)

    ws.freeze_panes = "A2"
    set_widths(ws, {1: 10, 2: 110})
    add_table(ws, "AichiTokuyoNotes")


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "year",
        "wide_area_capacity",
        "community_capacity",
        "total_capacity",
        "capacity_definition",
        "applicant_survey_date",
        "applicants",
        "applicant_definition",
        "pressure_ratio",
        "pressure_definition",
        "forecast_applicants_linear",
        "forecast_pressure_ratio_linear",
        "forecast_method",
        "capacity_source_id",
        "applicant_source_id",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    write_trend_sheet(wb, rows)
    write_applicant_sheet(wb)
    write_source_sheet(wb)
    write_note_sheet(wb)
    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build Aichi Tokuyo capacity/applicant/pressure dataset."
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX_PATH,
        help="Output xlsx path.",
    )
    parser.add_argument(
        "--csv",
        type=Path,
        default=DEFAULT_CSV_PATH,
        help="Output csv path.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows = build_trend_rows()
    write_workbook(args.xlsx, rows)
    write_csv(args.csv, rows)
    available_ratios = [row["pressure_ratio"] for row in rows if row["pressure_ratio"] is not None]
    print(f"Trend rows: {len(rows)}")
    print(f"Applicant snapshots: {len(APPLICANT_POINTS)}")
    print(f"Pressure ratios available: {len(available_ratios)}")
    print(f"XLSX: {args.xlsx}")
    print(f"CSV: {args.csv}")


if __name__ == "__main__":
    main()
