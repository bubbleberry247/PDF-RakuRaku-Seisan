from __future__ import annotations

import argparse
import re
import unicodedata
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SOURCE_PAGE_URL = "https://www.pref.aichi.jp/soshiki/korei/tokuyou-itiran.html"
PDF_URL = "https://www.pref.aichi.jp/uploaded/attachment/578325.pdf"
EXPECTED_RECORD_COUNT = 437
DEFAULT_WORK_DIR = Path("work") / "aichi_tokuyo"
DEFAULT_PDF_PATH = DEFAULT_WORK_DIR / "aichi_tokuyo_official.pdf"
DEFAULT_OUTPUT_PATH = DEFAULT_WORK_DIR / "愛知県特養施設リスト.xlsx"

HEADER_FILL = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
SUMMARY_FILL = PatternFill(fill_type="solid", start_color="EAF4E2", end_color="EAF4E2")
HEADER_FONT = Font(name="Meiryo", bold=True, size=10)
BODY_FONT = Font(name="Meiryo", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

MUNICIPALITY_PATTERNS = (
    re.compile(r"^(名古屋市[^0-9０-９]+?区)"),
    re.compile(r"^([^0-9０-９]+?市)"),
    re.compile(r"^([^0-9０-９]+?郡[^0-9０-９]+?[町村])"),
    re.compile(r"^([^0-9０-９]+?町)"),
    re.compile(r"^([^0-9０-９]+?村)"),
)


@dataclass(frozen=True)
class FacilityRecord:
    no: int
    region: str
    facility_name: str
    operator: str
    capacity: int
    address: str
    municipality: str
    phone: str
    notes: str


def download_pdf(url: str, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0"},
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        output_path.write_bytes(response.read())
    return output_path


def should_join_tokens(tokens: list[str]) -> bool:
    if len(tokens) <= 1:
        return False
    short_token_ratio = sum(len(token) <= 2 for token in tokens) / len(tokens)
    return short_token_ratio >= 0.7


def normalize_cell(
    value: str | None,
    *,
    join_lines_without_space: bool = True,
    force_join_tokens: bool = False,
) -> str:
    if not value:
        return ""

    normalized = unicodedata.normalize("NFKC", value).replace("\u3000", " ")
    lines = [line.strip() for line in normalized.splitlines() if line.strip()]
    if not lines:
        return ""

    cleaned_lines: list[str] = []
    for line in lines:
        tokens = line.split()
        if not tokens:
            continue
        if force_join_tokens or should_join_tokens(tokens):
            cleaned_lines.append("".join(tokens))
        else:
            cleaned_lines.append(" ".join(tokens))

    if not cleaned_lines:
        return ""

    if join_lines_without_space:
        combined = cleaned_lines[0]
        for line in cleaned_lines[1:]:
            if combined[-1].isascii() and line[0].isascii():
                combined = f"{combined} {line}"
            else:
                combined = f"{combined}{line}"
    else:
        combined = " ".join(cleaned_lines)

    combined = re.sub(r"([(\uff08])\s+", r"\1", combined)
    combined = re.sub(r"\s+([)\uff09])", r"\1", combined)
    combined = re.sub(r"([)\uff09])\s+(?=[A-Za-z0-9\u3040-\u30ff\u3400-\u9fff])", r"\1", combined)
    combined = re.sub(r"\s+([、。,])", r"\1", combined)
    return combined.strip()


def extract_municipality(address: str) -> str:
    normalized_address = normalize_cell(
        address,
        join_lines_without_space=True,
        force_join_tokens=True,
    )
    for pattern in MUNICIPALITY_PATTERNS:
        match = pattern.match(normalized_address)
        if match:
            return match.group(1)
    raise ValueError(f"Failed to extract municipality from address: {address}")


def extract_records(pdf_path: Path) -> list[FacilityRecord]:
    document = fitz.open(pdf_path)
    records: list[FacilityRecord] = []
    current_region = ""

    try:
        for page in document:
            tables = page.find_tables()
            if not tables.tables:
                raise RuntimeError(f"No table found on page {page.number + 1}")

            rows = tables.tables[0].extract()
            for raw_row in rows[1:]:
                if not raw_row or len(raw_row) < 7:
                    continue

                capacity_text = normalize_cell(raw_row[3], force_join_tokens=True)
                address = normalize_cell(
                    raw_row[4],
                    join_lines_without_space=True,
                    force_join_tokens=True,
                )
                if not capacity_text.isdigit() or not address:
                    continue

                region_text = normalize_cell(
                    raw_row[0],
                    join_lines_without_space=True,
                    force_join_tokens=True,
                )
                if region_text:
                    current_region = region_text
                if not current_region:
                    raise RuntimeError(f"Region not found before page {page.number + 1}")

                facility_name = normalize_cell(
                    raw_row[1],
                    join_lines_without_space=True,
                )
                operator = normalize_cell(
                    raw_row[2],
                    join_lines_without_space=True,
                )
                phone = normalize_cell(
                    raw_row[5],
                    join_lines_without_space=True,
                    force_join_tokens=True,
                )
                notes = normalize_cell(
                    raw_row[6],
                    join_lines_without_space=True,
                )
                municipality = extract_municipality(address)

                records.append(
                    FacilityRecord(
                        no=len(records) + 1,
                        region=current_region,
                        facility_name=facility_name,
                        operator=operator,
                        capacity=int(capacity_text),
                        address=address,
                        municipality=municipality,
                        phone=phone,
                        notes=notes,
                    )
                )
    finally:
        document.close()

    if len(records) != EXPECTED_RECORD_COUNT:
        raise RuntimeError(
            f"Expected {EXPECTED_RECORD_COUNT} records but extracted {len(records)}"
        )

    return records


def set_column_widths(worksheet, widths: dict[int, float]) -> None:
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def apply_header_style(worksheet, row_number: int, end_column: int, fill: PatternFill) -> None:
    for column_index in range(1, end_column + 1):
        cell = worksheet.cell(row=row_number, column=column_index)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER_ALIGN


def apply_body_style(worksheet, row_number: int, end_column: int) -> None:
    for column_index in range(1, end_column + 1):
        cell = worksheet.cell(row=row_number, column=column_index)
        cell.font = BODY_FONT
        cell.alignment = LEFT_ALIGN


def write_facility_sheet(workbook: Workbook, records: list[FacilityRecord]) -> None:
    worksheet = workbook.active
    worksheet.title = "施設一覧"

    headers = [
        "No",
        "圏域",
        "施設名",
        "設置者",
        "定員",
        "所在地",
        "市区町村",
        "電話番号",
        "備考",
    ]
    worksheet.append(headers)
    apply_header_style(worksheet, 1, len(headers), HEADER_FILL)

    for record in records:
        worksheet.append(
            [
                record.no,
                record.region,
                record.facility_name,
                record.operator,
                record.capacity,
                record.address,
                record.municipality,
                record.phone,
                record.notes,
            ]
        )
        apply_body_style(worksheet, worksheet.max_row, len(headers))
        worksheet.cell(row=worksheet.max_row, column=1).alignment = CENTER_ALIGN
        worksheet.cell(row=worksheet.max_row, column=5).alignment = CENTER_ALIGN

    table = Table(displayName="FacilityList", ref=f"A1:I{worksheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.auto_filter.ref = table.ref
    worksheet.freeze_panes = "A2"
    set_column_widths(
        worksheet,
        {
            1: 7,
            2: 18,
            3: 28,
            4: 24,
            5: 8,
            6: 34,
            7: 18,
            8: 16,
            9: 34,
        },
    )


def write_summary_sheet(workbook: Workbook, records: list[FacilityRecord]) -> None:
    worksheet = workbook.create_sheet("市区町村別集計")
    headers = ["市区町村", "施設数", "定員合計", "圏域"]
    worksheet.append(headers)
    apply_header_style(worksheet, 1, len(headers), SUMMARY_FILL)

    counts: dict[str, int] = defaultdict(int)
    capacities: dict[str, int] = defaultdict(int)
    regions: dict[str, set[str]] = defaultdict(set)

    for record in records:
        counts[record.municipality] += 1
        capacities[record.municipality] += record.capacity
        regions[record.municipality].add(record.region)

    for municipality in sorted(counts):
        worksheet.append(
            [
                municipality,
                counts[municipality],
                capacities[municipality],
                "、".join(sorted(regions[municipality])),
            ]
        )
        apply_body_style(worksheet, worksheet.max_row, len(headers))
        worksheet.cell(row=worksheet.max_row, column=2).alignment = CENTER_ALIGN
        worksheet.cell(row=worksheet.max_row, column=3).alignment = CENTER_ALIGN

    table = Table(displayName="MunicipalitySummary", ref=f"A1:D{worksheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.auto_filter.ref = table.ref
    worksheet.freeze_panes = "A2"
    set_column_widths(worksheet, {1: 18, 2: 10, 3: 12, 4: 22})


def write_source_sheet(workbook: Workbook, records: list[FacilityRecord], pdf_path: Path) -> None:
    worksheet = workbook.create_sheet("ソース情報")
    rows = [
        ("生成日時UTC", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("抽出件数", str(len(records))),
        ("想定件数", str(EXPECTED_RECORD_COUNT)),
        ("PDF保存先", str(pdf_path.resolve())),
        ("PDF URL", PDF_URL),
        ("案内ページ", SOURCE_PAGE_URL),
        ("出典メモ", "愛知県庁公式PDF「（3）特別養護老人ホーム(介護老人福祉施設）」"),
    ]

    for key, value in rows:
        worksheet.append([key, value])

    for row_number in range(1, worksheet.max_row + 1):
        worksheet.cell(row=row_number, column=1).font = HEADER_FONT
        worksheet.cell(row=row_number, column=1).fill = HEADER_FILL
        worksheet.cell(row=row_number, column=1).alignment = LEFT_ALIGN
        worksheet.cell(row=row_number, column=2).font = BODY_FONT
        worksheet.cell(row=row_number, column=2).alignment = LEFT_ALIGN

    set_column_widths(worksheet, {1: 16, 2: 120})


def write_workbook(records: list[FacilityRecord], output_path: Path, pdf_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    write_facility_sheet(workbook, records)
    write_summary_sheet(workbook, records)
    write_source_sheet(workbook, records, pdf_path)
    workbook.save(output_path)
    return output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extract Aichi special nursing home facilities from the official PDF."
    )
    parser.add_argument(
        "--pdf-url",
        default=PDF_URL,
        help="Official PDF URL.",
    )
    parser.add_argument(
        "--pdf-path",
        default=str(DEFAULT_PDF_PATH),
        help="Local path where the PDF will be stored.",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT_PATH),
        help="Output Excel path.",
    )
    parser.add_argument(
        "--skip-download",
        action="store_true",
        help="Use an existing local PDF instead of downloading again.",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    pdf_path = Path(args.pdf_path)
    output_path = Path(args.output)

    if not args.skip_download or not pdf_path.exists():
        download_pdf(args.pdf_url, pdf_path)

    records = extract_records(pdf_path)
    workbook_path = write_workbook(records, output_path, pdf_path)

    municipality_count = len({record.municipality for record in records})
    print(f"PDF: {pdf_path.resolve()}")
    print(f"Records extracted: {len(records)}")
    print(f"Municipalities extracted: {municipality_count}")
    print(f"Workbook: {workbook_path.resolve()}")


if __name__ == "__main__":
    main()
