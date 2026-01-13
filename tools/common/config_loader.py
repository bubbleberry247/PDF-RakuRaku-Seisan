# -*- coding: utf-8 -*-
"""
設定ファイル読込モジュール

Usage:
    from common.config_loader import load_config, get_env

    env = get_env()  # LOCAL or PROD
    config = load_config(env)
    print(config['FAX_FOLDER'])
"""
import os
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

# ベースディレクトリ
BASE_DIR = Path(r"C:\ProgramData\RK10\Robots\44PDF一般経費楽楽精算申請")
CONFIG_DIR = BASE_DIR / "config"


def get_env() -> str:
    """環境識別子を取得（LOCAL or PROD）

    Returns:
        str: 'LOCAL' または 'PROD'
    """
    env_file = CONFIG_DIR / "env.txt"
    if env_file.exists():
        env = env_file.read_text(encoding='utf-8').strip().upper()
        if env in ('LOCAL', 'PROD'):
            return env
    return 'LOCAL'


def load_config(env: Optional[str] = None) -> dict:
    """設定ファイルを読み込む

    Args:
        env: 環境識別子（省略時はenv.txtから読み込み）

    Returns:
        dict: 設定値の辞書
    """
    if env is None:
        env = get_env()

    config_file = CONFIG_DIR / f"RK10_config_{env}.xlsx"
    if not config_file.exists():
        raise FileNotFoundError(f"設定ファイルが見つかりません: {config_file}")

    wb = load_workbook(config_file, data_only=True)
    ws = wb.active

    config = {}
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        if key:
            config[key] = value if value is not None else ""

    wb.close()

    # 環境識別子を追加
    config['ENV'] = env

    return config


if __name__ == "__main__":
    # テスト実行
    import sys
    sys.stdout.reconfigure(encoding='utf-8')

    print(f"環境: {get_env()}")
    print()

    config = load_config()
    print("=== 設定値 ===")
    for key, value in config.items():
        print(f"  {key}: {value}")
