# -*- coding: utf-8 -*-
"""scenario70 configuration loader."""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


DEFAULT_CONFIG_DIR = Path(r"C:\ProgramData\RK10\Robots\70楽楽精算支払い確定\config")


def get_env(config_dir: Path | str = DEFAULT_CONFIG_DIR, env_override: Optional[str] = None) -> str:
    if env_override:
        return env_override.strip().upper()
    config_path = Path(config_dir)
    env_file = config_path / "env.txt"
    if env_file.exists():
        value = env_file.read_text(encoding="utf-8").strip().upper()
        if value.startswith("ENV="):
            value = value.split("=", 1)[1].strip()
        if value in {"LOCAL", "PROD"}:
            return value
    return "LOCAL"


def load_config(config_dir: Path | str = DEFAULT_CONFIG_DIR, env: Optional[str] = None) -> dict[str, str]:
    config_path = Path(config_dir)
    env_name = get_env(config_path, env)
    workbook_path = config_path / f"RK10_config_{env_name}.xlsx"
    if not workbook_path.exists():
        raise FileNotFoundError(f"config file not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb.active
    config: dict[str, str] = {}
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        if key:
            config[str(key).strip()] = "" if value is None else str(value).strip()
    wb.close()
    config["ENV"] = env_name
    config["CONFIG_DIR"] = str(config_path)
    return config


def split_mail_list(value: str) -> list[str]:
    if not value:
        return []
    normalized = value.replace(";", ",")
    return [item.strip() for item in normalized.split(",") if item.strip()]
