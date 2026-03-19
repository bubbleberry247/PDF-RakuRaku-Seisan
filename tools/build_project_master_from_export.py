# -*- coding: utf-8 -*-
"""Build normalized project_master workbook from a Karuwaza/Keigyo export.

Supported inputs:
- .xlsx
- .csv
- .tsv

Output columns:
- 工事番号
- 工事名称
- キーワード
- 部署
- status
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Iterable
from unicodedata import normalize

import openpyxl


DEFAULT_NUMBER_COLUMNS = ("工事番号", "工事No", "案件番号", "番号", "物件番号")
DEFAULT_NAME_COLUMNS = ("工事名称", "現場名", "案件名", "物件名", "工事名")
DEFAULT_BUSHO_COLUMNS = ("部署", "部署コード", "工事区分", "区分")
DEFAULT_KEYWORD_COLUMNS = (
    "キーワード",
    "検索用",
    "略称",
    "現場略称",
    "工事名称",
    "現場名",
    "顧客",
    "顧客略称",
    "現場住所1",
    "現場住所2",
    "工種名称",
)
REPAIR_NAME_TOKENS = ("修繕", "改修", "修理", "取替", "交換", "更新", "補修", "是正", "改造", "撤去", "営繕")
NEW_NAME_TOKENS = ("新築", "計画", "建築", "新設", "増築", "造成")


def _normalize_header(value: object) -> str:
    return normalize("NFKC", str(value or "")).strip().lower()


def _stringify(value: object) -> str:
    if value is None:
        return ""
    return normalize("NFKC", str(value)).strip()


def _split_list(value: str) -> list[str]:
    parts = []
    for raw in (value or "").split(","):
        item = _stringify(raw)
        if item:
            parts.append(item)
    return parts


def _load_mapping_json(path: Path) -> dict[str, object]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(data, dict):
        raise ValueError(f"mapping json must be an object: {path}")
    return data


def _iter_source_rows(path: Path, sheet_name: str | None) -> tuple[list[str], list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            headers = [_stringify(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows: list[dict[str, str]] = []
            for values in ws.iter_rows(min_row=2, values_only=True):
                row = {headers[idx]: _stringify(value) for idx, value in enumerate(values) if idx < len(headers)}
                rows.append(row)
            return headers, rows
        finally:
            wb.close()

    if suffix not in {".csv", ".tsv"}:
        raise ValueError(f"unsupported input format: {path}")

    delimiter = "\t" if suffix == ".tsv" else ","
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        headers = list(reader.fieldnames or [])
        rows = [{header: _stringify(value) for header, value in row.items()} for row in reader]
    return headers, rows


def _resolve_column(
    headers: Iterable[str],
    *,
    explicit: str | None,
    candidates: Iterable[str],
    required: bool,
) -> str | None:
    header_map = {_normalize_header(header): header for header in headers}
    if explicit:
        explicit_norm = _normalize_header(explicit)
        if explicit_norm in header_map:
            return header_map[explicit_norm]
        raise ValueError(f"column not found: {explicit}")
    for candidate in candidates:
        candidate_norm = _normalize_header(candidate)
        if candidate_norm in header_map:
            return header_map[candidate_norm]
    if required:
        raise ValueError(f"required column not found. tried: {', '.join(candidates)}")
    return None


def _auto_keyword_columns(headers: Iterable[str], *, excluded: set[str]) -> list[str]:
    header_map = {_normalize_header(header): header for header in headers}
    found: list[str] = []
    for candidate in DEFAULT_KEYWORD_COLUMNS:
        candidate_norm = _normalize_header(candidate)
        resolved = header_map.get(candidate_norm)
        if not resolved or resolved in excluded:
            continue
        if resolved not in found:
            found.append(resolved)
    return found


def _normalize_busho(
    raw: str,
    *,
    repair_values: set[str],
    new_values: set[str],
) -> tuple[str, str, str]:
    value = _stringify(raw)
    if not value:
        return "", "inactive", "missing"
    if value in repair_values or any(token in value for token in ("修", "営")):
        return "修繕", "active", "explicit"
    if value in new_values or "新" in value:
        return "新築", "active", "explicit"
    return value, "inactive", "explicit_unknown"


def _infer_busho_from_name(name: str) -> tuple[str, str, str]:
    project_name = _stringify(name)
    if not project_name:
        return "", "inactive", "name_missing"
    if any(token in project_name for token in REPAIR_NAME_TOKENS):
        return "修繕", "active", "name_inference"
    if any(token in project_name for token in NEW_NAME_TOKENS):
        return "新築", "active", "name_inference"
    return "", "inactive", "name_ambiguous"


def _merge_keywords(*values: str) -> str:
    seen: OrderedDict[str, None] = OrderedDict()
    for value in values:
        normalized = re.sub(r"(?<=\d),(?=\d)", "", value.replace("、", ",").replace(";", ","))
        for part in _split_list(normalized):
            normalized_part = normalize("NFKC", part).strip()
            if len(normalized_part) < 3:
                continue
            if normalized_part.isdigit() and len(normalized_part) < 4:
                continue
            if part not in seen:
                seen[part] = None
    return ",".join(seen.keys())


def _write_output(path: Path, rows: list[tuple[str, str, str, str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "project_master"
    ws.append(["工事番号", "工事名称", "キーワード", "部署", "status"])
    for row in rows:
        ws.append(list(row))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _load_master_rows(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: list[dict[str, str]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append(
                {
                    "工事番号": _stringify(r[0]),
                    "工事名称": _stringify(r[1]),
                    "キーワード": _stringify(r[2]),
                    "部署": _stringify(r[3]),
                    "status": _stringify(r[4]) or "active",
                }
            )
        return rows
    finally:
        wb.close()


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Source export file (.xlsx/.csv/.tsv)")
    ap.add_argument("--output", required=False, help="Normalized project master output path")
    ap.add_argument("--sheet", default=None, help="Sheet name for xlsx input")
    ap.add_argument("--mapping-json", default=None, help="JSON file with column/value mapping")
    ap.add_argument("--list-columns", action="store_true", help="Print source headers and exit")
    ap.add_argument("--base-master", default=None, help="Existing normalized project_master xlsx to inherit by 工事番号")
    ap.add_argument(
        "--infer-busho-from-name",
        action="store_true",
        help="If no explicit usable busho is available, infer only obvious 新築/修繕 cases from project name",
    )
    ap.add_argument("--number-column", default=None, help="Source column name for project number")
    ap.add_argument("--name-column", default=None, help="Source column name for project name")
    ap.add_argument("--busho-column", default=None, help="Source column name for department/classification")
    ap.add_argument(
        "--keyword-columns",
        default="",
        help="Comma-separated extra source columns to merge into キーワード",
    )
    ap.add_argument(
        "--repair-values",
        default="修繕,営繕",
        help="Comma-separated source values mapped to 修繕",
    )
    ap.add_argument(
        "--new-values",
        default="新築",
        help="Comma-separated source values mapped to 新築",
    )
    args = ap.parse_args(argv)

    src = Path(args.input)
    if not src.exists():
        raise FileNotFoundError(src)

    output = Path(args.output) if args.output else (
        src.parent / f"project_master_v{datetime.now():%Y%m%d}.xlsx"
    )
    headers, source_rows = _iter_source_rows(src, args.sheet)
    if args.list_columns:
        for idx, header in enumerate(headers, start=1):
            print(f"{idx}\t{header}")
        return 0

    mapping = _load_mapping_json(Path(args.mapping_json)) if args.mapping_json else {}
    number_column = args.number_column or str(mapping.get("number_column") or "") or None
    name_column = args.name_column or str(mapping.get("name_column") or "") or None
    busho_column = args.busho_column or str(mapping.get("busho_column") or "") or None
    keyword_columns_raw = args.keyword_columns
    if not keyword_columns_raw and isinstance(mapping.get("keyword_columns"), list):
        keyword_columns_raw = ",".join([_stringify(v) for v in mapping["keyword_columns"]])
    repair_values_raw = args.repair_values
    if repair_values_raw == "修繕,営繕" and isinstance(mapping.get("repair_values"), list):
        repair_values_raw = ",".join([_stringify(v) for v in mapping["repair_values"]])
    new_values_raw = args.new_values
    if new_values_raw == "新築" and isinstance(mapping.get("new_values"), list):
        new_values_raw = ",".join([_stringify(v) for v in mapping["new_values"]])

    number_col = _resolve_column(
        headers,
        explicit=number_column,
        candidates=DEFAULT_NUMBER_COLUMNS,
        required=False,
    )
    name_col = _resolve_column(
        headers,
        explicit=name_column,
        candidates=DEFAULT_NAME_COLUMNS,
        required=True,
    )
    busho_col = _resolve_column(
        headers,
        explicit=busho_column,
        candidates=DEFAULT_BUSHO_COLUMNS,
        required=False,
    )
    keyword_cols: list[str] = []
    for raw in _split_list(keyword_columns_raw):
        resolved = _resolve_column(headers, explicit=raw, candidates=(), required=False)
        if resolved:
            keyword_cols.append(resolved)
    if not keyword_cols:
        keyword_cols = _auto_keyword_columns(
            headers,
            excluded={col for col in (number_col, name_col, busho_col) if col},
        )

    repair_values = set(_split_list(repair_values_raw))
    new_values = set(_split_list(new_values_raw))

    merged: OrderedDict[tuple[str, str], dict[str, str]] = OrderedDict()
    unknown_busho = 0
    inferred_busho = 0
    explicit_busho = 0
    for row in source_rows:
        name = _stringify(row.get(name_col, ""))
        if not name:
            continue
        number = _stringify(row.get(number_col, "")) if number_col else ""
        busho_raw = _stringify(row.get(busho_col, "")) if busho_col else ""
        busho, status, source = _normalize_busho(
            busho_raw,
            repair_values=repair_values,
            new_values=new_values,
        )
        if status != "active" and args.infer_busho_from_name:
            busho, status, source = _infer_busho_from_name(name)
        if status != "active":
            unknown_busho += 1
        elif source == "name_inference":
            inferred_busho += 1
        else:
            explicit_busho += 1
        keyword_parts = [name]
        for col in keyword_cols:
            keyword_parts.append(_stringify(row.get(col, "")))
        keywords = _merge_keywords(*keyword_parts)
        key = (number, name)
        current = merged.get(key)
        if current is None:
            merged[key] = {
                "工事番号": number,
                "工事名称": name,
                "キーワード": keywords,
                "部署": busho,
                "status": status,
            }
            continue
        current["キーワード"] = _merge_keywords(current["キーワード"], keywords)
        if current["status"] != "active" and status == "active":
            current["部署"] = busho
            current["status"] = status

    inherited_count = 0
    carried_from_base_only = 0
    if args.base_master:
        base_rows = _load_master_rows(Path(args.base_master))
        base_by_number = {
            row["工事番号"]: row
            for row in base_rows
            if row["工事番号"]
        }
        for key, item in list(merged.items()):
            number = item["工事番号"]
            base = base_by_number.get(number)
            if not base:
                continue
            inherited_count += 1
            item["工事名称"] = base["工事名称"] or item["工事名称"]
            item["キーワード"] = _merge_keywords(
                base["工事名称"],
                base["キーワード"],
                item["工事名称"],
                item["キーワード"],
            )
            if base["status"] == "active":
                item["部署"] = base["部署"]
                item["status"] = base["status"]
        existing_numbers = {item["工事番号"] for item in merged.values() if item["工事番号"]}
        for row in base_rows:
            number = row["工事番号"]
            if number and number in existing_numbers:
                continue
            carried_from_base_only += 1
            merged[(row["工事番号"], row["工事名称"])] = dict(row)

    rows = []
    for item in merged.values():
        rows.append(
            (
                item["工事番号"],
                item["工事名称"],
                item["キーワード"],
                item["部署"],
                item["status"],
            )
        )
    _write_output(output, rows)
    print(
        f"wrote {len(rows)} rows to {output} "
        f"(unknown_busho={unknown_busho}, "
        f"explicit_busho={explicit_busho}, inferred_busho={inferred_busho}, "
        f"inherited_from_base={inherited_count}, carried_base_only={carried_from_base_only}, "
        f"number_column={number_col or '-'}, name_column={name_col}, "
        f"busho_column={busho_col or '-'}, keyword_columns={','.join(keyword_cols) or '-'})"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
