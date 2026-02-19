"""
Scenario 56: Inspect mock Excel files to understand data structure.
"""
import openpyxl
from pathlib import Path
from datetime import datetime

def inspect_input_file():
    """入金一覧 (Input file)"""
    path = Path(r"C:\RK10_MOCK\ユニオン共有\5F\15全社資金繰表\建設事業部\全社入金予定表.xlsx")
    print("=" * 80)
    print("INPUT FILE: 全社入金予定表.xlsx")
    print("=" * 80)

    if not path.exists():
        print(f"ERROR: File not found: {path}")
        return

    wb = openpyxl.load_workbook(path, data_only=True)

    # 1. List all sheet names
    print(f"\n[Sheet Names] ({len(wb.sheetnames)} sheets)")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {name}")

    # 2. Examine sheet "全社入金予定表"
    if "全社入金予定表" in wb.sheetnames:
        ws = wb["全社入金予定表"]
        print(f"\n[Sheet: 全社入金予定表]")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

        # Show rows 1-6
        print("\n[Rows 1-6: Column Layout]")
        for row_idx in range(1, 7):
            row_data = []
            for col_idx in range(1, 11):  # A-J columns
                cell = ws.cell(row_idx, col_idx)
                value = cell.value
                if value is None:
                    value = "(empty)"
                elif isinstance(value, datetime):
                    value = f"{value.strftime('%Y-%m-%d')}"
                else:
                    value = str(value)[:20]  # Truncate long values
                row_data.append(f"{openpyxl.utils.get_column_letter(col_idx)}:{value}")
            print(f"  Row {row_idx}: {' | '.join(row_data)}")

        # Sample G, H, I columns
        print("\n[Sample Data: G(date), H(金額), I(未収金額)]")
        sample_rows = [7, 8, 9, 10, 20, 30, 40]
        for row_idx in sample_rows:
            if row_idx > ws.max_row:
                break
            g_val = ws.cell(row_idx, 7).value  # G
            h_val = ws.cell(row_idx, 8).value  # H
            i_val = ws.cell(row_idx, 9).value  # I

            g_str = f"{g_val}" if g_val else "(empty)"
            h_str = f"{h_val}" if h_val else "(empty)"
            i_str = f"{i_val}" if i_val else "(empty)"

            print(f"  Row {row_idx}: G={g_str}, H={h_str}, I={i_str}")

        # Count data rows
        data_count = 0
        for row_idx in range(1, ws.max_row + 1):
            if ws.cell(row_idx, 7).value is not None:  # G column has data
                data_count += 1
        print(f"\n[Total rows with G column data]: {data_count}")
    else:
        print("\nWARNING: Sheet '全社入金予定表' not found")

    wb.close()

def inspect_output_file():
    """資金繰表 (Output file)"""
    path = Path(r"C:\RK10_MOCK\ユニオン共有\5F\15全社資金繰表\建設事業部\第26期TIC資金繰表.xlsx")
    print("\n" + "=" * 80)
    print("OUTPUT FILE: 第26期TIC資金繰表.xlsx")
    print("=" * 80)

    if not path.exists():
        print(f"ERROR: File not found: {path}")
        return

    wb = openpyxl.load_workbook(path, data_only=True)

    # 1. List all sheet names
    print(f"\n[Sheet Names] ({len(wb.sheetnames)} sheets)")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {name}")

    # 2. Pick one sheet that looks like "7.11" pattern
    target_sheet = None
    for name in wb.sheetnames:
        if "." in name and any(c.isdigit() for c in name):
            target_sheet = name
            break

    if not target_sheet:
        print("\nWARNING: No sheet matching date pattern found")
        wb.close()
        return

    ws = wb[target_sheet]
    print(f"\n[Examining Sheet: {target_sheet}]")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

    # Show first 20 rows: A (date), F (label), G (value)
    print("\n[First 20 rows: A(date), F(label), G(value)]")
    for row_idx in range(1, 21):
        a_val = ws.cell(row_idx, 1).value  # A
        f_val = ws.cell(row_idx, 6).value  # F
        g_val = ws.cell(row_idx, 7).value  # G

        # Check A column type
        a_cell = ws.cell(row_idx, 1)
        a_type = type(a_val).__name__
        if isinstance(a_val, datetime):
            a_str = f"{a_val.strftime('%Y-%m-%d')} ({a_type})"
        elif isinstance(a_val, (int, float)):
            a_str = f"{a_val} ({a_type}, serial?)"
        else:
            a_str = f"{a_val} ({a_type})"

        f_str = f"{f_val}" if f_val else "(empty)"
        g_str = f"{g_val}" if g_val else "(empty)"

        # Highlight "工事金" rows
        marker = " *** 工事金 ***" if f_val and "工事金" in str(f_val) else ""
        print(f"  Row {row_idx}: A={a_str}, F={f_str}, G={g_str}{marker}")

    # Search for more "工事金" rows
    print("\n[Search for '工事金' rows in entire sheet]")
    kojikin_rows = []
    for row_idx in range(1, ws.max_row + 1):
        f_val = ws.cell(row_idx, 6).value
        if f_val and "工事金" in str(f_val):
            kojikin_rows.append(row_idx)

    print(f"Found {len(kojikin_rows)} rows with '工事金' in F column")
    if kojikin_rows:
        print(f"Row numbers: {kojikin_rows[:10]}")  # Show first 10

    wb.close()

if __name__ == "__main__":
    try:
        inspect_input_file()
        inspect_output_file()
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
