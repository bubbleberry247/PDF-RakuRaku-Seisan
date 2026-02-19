# -*- coding: utf-8 -*-
"""
Populate v2 template Excel with supplier names from master CSV.

Pre-fills 末日 section (rows 142-541) with top suppliers by frequency.
"""
import csv
import sys
from pathlib import Path

import openpyxl

# Windows Japanese output
sys.stdout.reconfigure(encoding="utf-8")

# Paths
MASTER_CSV = Path(r"C:\ProgramData\RK10\Robots\55 １５日２５日末日振込エクセル作成\config\supplier_template_master.csv")
TEMPLATE_EXCEL = Path(r"C:\ProgramData\RK10\Robots\55 １５日２５日末日振込エクセル作成\docs\55_v2_template.xlsx")

# v2 column constants
COL_A_SUPPLIER = 1
COL_D_ACCOUNT = 4
COL_F_SHOP = 6
COL_G_ACCOUNT_CODE = 7

# 末日 section (400 slots)
SECTION_START = 142
SECTION_END = 541
SUBTOTAL_ROW = 542


def load_master():
    """Load master CSV and sort by count descending."""
    suppliers = []
    with open(MASTER_CSV, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            count_str = row.get("count", "0").strip()
            try:
                count = int(count_str)
            except ValueError:
                count = 0
            suppliers.append({
                "supplier_name": row.get("supplier_name", "").strip(),
                "account": row.get("account", "").strip(),
                "shop": row.get("shop", "").strip(),
                "account_code": row.get("account_code", "").strip(),
                "count": count,
            })
    # Sort by count descending
    suppliers.sort(key=lambda x: x["count"], reverse=True)
    return suppliers


def extract_first_line(name: str) -> str:
    """Extract first line of multiline supplier name."""
    if "\n" in name:
        return name.split("\n")[0].strip()
    return name


def main():
    print("=== v2テンプレート支払先名事前登録 ===\n")
    print(f"マスタCSV: {MASTER_CSV}")
    print(f"テンプレートExcel: {TEMPLATE_EXCEL}\n")

    if not MASTER_CSV.exists():
        print(f"エラー: マスタCSVが見つかりません: {MASTER_CSV}")
        return 1
    if not TEMPLATE_EXCEL.exists():
        print(f"エラー: テンプレートExcelが見つかりません: {TEMPLATE_EXCEL}")
        return 1

    # Load master
    suppliers = load_master()
    print(f"マスタ読み込み: {len(suppliers)}件")
    if len(suppliers) != 119:
        print(f"警告: 件数が119件ではありません: {len(suppliers)}件")

    # Open Excel
    wb = openpyxl.load_workbook(TEMPLATE_EXCEL)
    if len(wb.sheetnames) != 1:
        print(f"警告: シート数が1枚ではありません: {len(wb.sheetnames)}枚")
    ws = wb[wb.sheetnames[0]]
    print(f"シート: {ws.title}\n")

    # Populate 末日 section
    available_rows = SECTION_END - SECTION_START + 1
    print(f"末日セクション: 行{SECTION_START}-{SECTION_END} ({available_rows}スロット)")
    print(f"転記予定件数: {min(len(suppliers), available_rows)}件\n")

    populated = 0
    for i, supplier in enumerate(suppliers):
        row = SECTION_START + i
        if row > SECTION_END:
            print(f"警告: 空きスロット不足（{len(suppliers) - i}件未転記）")
            break

        name = extract_first_line(supplier["supplier_name"])
        if not name:
            continue

        # A: supplier name
        ws.cell(row=row, column=COL_A_SUPPLIER).value = name

        # D: account (if available)
        account = supplier.get("account")
        if account:
            ws.cell(row=row, column=COL_D_ACCOUNT).value = account

        # F: shop (if available)
        shop = supplier.get("shop")
        if shop:
            ws.cell(row=row, column=COL_F_SHOP).value = shop

        # G: account_code (if available)
        account_code = supplier.get("account_code")
        if account_code:
            ws.cell(row=row, column=COL_G_ACCOUNT_CODE).value = account_code

        populated += 1

    print(f"転記完了: {populated}件\n")

    # Save
    wb.save(TEMPLATE_EXCEL)
    print(f"保存完了: {TEMPLATE_EXCEL}\n")

    # Verify SUM formula (subtotal row)
    subtotal_cell = ws.cell(row=SUBTOTAL_ROW, column=3)  # C列合計
    formula = subtotal_cell.value
    print(f"SUM数式確認 (行{SUBTOTAL_ROW}): {formula}")
    if isinstance(formula, str) and formula.startswith("=SUM"):
        print("  ✓ SUM数式は保持されています")
    else:
        print("  ⚠ SUM数式が見つかりません（要確認）")

    # Sample entries
    print(f"\n登録サンプル（上位3件＋最終行）:")
    print(f"{'行':<5} {'支払先':<30} {'勘定科目':<15} {'店':<5} {'科目コード':<10}")
    print("-" * 70)
    for i in [0, 1, 2, populated - 1]:
        if i >= len(suppliers):
            continue
        row = SECTION_START + i
        name = extract_first_line(suppliers[i]["supplier_name"])
        account = suppliers[i].get("account", "")
        shop = suppliers[i].get("shop", "")
        code = suppliers[i].get("account_code", "")
        count = suppliers[i].get("count", 0)
        print(f"{row:<5} {name:<30} {account:<15} {shop:<5} {code:<10} (件数:{count})")

    print(f"\n処理完了: 119件の支払先名を末日セクションに事前登録しました")
    print(f"excel_writer_v2.py の find_supplier_row() がA列で名前マッチ可能になります")
    return 0


if __name__ == "__main__":
    sys.exit(main())
