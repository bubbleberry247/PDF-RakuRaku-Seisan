"""
Inspect December Excel file structure row-by-row
"""
import openpyxl
from pathlib import Path

def inspect_file(file_path: str, sheet_name: str, label: str):
    print(f"\n{'='*80}")
    print(f"{label}: {file_path}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*80}\n")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return

    ws = wb[sheet_name]

    # 1. Print rows 1-20, columns A-K
    print("ROWS 1-20 STRUCTURE:")
    print("-" * 80)
    for row_num in range(1, 21):
        cells = []
        for col_num in range(1, 12):  # A-K = 1-11
            cell = ws.cell(row_num, col_num)
            value = cell.value
            if value is None:
                value_str = "None"
            elif isinstance(value, str):
                value_str = repr(value)  # Show with quotes and escape chars
            else:
                value_str = str(value)

            col_letter = openpyxl.utils.get_column_letter(col_num)
            cells.append(f"{col_letter}={value_str}")

        print(f"Row {row_num:2d}: {' | '.join(cells)}")

    # 2. Find HEADER row
    print("\n" + "-" * 80)
    header_keywords = ["業者番号", "支払先", "摘要"]
    header_row = None
    for row_num in range(1, 30):
        for col_num in range(1, 12):
            value = ws.cell(row_num, col_num).value
            if value and any(kw in str(value) for kw in header_keywords):
                header_row = row_num
                print(f"HEADER ROW = {header_row}")
                break
        if header_row:
            break

    if not header_row:
        print("HEADER ROW = NOT FOUND")

    # 3. Find FIRST DATA row
    print("-" * 80)
    first_data_row = None
    if header_row:
        for row_num in range(header_row + 1, min(header_row + 20, ws.max_row + 1)):
            b_value = ws.cell(row_num, 2).value  # B column
            if b_value and str(b_value).strip():
                first_data_row = row_num
                print(f"FIRST DATA ROW = {first_data_row}")
                print(f"  B{first_data_row} = {repr(b_value)}")
                break

    if not first_data_row:
        print("FIRST DATA ROW = NOT FOUND")

    # 4. Check autofilter
    print("-" * 80)
    if ws.auto_filter and ws.auto_filter.ref:
        print(f"AUTOFILTER = {ws.auto_filter.ref}")
    else:
        print("NO AUTOFILTER")

    # 5. Check row heights
    print("-" * 80)
    print("ROW HEIGHTS (1-20):")
    for row_num in range(1, 21):
        height = ws.row_dimensions[row_num].height
        print(f"  Row {row_num:2d} height = {height}")

    # 6. Find 25日/末日 subtotal rows
    print("-" * 80)
    print("SEARCHING FOR SUBTOTAL ROWS:")
    subtotal_keywords = ["25日支払小計", "末日支払小計"]
    found_subtotals = []

    for row_num in range(1, min(ws.max_row + 1, 200)):
        for col_num in range(1, 12):  # A-K
            value = ws.cell(row_num, col_num).value
            if value:
                value_str = str(value)
                for kw in subtotal_keywords:
                    if kw in value_str:
                        col_letter = openpyxl.utils.get_column_letter(col_num)
                        found_subtotals.append((row_num, col_letter, kw, value_str))
                        print(f"  Found '{kw}' at Row {row_num}, Column {col_letter}: {repr(value_str)}")

    if not found_subtotals:
        print("  NO SUBTOTAL ROWS FOUND")

    wb.close()


if __name__ == "__main__":
    # Original December file
    original_file = r"C:\Users\masam\Desktop\KeyenceRK\ロボット作成資料\55 総合振込計上・支払伝票 12月分有.xlsx"

    # Comparison file
    comparison_file = r"C:\ProgramData\RK10\Robots\55 １５日２５日末日振込エクセル作成\docs\55 総合振込計上・支払伝票_12月比較用.xlsx"

    # Inspect original
    inspect_file(original_file, "R7.12月分", "ORIGINAL DECEMBER FILE")

    # Inspect comparison - R7.12月分
    inspect_file(comparison_file, "R7.12月分", "COMPARISON FILE - Original Sheet")

    # Inspect comparison - R7.12月分1
    inspect_file(comparison_file, "R7.12月分1", "COMPARISON FILE - RPA Modified Sheet")
