"""
Investigate the 6 amount mismatches - v2 with better matching.
"""
import openpyxl
from pathlib import Path

excel_path = Path(r"C:\ProgramData\RK10\Robots\55 １５日２５日末日振込エクセル作成\docs\55 総合振込計上・支払伝票.xlsx")

# Mismatches to investigate
mismatches = [
    {"month": "R7.6", "sheet": "R7.6月分", "supplier": "スマイルテクノロジー", "group": "25日", "orig": 29517400, "rpa": 242000},
    {"month": "R7.7", "sheet": "R7.7月分", "supplier": "ワークマン", "group": "末日", "orig": 67734, "rpa": 135468},
    {"month": "R7.8", "sheet": "R7.8月分", "supplier": "ライフテック", "group": "末日", "orig": 99660, "rpa": 64680},
    {"month": "R7.8", "sheet": "R7.8月分", "supplier": "山庄", "group": "末日", "orig": 2346189, "rpa": 2365219},
    {"month": "R7.9", "sheet": "R7.9月分", "supplier": "山庄", "group": "末日", "orig": 103910, "rpa": 84880},
    {"month": "R7.10", "sheet": "R7.10月分", "supplier": "ワークマン", "group": "末日", "orig": 5000, "rpa": 47200},
]

wb = openpyxl.load_workbook(excel_path, data_only=True)

for idx, mismatch in enumerate(mismatches, 1):
    print(f"\n{'='*80}")
    print(f"Mismatch {idx}: {mismatch['supplier']} ({mismatch['month']} {mismatch['group']})")
    print(f"{'='*80}")

    sheet_name = mismatch['sheet']
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet {sheet_name} not found!")
        continue

    ws = wb[sheet_name]
    supplier = mismatch['supplier']

    # Search for supplier in column B (odd rows only = data rows)
    found_rows = []
    for row_idx in range(1, ws.max_row + 1):
        cell_b = ws.cell(row_idx, 2)  # Column B
        if cell_b.value and supplier in str(cell_b.value):
            cell_i = ws.cell(row_idx, 9)  # Column I (amount)
            cell_a = ws.cell(row_idx, 1)  # Column A (supplier code)
            cell_c = ws.cell(row_idx, 3)  # Column C (memo)
            found_rows.append({
                "row": row_idx,
                "supplier": cell_b.value,
                "amount": cell_i.value,
                "code": cell_a.value,
                "memo": cell_c.value,
            })

    if not found_rows:
        print(f"NOT FOUND in original Excel!")
    else:
        print(f"Found {len(found_rows)} row(s) in original Excel:")
        total = 0
        for r in found_rows:
            amt_str = f"{r['amount']:,}" if isinstance(r['amount'], (int, float)) else str(r['amount'])
            memo_str = str(r['memo'])[:50] if r['memo'] else ""
            print(f"  Row {r['row']:3d}: Code={r['code']}, Amount={amt_str:>15}, Memo={memo_str}")
            if isinstance(r['amount'], (int, float)):
                total += r['amount']
        print(f"  Total amount: {total:,}")

    print(f"\nExpected original: {mismatch['orig']:,}")
    print(f"RPA wrote: {mismatch['rpa']:,}")
    print(f"Difference: {mismatch['rpa'] - mismatch['orig']:,}")

wb.close()
print("\n" + "="*80)
print("DONE")
